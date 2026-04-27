"""Read-only spike: parse prefecture-level aggregation PDFs.

Goal: estimate yield of "Layer 0 prefecture aggregator" approach without
touching the DB. For each prefecture PDF, extract (school_name, address,
operator, disclosure_url), match against existing `school` table, and
emit JSON metrics.

Output: output/pref-aggregator/{tokyo,kanagawa,saitama}.json

Usage:
    uv run python scripts/spike_pref_aggregator.py
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path

import pdfplumber

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from eidp.db.models import School, SchoolSite  # noqa: E402
from eidp.db.session import SessionLocal  # noqa: E402


# --- Domain types ---------------------------------------------------------

@dataclass
class PrefSchool:
    pref: str
    school_name_raw: str
    school_name_norm: str
    address: str
    operator_kind: str
    operator_name: str
    operator_address: str
    disclosure_url: str | None
    school_code: str | None = None  # MEXT 学校番号 (Osaka has it; gold-standard match)


@dataclass
class MatchResult:
    pref_school: PrefSchool
    db_school_id: int | None
    match_strategy: str  # exact | nfkc | operator_pref | none
    db_school_name: str | None
    has_existing_url: bool
    is_new_url_candidate: bool


@dataclass
class PrefReport:
    pref: str
    pdf_path: str
    extracted_total: int = 0
    extracted_with_url: int = 0
    db_matched: int = 0
    db_unmatched: int = 0
    match_by_strategy: dict[str, int] = field(default_factory=dict)
    existing_url_coverage: int = 0  # matched school already has URL in school_site
    new_url_candidates: int = 0  # matched school + URL in PDF + no existing URL
    url_quality_distribution: dict[str, int] = field(default_factory=dict)
    sample_unmatched: list[dict] = field(default_factory=list)
    sample_new_urls: list[dict] = field(default_factory=list)
    records: list[dict] = field(default_factory=list)  # per-school audit trail


# URL quality classifier — single source of truth, addresses Codex audit gap
# Order matters: more specific patterns checked first
DISCLOSURE_KEYWORDS = (
    "disclos", "joho", "kyufu", "shien", "youken", "yoken",
    "info", "support", "schoolguide", "valuation",
    "kakunin", "shinsei", "musho",
)

def classify_url_quality(url: str | None) -> str:
    """Classify a candidate URL into one of: direct_pdf, disclosure, homepage, none."""
    if not url:
        return "none"
    lo = url.lower()
    if lo.endswith(".pdf"):
        return "direct_pdf"
    if any(k in lo for k in DISCLOSURE_KEYWORDS):
        return "disclosure"
    return "homepage"


# --- Helpers --------------------------------------------------------------

def norm(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", "", s)
    return s.strip()


URL_RE = re.compile(r"https?://[^\s　　]+")


def extract_url(cell: str | None) -> str | None:
    if not cell:
        return None
    m = URL_RE.search(cell)
    return m.group(0) if m else None


def extract_pdf_annotation_links(pdf_path: Path) -> dict[str, str]:
    """Extract hyperlink annotations keyed by the anchor text (school name).

    Some prefecture PDFs (Saitama, partially Aichi) embed URLs as clickable
    hyperlink annotations on the school name cell, NOT as plain text in
    the 備考 column. pdfplumber text extraction misses these entirely;
    PyMuPDF exposes them via page.get_links().
    """
    import fitz  # pymupdf
    out: dict[str, str] = {}
    doc = fitz.open(pdf_path)
    try:
        for page in doc:
            for link in page.get_links():
                uri = link.get("uri") or ""
                rect = link.get("from")
                if not uri or not rect:
                    continue
                anchor_text = page.get_text("text", clip=rect).strip()
                if not anchor_text:
                    continue
                key = norm(anchor_text)
                if key and key not in out:
                    out[key] = uri
    finally:
        doc.close()
    return out


def is_header(row: list[str | None]) -> bool:
    """Detect repeated header rows / pref note rows."""
    joined = "".join(str(c) for c in row if c)
    return any(k in joined for k in ("確認大学等", "学 校 名", "学校名"))


# --- Per-prefecture parsers -----------------------------------------------

def parse_tokyo(pdf_path: Path) -> list[PrefSchool]:
    """Tokyo: 8-column tables. URL = col[7]."""
    out: list[PrefSchool] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or len(row) < 8:
                        continue
                    if is_header(row):
                        continue
                    school_name = (row[2] or "").strip()
                    if not school_name or school_name.isdigit():
                        continue
                    out.append(PrefSchool(
                        pref="tokyo",
                        school_name_raw=school_name,
                        school_name_norm=norm(school_name),
                        address=(row[3] or "").strip(),
                        operator_kind=(row[4] or "").strip(),
                        operator_name=(row[5] or "").strip(),
                        operator_address=(row[6] or "").strip(),
                        disclosure_url=extract_url(row[7]),
                    ))
    return out


def parse_osaka_xlsx(xlsx_path: Path) -> list[PrefSchool]:
    """Osaka: Excel with 8 columns + MEXT 学校番号 in col 2."""
    import openpyxl
    out: list[PrefSchool] = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True, min_row=8):
            if not row or len(row) < 8:
                continue
            num, fy, school_code, name, addr, op_name, op_addr, biko = row[:8]
            if not name or not str(name).strip():
                continue
            if "確認大学等" in str(name) or "学校名" in str(name):
                continue
            out.append(PrefSchool(
                pref="osaka",
                school_name_raw=str(name).strip(),
                school_name_norm=norm(str(name)),
                address=str(addr or "").strip(),
                operator_kind="",
                operator_name=str(op_name or "").strip(),
                operator_address=str(op_addr or "").strip(),
                disclosure_url=extract_url(str(biko or "")),
                school_code=str(school_code).strip() if school_code else None,
            ))
    return out


def parse_7col_hokkaido(pdf_path: Path, pref: str = "hokkaido") -> list[PrefSchool]:
    """Hokkaido: 7-col table with URL in col[5].

    Header row:
      ['確認大学等の名称', '確認大学等の所在地', '設置者の名称',
       '設置者の主たる事務所の所在地', '機関要件\n確認日',
       'ホームページURL', '備考']
    """
    out: list[PrefSchool] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or len(row) < 7:
                        continue
                    if is_header(row):
                        continue
                    school_name = (row[0] or "").strip()
                    if not school_name or "確認大学" in school_name:
                        continue
                    url = extract_url(row[5])
                    out.append(PrefSchool(
                        pref=pref,
                        school_name_raw=school_name,
                        school_name_norm=norm(school_name),
                        address=(row[1] or "").strip(),
                        operator_kind="",
                        operator_name=(row[2] or "").strip(),
                        operator_address=(row[3] or "").strip(),
                        disclosure_url=url,
                    ))
    return out


def parse_5col(pdf_path: Path, pref: str) -> list[PrefSchool]:
    """Kanagawa / Saitama: 5-column tables. URL from col[4] 備考 OR hyperlink annotation on school name."""
    # Pre-extract hyperlink annotations (Saitama-style hidden URLs)
    annotation_links = extract_pdf_annotation_links(pdf_path)

    out: list[PrefSchool] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or len(row) < 5:
                        continue
                    if is_header(row):
                        continue
                    school_name = (row[0] or "").strip()
                    if not school_name:
                        continue

                    # URL priority: (1) text in 備考 col, (2) hyperlink annotation on name
                    url = extract_url(row[4])
                    if not url:
                        url = annotation_links.get(norm(school_name))

                    out.append(PrefSchool(
                        pref=pref,
                        school_name_raw=school_name,
                        school_name_norm=norm(school_name),
                        address=(row[1] or "").strip(),
                        operator_kind="",
                        operator_name=(row[2] or "").strip(),
                        operator_address=(row[3] or "").strip(),
                        disclosure_url=url,
                    ))
    return out


# --- Matching to existing DB schools --------------------------------------

PREF_KEY_TO_DB = {
    "tokyo": "東京都",
    "kanagawa": "神奈川県",
    "saitama": "埼玉県",
    "osaka": "大阪府",
    "fukuoka": "福岡県",
    "hyogo": "兵庫県",
    "shizuoka": "静岡県",
    "okinawa": "沖縄県",
    "miyagi": "宮城県",
    "hokkaido": "北海道",
    "niigata": "新潟県",
    "aichi": "愛知県",
}


def match_school(session, ps: PrefSchool, school_idx: dict, site_idx: dict) -> MatchResult:
    """Per-pref school/site indices pre-built by caller to avoid N+1 queries."""
    # Strategy 0: school_code (MEXT 学校番号) — gold standard, highest priority
    match = None
    strategy = "none"
    if ps.school_code:
        by_code = school_idx.get(f"code:{ps.school_code}")
        if by_code:
            match = by_code
            strategy = "school_code"

    # Strategy 1: exact name match within prefecture
    if not match:
        exact = school_idx.get(f"name:{ps.school_name_raw}")
        if exact:
            match = exact
            strategy = "exact"

    # Strategy 2: NFKC normalized name
    if not match:
        nfkc = school_idx.get(f"nfkc:{ps.school_name_norm}")
        if nfkc:
            match = nfkc
            strategy = "nfkc"

    # Strategy 3: same operator + substring overlap (linear scan fallback, rare)
    if not match and ps.operator_name:
        op_norm = norm(ps.operator_name)
        cand = [
            s for s in school_idx["_all"]
            if norm(s.corporation_name) == op_norm
            and (norm(s.school_name) in ps.school_name_norm or ps.school_name_norm in norm(s.school_name))
        ]
        if len(cand) == 1:
            match = cand[0]
            strategy = "operator_pref"

    existing_urls = site_idx.get(match.id, []) if match else []
    has_url = bool(existing_urls)
    is_new = bool(ps.disclosure_url and match and not has_url)

    return MatchResult(
        pref_school=ps,
        db_school_id=match.id if match else None,
        match_strategy=strategy,
        db_school_name=match.school_name if match else None,
        has_existing_url=has_url,
        is_new_url_candidate=is_new,
    )


def build_indices(session, pref_key: str) -> tuple[dict, dict]:
    """Pre-build per-pref school + site indices (addresses Codex finding #5 N+1)."""
    db_pref = PREF_KEY_TO_DB[pref_key]
    schools = (
        session.query(School)
        .filter(School.prefecture == db_pref, School.status == "active")
        .all()
    )
    school_idx: dict = {"_all": schools}
    for s in schools:
        school_idx[f"name:{s.school_name}"] = s
        school_idx[f"nfkc:{norm(s.school_name)}"] = s
        if s.school_code:
            school_idx[f"code:{s.school_code}"] = s

    school_ids = [s.id for s in schools]
    sites = (
        session.query(SchoolSite)
        .filter(SchoolSite.school_id.in_(school_ids))
        .all()
    ) if school_ids else []
    site_idx: dict = {}
    for site in sites:
        site_idx.setdefault(site.school_id, []).append(site.url)
    return school_idx, site_idx


# --- Report ---------------------------------------------------------------

QUALITY_RANK = {"direct_pdf": 3, "disclosure": 2, "homepage": 1, "none": 0}


def recommend_action(pref_quality: str, existing_qualities: list[str]) -> str:
    """Decide Stage-2 writer action per-school.

    - review: PDF has URL but no DB match (caller handles; here we only see matched)
    - add   : matched + PDF URL present + no existing URL
    - upgrade: matched + PDF URL strictly better than best existing (rank-wise)
    - noop  : matched + PDF URL same or worse than existing (or PDF URL missing)
    """
    if pref_quality == "none":
        return "noop"
    best_existing = max((QUALITY_RANK.get(q, 0) for q in existing_qualities), default=0)
    if best_existing == 0:
        return "add"
    if QUALITY_RANK[pref_quality] > best_existing:
        return "upgrade"
    return "noop"


def build_report(
    pref: str,
    pdf_path: Path,
    results: list[MatchResult],
    site_idx: dict,
) -> PrefReport:
    rep = PrefReport(pref=pref, pdf_path=str(pdf_path))
    by_strat: dict[str, int] = {}
    quality_dist: dict[str, int] = {}
    action_dist: dict[str, int] = {}
    upgrade_candidate_count = 0
    unmatched_samples: list[dict] = []
    new_url_samples: list[dict] = []

    for r in results:
        rep.extracted_total += 1
        if r.pref_school.disclosure_url:
            rep.extracted_with_url += 1
        if r.db_school_id:
            rep.db_matched += 1
        else:
            rep.db_unmatched += 1
        by_strat[r.match_strategy] = by_strat.get(r.match_strategy, 0) + 1

        url_quality = classify_url_quality(r.pref_school.disclosure_url)
        quality_dist[url_quality] = quality_dist.get(url_quality, 0) + 1

        if r.has_existing_url:
            rep.existing_url_coverage += 1

        existing_urls = site_idx.get(r.db_school_id, []) if r.db_school_id else []
        existing_qualities = [classify_url_quality(u) for u in existing_urls]

        if r.match_strategy == "none":
            action = "review"
        else:
            action = recommend_action(url_quality, existing_qualities)
        action_dist[action] = action_dist.get(action, 0) + 1

        quality_upgrade = action == "upgrade"
        if quality_upgrade:
            upgrade_candidate_count += 1

        if r.is_new_url_candidate:
            rep.new_url_candidates += 1
            if len(new_url_samples) < 5:
                new_url_samples.append({
                    "school": r.db_school_name,
                    "url": r.pref_school.disclosure_url,
                })
        if r.match_strategy == "none" and len(unmatched_samples) < 5:
            unmatched_samples.append({
                "pdf_name": r.pref_school.school_name_raw,
                "operator": r.pref_school.operator_name,
            })

        rep.records.append({
            "db_school_id": r.db_school_id,
            "db_school_name": r.db_school_name,
            "pdf_school_name": r.pref_school.school_name_raw,
            "pdf_school_code": r.pref_school.school_code,
            "pdf_address": r.pref_school.address,
            "pdf_operator": r.pref_school.operator_name,
            "pref_url": r.pref_school.disclosure_url,
            "url_quality": url_quality,
            "match_strategy": r.match_strategy,
            "existing_urls": existing_urls,
            "existing_url_quality": existing_qualities,
            "is_new_url_candidate": r.is_new_url_candidate,
            "quality_upgrade_candidate": quality_upgrade,
            "recommended_action": action,
        })

    rep.match_by_strategy = by_strat
    rep.url_quality_distribution = quality_dist
    rep.sample_unmatched = unmatched_samples
    rep.sample_new_urls = new_url_samples
    # store action distribution on the report for summary rollup
    rep.__dict__["action_distribution"] = action_dist
    rep.__dict__["quality_upgrade_candidates"] = upgrade_candidate_count
    return rep


# --- Driver ---------------------------------------------------------------

PARSERS = {
    "tokyo": (parse_tokyo, Path("/tmp/eidp_pref_pdfs/tokyo.pdf")),
    "kanagawa": (lambda p: parse_5col(p, "kanagawa"), Path("/tmp/eidp_pref_pdfs/kanagawa.pdf")),
    "saitama": (lambda p: parse_5col(p, "saitama"), Path("/tmp/eidp_pref_pdfs/saitama.pdf")),
    "osaka": (parse_osaka_xlsx, Path("/tmp/eidp_pref_pdfs/osaka_r8.xlsx")),
    "fukuoka": (lambda p: parse_5col(p, "fukuoka"), Path("/tmp/eidp_pref_pdfs/fukuoka.pdf")),
    "hyogo": (lambda p: parse_5col(p, "hyogo"), Path("/tmp/eidp_pref_pdfs/hyogo.pdf")),
    "shizuoka": (lambda p: parse_5col(p, "shizuoka"), Path("/tmp/eidp_pref_pdfs/shizuoka.pdf")),
    "okinawa": (lambda p: parse_5col(p, "okinawa"), Path("/tmp/eidp_pref_pdfs/okinawa.pdf")),
    "miyagi": (lambda p: parse_5col(p, "miyagi"), Path("/tmp/eidp_pref_pdfs/miyagi.pdf")),
    "hokkaido": (parse_7col_hokkaido, Path("/tmp/eidp_pref_pdfs/hokkaido.pdf")),
}


def main() -> None:
    out_dir = REPO_ROOT / "output" / "pref-aggregator"
    out_dir.mkdir(parents=True, exist_ok=True)

    summary = {"prefectures": {}}

    session = SessionLocal()
    try:
        for pref, (parser, pdf_path) in PARSERS.items():
            if not pdf_path.exists():
                print(f"[skip] {pref}: PDF missing at {pdf_path}", flush=True)
                continue

            schools = parser(pdf_path)
            school_idx, site_idx = build_indices(session, pref)
            results = [match_school(session, s, school_idx, site_idx) for s in schools]
            rep = build_report(pref, pdf_path, results, site_idx)

            out_path = out_dir / f"{pref}.json"
            out_path.write_text(json.dumps(asdict(rep), ensure_ascii=False, indent=2))
            print(f"[done] {pref}: extracted={rep.extracted_total} matched={rep.db_matched} "
                  f"upgrade_candidates={rep.__dict__['quality_upgrade_candidates']} "
                  f"actions={rep.__dict__['action_distribution']} -> {out_path}", flush=True)

            summary["prefectures"][pref] = {
                "extracted_total": rep.extracted_total,
                "extracted_with_url": rep.extracted_with_url,
                "db_matched": rep.db_matched,
                "db_unmatched": rep.db_unmatched,
                "match_by_strategy": rep.match_by_strategy,
                "existing_url_coverage": rep.existing_url_coverage,
                "new_url_candidates": rep.new_url_candidates,
                "url_quality_distribution": rep.url_quality_distribution,
                "action_distribution": rep.__dict__["action_distribution"],
                "quality_upgrade_candidates": rep.__dict__["quality_upgrade_candidates"],
            }
    finally:
        session.close()

    summary_path = out_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"\n=== Summary -> {summary_path} ===")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
