"""Investigate unmatched schools - check against target institution list."""

import csv
import json
import re
import unicodedata
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

warnings.filterwarnings('ignore')

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / 'data' / 'mext'


def normalize(s):
    if not isinstance(s, str):
        return str(s) if s is not None else ''
    return unicodedata.normalize('NFKC', re.sub(r'\s+', '', s or ''))


with open(DATA_DIR / 'unmatched_schools.json', 'r') as f:
    unmatched = json.load(f)

# Load target institution list vocational schools
wb = openpyxl.load_workbook(DATA_DIR / 'target_institutions.xlsx', data_only=True)
ws = wb['20260401']

target_names = {}
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    if row[2] == '専門学校' and row[3]:
        target_names[normalize(row[3])] = {
            'code': row[0],
            'name': row[3],
            'pref': row[6]
        }

print(f'Total vocational schools in target list: {len(target_names)}')

# Match unmatched against target list
target_matched = 0
target_unmatched_list = []
for pref, corp, name in unmatched:
    norm = normalize(name)
    if norm in target_names:
        t = target_names[norm]
        target_matched += 1
    else:
        target_unmatched_list.append((pref, corp, name))

print(f'\nOf {len(unmatched)} unmatched schools:')
print(f'  Found in target institution list (by name): {target_matched}')
print(f'  Not found in either: {len(target_unmatched_list)}')

# Load MEXT school code CSV for cross-check
mext_by_pref = defaultdict(list)
for fname in ['school_code_east.csv', 'school_code_west.csv']:
    with open(DATA_DIR / fname, 'r', encoding='cp932') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if len(row) < 6 or 'H1' not in row[1]:
                continue
            pref_match = re.search(r'\((.+?)\)', row[2])
            pref = pref_match.group(1) if pref_match else row[2]
            mext_by_pref[pref].append({
                'code': row[0],
                'name': row[5],
                'abolished': row[9] if len(row) > 9 else ''
            })

# For remaining unmatched, try to find abolished entries that might match
print(f'\n=== Possible renamed schools (checking abolished entries) ===')
renamed_found = 0
for pref, corp, name in target_unmatched_list[:50]:
    norm_ours = normalize(name)
    candidates = mext_by_pref.get(pref, [])
    # Check abolished schools with similar names
    for c in candidates:
        if c['abolished']:
            norm_mext = normalize(c['name'])
            # Check substring match
            common_chars = sum(1 for a, b in zip(norm_ours, norm_mext) if a == b)
            ratio = common_chars / max(len(norm_ours), len(norm_mext), 1)
            if ratio > 0.5 or norm_ours[:5] == norm_mext[:5]:
                print(f'  [{pref}] Our: {name}')
                print(f'          MEXT(abolished): {c["name"]} (code: {c["code"]}, abolished: {c["abolished"]})')
                renamed_found += 1
                break

print(f'\n  Found {renamed_found} potential renames among abolished entries')

print(f'\n=== Schools not found anywhere (first 30) ===')
count = 0
for pref, corp, name in target_unmatched_list:
    print(f'  [{pref}] {corp} / {name}')
    count += 1
    if count >= 30:
        break
