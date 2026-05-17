"""Analyze the MEXT target institution list (修学支援新制度 対象機関リスト)."""

import openpyxl
import warnings
from collections import Counter
from pathlib import Path

warnings.filterwarnings('ignore')

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / 'data' / 'mext'

wb = openpyxl.load_workbook(DATA_DIR / 'target_institutions.xlsx', data_only=True)
ws = wb['20260401']

type_counts = Counter()
category_counts = Counter()
all_entries = []

for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    school_code = row[0]
    category = row[1]
    school_type = row[2]
    name = row[3]
    if school_type:
        type_counts[school_type] += 1
        category_counts[f'{school_type}({category})'] += 1
        all_entries.append({
            'code': school_code,
            'category': category,
            'type': school_type,
            'name': name,
            'address': row[4],
            'pref_code': row[5],
            'pref_name': row[6]
        })

print(f'Total target institutions: {len(all_entries)}')
print(f'\nBy school type:')
for t, c in type_counts.most_common():
    print(f'  {t}: {c}')

print(f'\nBy school type and category:')
for t, c in category_counts.most_common():
    print(f'  {t}: {c}')

univ_count = sum(c for t, c in type_counts.items() if '大学' in t and '短' not in t)
print(f'\nUniversities (大学) total: {univ_count}')

senmon_count = type_counts.get('専門学校', 0)
print(f'Vocational schools (専門学校) total: {senmon_count}')
