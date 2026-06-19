"""Verify Windows distribution ZIPs before the Windows VM gate.

This verifier is intentionally stricter than a manifest smoke test. It
answers a release question: "is this ZIP complete enough to hand to the
Windows VM offline validation step?"

It does not execute Windows binaries. That remains the Sprint 8.5.b VM gate.
"""

from __future__ import annotations

import argparse
import ast
import csv
import getpass
import hashlib
import importlib.util
import json
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from _packaging_lib import sha256_file  # noqa: E402,F401  (re-exported for tests/callers)
from windows_path_safety import check_windows_safe_paths, issues_by_kind  # noqa: E402


def _load_accepted_wheel_suffixes() -> tuple[str, ...]:
    """Load the wheel suffix contract from build_windows_zip.py.

    This script can be executed directly from any cwd, so do a file-based
    import instead of relying on PYTHONPATH.
    """
    build_script = SCRIPT_DIR / "build_windows_zip.py"
    spec = importlib.util.spec_from_file_location("build_windows_zip_for_verify", build_script)
    if not spec or not spec.loader:
        raise RuntimeError(f"cannot load {build_script}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return tuple(module.ACCEPTED_WHEEL_SUFFIXES)


ACCEPTED_WHEEL_SUFFIXES = _load_accepted_wheel_suffixes()

MIN_SUPPORTED_TARGET_FISCAL_YEAR = 2019
MAX_SUPPORTED_TARGET_FISCAL_YEAR = 2099
SUPPORTED_TARGET_FISCAL_YEAR_RANGE_LABEL = (
    f"[{MIN_SUPPORTED_TARGET_FISCAL_YEAR}, {MAX_SUPPORTED_TARGET_FISCAL_YEAR}]"
)
WINDOWS_ZIP_VERSION_RE = re.compile(r"(?:dist/|C:\\EIDP-staging\\)?eidp-windows-v\d+\.zip", re.IGNORECASE)
SHA256_HEX_RE = re.compile(r"\b[0-9a-fA-F]{64}\b")
GENERIC_LOCAL_USERS = {"", "root", "runner", "vscode", "codespace"}
WINDOWS_REAL_USER_PATH_RE = re.compile(r"C:[\\/]+Users[\\/]+(?!<user>(?:[\\/]|$))[^\\/\s`\"'<>]+", re.IGNORECASE)
MAC_REAL_USER_PATH_RE = re.compile(r"/Users/(?!<user>(?:/|$))[^/\s`\"'<>]+")


@dataclass
class ZipCheck:
    name: str
    path: Path
    ok: bool = True
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    details: dict[str, Any] = field(default_factory=dict)

    def fail(self, message: str) -> None:
        self.ok = False
        self.errors.append(message)

    def warn(self, message: str) -> None:
        self.warnings.append(message)


CORE_REQUIRED_EXACT = (
    "BUILD_INFO.json",
    ".streamlit/config.toml",
    "EIDP-setup.bat",
    "EIDP-start.bat",
    "EIDP-diagnose.bat",
    "EIDP-stage6-evidence.bat",
    "EIDP-stage6-verify-evidence.bat",
    "EIDP-stage6-recovery.bat",
    "EIDP-repair-launcher.bat",
    "README.md",
    "requirements-windows.txt",
    "pyproject.toml",
    "alembic.ini",
    "docs/runbooks/eidp-windows.md",
    "docs/runbooks/eidp-operator-e2e-template.md",
    "scripts/first_setup.bat",
    "scripts/launch.bat",
    "scripts/weekly_run.bat",
    "scripts/bootstrap_pdfs.bat",
    "scripts/diagnose.bat",
    "scripts/collect_stage6_evidence.bat",
    "scripts/collect_bug_report.bat",
    "scripts/verify_stage6_evidence.bat",
    "scripts/uninstall.bat",
    "scripts/validate_install.bat",
    "scripts/stage6_recovery_check.bat",
    "scripts/stage6_residual_cleanup.bat",
    "scripts/offline_pip_install.py",
    "scripts/atomic_write.py",
    "scripts/run_weekly_target_year_discovery.py",
    "scripts/run_r8_rediscovery_weekly.py",
    "scripts/validate_windows_install.py",
    "scripts/collect_stage6_evidence.py",
    "scripts/collect_bug_report.py",
    "scripts/verify_stage6_evidence.py",
    "scripts/verify_stage6_return.py",
    "scripts/build_mature_year_acquisition_proof.py",
    "scripts/evaluate_strict_yield_bound.py",
    "scripts/repair_streamlit_launcher.py",
    "scripts/stage6_recovery_check.py",
    "scripts/stage6_residual_cleanup.py",
    "scripts/bootstrap_pdf_pipeline.py",
    "scripts/ship_gate_contract.py",
    "scripts/download_prefecture_artifacts.py",
    "scripts/prune_release_artifacts.py",
    "scripts/rotate_audit_outbox.py",
    "scripts/prune_pdf_storage.py",
    "scripts/disk_health_check.py",
    "data/authority-index/sources.csv",
    "data/mext/target_institutions.xlsx",
    "data/mext/target_institutions_page.html",
    "data/prefecture-aggregators/seed.csv",
    "data/url-discovery/discovered-urls-50.csv",
    "data/url-discovery/corporation_domains.csv",
    "data/url-discovery/school_domain_overrides.csv",
    "sample/20250826更新版_競合校の在校生数.xlsx",
    "data/discovery-gold-set/README.md",
    "data/discovery-gold-set/schema.json",
    "data/discovery-gold-set/expected-predictions.jsonl",
    "src/eidp/review/app.py",
    "src/eidp/review/operator_pages.py",
    "src/eidp/review/_pages/bug_report.py",
    "src/eidp/bug_signals/detector.py",
    "src/eidp/bug_signals/bundle.py",
    "src/eidp/review/_pages/audit_log.py",
    "src/eidp/review/_pages/settings_page.py",
    "src/eidp/review/_pages/school_year_tasks.py",
    "src/eidp/review/_pages/pdf_manual_entry.py",
    "src/eidp/review/_pages/prefecture_remarks.py",
    "src/eidp/review/_pages/url_candidate_review.py",
    "src/eidp/review/_pages/fiscal_year_override.py",
    "src/eidp/review/_pages/excel_preview.py",
    "src/eidp/review/school_scope.py",
    "src/eidp/cli.py",
    "src/eidp/cli_discovery.py",
    "src/eidp/cli_reports.py",
    "src/eidp/cli_tools.py",
    "src/eidp/windows_platform.py",
    "src/sitecustomize.py",
    "src/eidp/excel/exporter.py",
    "src/eidp/excel/competition_exporter.py",
    "src/eidp/reports/coverage.py",
    "src/eidp/reports/gaps.py",
    "src/eidp/reports/ship_readiness.py",
    "src/eidp/db/audit.py",
    "src/eidp/db/audit_outbox.py",
    "src/eidp/db/sqlite_bootstrap.py",
    "src/eidp/scraper/pdf_discovery.py",
    "src/eidp/pipeline/ingest.py",
    "src/eidp/pipeline/manual_entry.py",
    "src/eidp/pipeline/fiscal_year_override.py",
    "src/eidp/ocr/tesseract.py",
    "src/eidp/ocr/availability.py",
    "src/eidp/scraper/prefecture_aggregator.py",
    "runtime/python/python.exe",
    "runtime/uv.exe",
)

CORE_REQUIRED_PREFIXES = (
    "src/eidp/",
    "migrations/",
    "wheelhouse/",
    "data/discovery-gold-set/entries/",
)

EXPECTED_PREFECTURE_KEYS = frozenset(
    {
        "aichi",
        "akita",
        "aomori",
        "chiba",
        "ehime",
        "fukui",
        "fukuoka",
        "fukushima",
        "gifu",
        "gunma",
        "hiroshima",
        "hokkaido",
        "hyogo",
        "ibaraki",
        "ishikawa",
        "iwate",
        "kagawa",
        "kagoshima",
        "kanagawa",
        "kochi",
        "kumamoto",
        "kyoto",
        "mie",
        "miyagi",
        "miyazaki",
        "nagano",
        "nagasaki",
        "nara",
        "niigata",
        "oita",
        "okayama",
        "okinawa",
        "osaka",
        "saga",
        "saitama",
        "shiga",
        "shimane",
        "shizuoka",
        "tochigi",
        "tokushima",
        "tokyo",
        "tottori",
        "toyama",
        "wakayama",
        "yamagata",
        "yamaguchi",
        "yamanashi",
    }
)

DOWNLOADABLE_PREFECTURE_STATUSES = frozenset({"spiked", "downloaded", "url_found"})
PREFECTURE_ARTIFACT_FORMATS = frozenset({"pdf", "xlsx", "html", "htm"})
DISCOVERY_GOLD_REQUIRED_OUTCOMES = frozenset(
    {
        "accepted_target_pdf",
        "needs_operator_review",
        "no_target_candidate_found",
        "publication_lag_latest_public",
        "site_fetch_error",
    }
)
DISCOVERY_GOLD_ALLOWED_OUTCOMES = DISCOVERY_GOLD_REQUIRED_OUTCOMES
DISCOVERY_GOLD_TRACKED_EXTRACTOR_SOURCES = (
    "embed",
    "wordpress_download_manager",
)
DISCOVERY_GOLD_EXPERIMENTAL_EXTRACTOR_SOURCES = (
    "data_attribute",
    "form_action",
    "input_control",
    "meta_refresh",
    "onclick",
    "select_option",
)

OCR_REQUIRED_EXACT = (
    "ocr-addon/tesseract/tesseract.exe",
    "ocr-addon/tessdata/jpn.traineddata",
    "ocr-addon/tessdata/configs/tsv",
    "ocr-addon/MANIFEST.json",
)

PLAYWRIGHT_REQUIRED_EXACT = (
    "playwright-addon/MANIFEST.json",
)

PLAYWRIGHT_REQUIRED_PREFIXES = (
    "playwright-addon/wheelhouse/",
    "playwright-addon/ms-playwright/",
)


def _read_zip_names(check: ZipCheck) -> set[str] | None:
    if not check.path.is_file():
        check.fail(f"zip does not exist: {check.path}")
        return None
    check.details["size_bytes"] = check.path.stat().st_size
    check.details["sha256"] = sha256_file(check.path)
    if not zipfile.is_zipfile(check.path):
        check.fail(f"not a zip file: {check.path}")
        return None
    with zipfile.ZipFile(check.path) as zf:
        names = zf.namelist()
    seen: set[str] = set()
    duplicates: list[str] = []
    for name in names:
        if name in seen:
            duplicates.append(name)
        else:
            seen.add(name)
    if duplicates:
        check.fail(f"zip contains duplicate entries: {duplicates[:5]}")
    # Reuse the dedupe set we already paid for; the prior `set(names)`
    # at the return site discarded that work and walked the list a second
    # time. Same membership semantics, half the iteration cost on a
    # 50k-entry Chromium ZIP.
    return seen


def _check_exact(check: ZipCheck, names: set[str], required: tuple[str, ...]) -> None:
    for entry in required:
        if entry not in names:
            check.fail(f"missing required entry: {entry}")


def _check_prefixes(check: ZipCheck, names: set[str], required: tuple[str, ...]) -> None:
    for prefix in required:
        if not any(name.startswith(prefix) and name != prefix for name in names):
            check.fail(f"missing required prefix content: {prefix}")


def _check_no_pycache(check: ZipCheck, names: set[str]) -> None:
    pycache = sorted(name for name in names if "__pycache__" in name)
    if pycache:
        check.fail(f"zip contains __pycache__ entries: {pycache[:5]}")


def _check_windows_safe_paths(check: ZipCheck, names: set[str]) -> None:
    """Reject names that are valid in ZIPs but unsafe on Windows.

    Windows treats paths case-insensitively and reserves device names
    like CON/PRN/AUX/NUL even when an extension is present. Catch these
    before handing artifacts to the VM.
    """
    grouped = issues_by_kind(check_windows_safe_paths(names))
    if collisions := grouped.get("case-collision"):
        check.fail(f"zip contains Windows case-insensitive path collisions: {collisions[:5]}")
    if reserved := grouped.get("reserved-name"):
        check.fail(f"zip contains Windows reserved path components: {reserved[:5]}")
    absolute_or_parent = [
        *grouped.get("absolute-path", []),
        *grouped.get("parent-directory", []),
    ]
    if absolute_or_parent:
        check.fail(f"zip contains absolute or parent-directory paths: {absolute_or_parent[:5]}")
    if trailing := grouped.get("trailing-dot-space"):
        check.fail(f"zip contains Windows trailing-dot/space path components: {trailing[:5]}")


def _check_no_runtime_data(check: ZipCheck, names: set[str]) -> None:
    """Reject mutable operator runtime data from the handoff ZIP.

    The core ZIP may ship static bootstrap inputs such as master.xlsx,
    prefecture seeds, URL seeds, and discovery gold-set fixtures. It must not
    carry a developer/operator SQLite DB, lock file, downloaded PDFs, audit
    outbox, or generated Excel/log state from the build host.
    """
    forbidden_exact = {
        "data/.lock",
        "data/eidp.sqlite3",
        "data/eidp.sqlite3-wal",
        "data/eidp.sqlite3-shm",
    }
    forbidden_prefixes = (
        "data/audit/",
        "data/output/",
        "data/pdfs/",
        "data/prefecture-aggregators/artifacts/",
    )
    runtime_entries = sorted(
        name
        for name in names
        if name in forbidden_exact or any(name.startswith(prefix) for prefix in forbidden_prefixes)
    )
    if runtime_entries:
        check.fail(f"zip contains mutable runtime data entries: {runtime_entries[:8]}")


def _check_no_secret_files(check: ZipCheck, names: set[str]) -> None:
    private_key_names = {
        "id_dsa",
        "id_ecdsa",
        "id_ed25519",
        "id_rsa",
    }
    secret_entries: list[str] = []
    for name in sorted(names):
        base_name = name.rsplit("/", 1)[-1].lower()
        if base_name == ".env" or (base_name.startswith(".env.") and base_name != ".env.example"):
            secret_entries.append(name)
            continue
        if base_name in private_key_names or base_name.endswith("_private_key.pem"):
            secret_entries.append(name)
    if secret_entries:
        check.fail(f"zip contains local secret/key files: {secret_entries[:8]}")


def _check_no_historical_runbooks(check: ZipCheck, names: set[str]) -> None:
    historical_runbooks = sorted(
        name for name in names
        if name.startswith("docs/runbooks/eidp-v")
    )
    if historical_runbooks:
        check.fail(
            "zip contains historical handoff runbooks; ship only current operator docs: "
            f"{historical_runbooks[:10]}"
        )


def _check_no_demo_prototype_runtime(check: ZipCheck, names: set[str]) -> None:
    """Reject design/demo prototype assets from the operator runtime ZIP.

    The operations-console HTML prototype is useful as a design contract, but
    it is not the Streamlit production UI. If generated runtime files or
    standalone demo HTML enter the Windows ZIP, operators can mistake a mock UI
    for the audited application.
    """
    forbidden_prefixes = (
        "UI-example/",
        "docs/design/operations-console-demo/",
    )
    forbidden_names = sorted(
        name
        for name in names
        if any(name.startswith(prefix) for prefix in forbidden_prefixes)
        or Path(name).name.lower() == "support.js"
        or Path(name).name.lower().endswith(".standalone.html")
    )
    if forbidden_names:
        check.fail(
            "zip contains demo/prototype UI assets; keep HTML prototypes under docs/design "
            f"and out of the Windows operator runtime: {forbidden_names[:8]}"
        )


def _check_wheelhouse(check: ZipCheck, names: set[str], *, require_project: bool) -> None:
    wheels = sorted(name for name in names if name.startswith("wheelhouse/") and name.endswith(".whl"))
    check.details["wheel_count"] = len(wheels)
    if not wheels:
        check.fail("wheelhouse contains no wheels")
        return

    rejected = [
        wheel for wheel in wheels
        if not any(Path(wheel).name.endswith(suffix) for suffix in ACCEPTED_WHEEL_SUFFIXES)
    ]
    if rejected:
        check.fail(f"wheelhouse contains rejected wheel names: {rejected[:5]}")

    by_distribution: dict[str, list[str]] = {}
    for wheel in wheels:
        distribution = Path(wheel).name.split("-", 1)[0].lower().replace("_", "-")
        by_distribution.setdefault(distribution, []).append(wheel)
    duplicates = {
        distribution: values
        for distribution, values in by_distribution.items()
        if len(values) > 1
    }
    if duplicates:
        sample = {
            distribution: values[:5]
            for distribution, values in sorted(duplicates.items())[:5]
        }
        check.fail(f"wheelhouse contains duplicate distributions: {sample}")

    project_wheels = [wheel for wheel in wheels if Path(wheel).name.startswith("eidp-")]
    check.details["project_wheel_count"] = len(project_wheels)
    if require_project and not project_wheels:
        check.fail("wheelhouse missing project wheel eidp-*.whl")
    if require_project and len(project_wheels) > 1:
        check.fail(f"wheelhouse contains multiple project wheels: {project_wheels[:5]}")


def _read_zip_text(check: ZipCheck, member: str) -> str | None:
    if not zipfile.is_zipfile(check.path):
        return None
    with zipfile.ZipFile(check.path) as zf:
        try:
            raw = zf.read(member)
        except KeyError:
            return None
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        check.fail(f"{member} is not UTF-8 text: {exc}")
        return None


def _read_zip_bytes(check: ZipCheck, member: str) -> bytes | None:
    if not zipfile.is_zipfile(check.path):
        return None
    with zipfile.ZipFile(check.path) as zf:
        try:
            return zf.read(member)
        except KeyError:
            return None


def _parser_keys_from_source(check: ZipCheck, source: str, member: str) -> set[str]:
    """Extract parser registry keys from the packaged source file.

    Do not import the local checkout here: the release question is whether
    the ZIP itself carries parser support for every official seed row.
    """
    try:
        tree = ast.parse(source)
    except SyntaxError as exc:
        check.fail(f"{member} cannot be parsed for prefecture parser registry: {exc}")
        return set()

    for node in tree.body:
        value: ast.expr | None = None
        if isinstance(node, ast.Assign) and any(
            isinstance(target, ast.Name) and target.id == "PARSERS"
            for target in node.targets
        ):
            value = node.value
        elif isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name) and node.target.id == "PARSERS":
            value = node.value
        if value is None:
            continue
        if not isinstance(value, ast.Dict):
            check.fail(f"{member} PARSERS must be a literal dict for ZIP verification")
            return set()
        keys: set[str] = set()
        for key_node in value.keys:
            if isinstance(key_node, ast.Constant) and isinstance(key_node.value, str):
                keys.add(key_node.value)
        return keys

    check.fail(f"{member} missing PARSERS registry")
    return set()


def _truthy_csv_value(value: str | None) -> bool:
    normalized = (value or "").strip().lower()
    return bool(normalized) and normalized not in {"no", "n/a", "unknown", "tbd", "false", "0"}


def _split_artifact_urls(raw: str | None) -> list[str]:
    return [
        part.strip()
        for part in (raw or "").replace("\n", "|").replace(";", "|").split("|")
        if part.strip()
    ]


def _check_prefecture_seed_contract(check: ZipCheck, names: set[str]) -> None:
    """Validate the official-prefecture bootstrap surface shipped in the ZIP.

    The product objective depends on starting from the 47 government/prefecture
    confirmation indexes. A ZIP can be structurally complete while silently
    shipping a partial seed or missing parser registration, so make that a
    release gate rather than a report-only observation.
    """
    seed_member = "data/prefecture-aggregators/seed.csv"
    parser_member = "src/eidp/scraper/prefecture_aggregator.py"
    if seed_member not in names or parser_member not in names:
        return

    seed_body = _read_zip_text(check, seed_member)
    parser_body = _read_zip_text(check, parser_member)
    if seed_body is None or parser_body is None:
        return

    records = list(csv.DictReader(seed_body.splitlines()))
    parser_keys = _parser_keys_from_source(check, parser_body, parser_member)
    if not records or not parser_keys:
        return

    pref_keys = [(record.get("pref_key") or "").strip() for record in records]
    pref_key_set = set(pref_keys)
    duplicates = sorted({pref for pref in pref_keys if pref and pref_keys.count(pref) > 1})
    if duplicates:
        check.fail(f"{seed_member} contains duplicate pref_key values: {duplicates}")

    missing_rows = sorted(EXPECTED_PREFECTURE_KEYS - pref_key_set)
    unexpected_rows = sorted(pref_key_set - EXPECTED_PREFECTURE_KEYS)
    if missing_rows or unexpected_rows or len(records) != len(EXPECTED_PREFECTURE_KEYS):
        check.fail(
            f"{seed_member} must contain exactly {len(EXPECTED_PREFECTURE_KEYS)} current prefecture rows; "
            f"missing={missing_rows} unexpected={unexpected_rows} actual={len(records)}"
        )

    unsupported = sorted(pref for pref in pref_key_set if pref and pref not in parser_keys)
    if unsupported:
        check.fail(f"{parser_member} PARSERS missing seed prefectures: {unsupported}")

    non_downloadable: list[str] = []
    missing_artifact_url: list[str] = []
    bad_artifact_format: list[str] = []
    with_school_link_signal = 0
    supplemental_rows = 0
    school_total = 0
    for record in records:
        pref = (record.get("pref_key") or "").strip()
        artifact_url = (record.get("artifact_url") or "").strip()
        status = (record.get("verified_status") or "").strip()
        artifact_format = (record.get("artifact_format") or "").strip().lower()
        if status not in DOWNLOADABLE_PREFECTURE_STATUSES:
            non_downloadable.append(f"{pref}:{status or '<blank>'}")
        if not artifact_url.startswith("http"):
            missing_artifact_url.append(pref or "<blank>")
        if artifact_format not in PREFECTURE_ARTIFACT_FORMATS:
            bad_artifact_format.append(f"{pref}:{artifact_format or '<blank>'}")
        if _truthy_csv_value(record.get("has_url_col")) or _truthy_csv_value(record.get("has_hyperlink_annot")):
            with_school_link_signal += 1
        if any(url.startswith("http") for url in _split_artifact_urls(record.get("supplemental_artifact_urls"))):
            supplemental_rows += 1
        raw_school_count = (record.get("schools_in_db") or "").strip()
        if raw_school_count and raw_school_count.lower() != "unknown":
            try:
                school_total += int(raw_school_count)
            except ValueError:
                check.warn(f"{seed_member} has non-integer schools_in_db for {pref}")

    if non_downloadable:
        check.fail(f"{seed_member} has non-downloadable prefecture statuses: {non_downloadable}")
    if missing_artifact_url:
        check.fail(f"{seed_member} has missing artifact URLs: {missing_artifact_url}")
    if bad_artifact_format:
        check.fail(f"{seed_member} has unsupported artifact formats: {bad_artifact_format}")

    check.details["prefecture_seed_rows"] = len(records)
    check.details["prefecture_seed_parser_supported"] = len(pref_key_set & parser_keys)
    check.details["prefecture_seed_downloadable"] = sum(
        1
        for record in records
        if (record.get("verified_status") or "").strip() in DOWNLOADABLE_PREFECTURE_STATUSES
        and (record.get("artifact_url") or "").strip().startswith("http")
    )
    check.details["prefecture_seed_school_rows_total"] = school_total
    check.details["prefecture_seed_with_school_link_signal"] = with_school_link_signal
    check.details["prefecture_seed_supplemental_rows"] = supplemental_rows


def _check_mext_authority_index_contract(check: ZipCheck, names: set[str]) -> None:
    """Validate the T0 MEXT authority-index surface shipped in the ZIP.

    The product objective covers both vocational/specialty schools and roughly
    700 universities. The prefecture seed proves the vocational official-index
    lane; this gate keeps the national MEXT confirmed-institution list from
    being dropped from release packages or replaced by broad search inputs.
    """

    catalog_member = "data/authority-index/sources.csv"
    page_member = "data/mext/target_institutions_page.html"
    workbook_member = "data/mext/target_institutions.xlsx"
    if catalog_member not in names or page_member not in names or workbook_member not in names:
        return

    catalog_body = _read_zip_text(check, catalog_member)
    page_body = _read_zip_text(check, page_member)
    workbook_body = _read_zip_bytes(check, workbook_member)
    if catalog_body is None or page_body is None or workbook_body is None:
        return

    records = list(csv.DictReader(catalog_body.splitlines()))
    mext_records = [
        row
        for row in records
        if (row.get("source_id") or "").strip() == "mext_target_institutions_20260401"
    ]
    if len(mext_records) != 1:
        check.fail(
            f"{catalog_member} must contain exactly one mext_target_institutions_20260401 row"
        )
        return
    mext_record = mext_records[0]

    source_url = (mext_record.get("source_url") or "").strip()
    artifact_path = (mext_record.get("artifact_path") or "").strip()
    trust_tier = (mext_record.get("trust_tier") or "").strip()
    authority_type = (mext_record.get("authority_type") or "").strip()
    institution_types = (mext_record.get("institution_types") or "").strip()
    auto_accept = (mext_record.get("auto_accept_allowed") or "").strip().lower()
    if not source_url.startswith("https://www.mext.go.jp/"):
        check.fail(f"{catalog_member} MEXT source_url must be an official mext.go.jp URL")
    if artifact_path != workbook_member:
        check.fail(f"{catalog_member} MEXT artifact_path must point to {workbook_member}")
    if trust_tier != "t0_mext":
        check.fail(f"{catalog_member} MEXT trust_tier must be t0_mext")
    if authority_type != "mext":
        check.fail(f"{catalog_member} MEXT authority_type must be mext")
    if auto_accept != "yes":
        check.fail(f"{catalog_member} MEXT auto_accept_allowed must be yes")
    for required_type in ("大学", "専門学校"):
        if required_type not in institution_types:
            check.fail(f"{catalog_member} MEXT institution_types missing {required_type}")

    if "mext.go.jp" not in page_body or "20260401" not in page_body or ".xlsx" not in page_body:
        check.fail(f"{page_member} does not look like the MEXT target-institutions index page")

    try:
        workbook = openpyxl.load_workbook(BytesIO(workbook_body), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception taxonomy is broad.
        check.fail(f"{workbook_member} cannot be read as XLSX: {exc}")
        return
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        counts: dict[str, int] = {}
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            school_code = str(row[0]).strip() if row and row[0] else ""
            school_type = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if not school_code or not school_type:
                continue
            counts[school_type] = counts.get(school_type, 0) + 1
    finally:
        workbook.close()

    total_rows = sum(counts.values())
    university_rows = counts.get("大学", 0)
    specialty_rows = counts.get("専門学校", 0)
    short_college_rows = counts.get("短期大学", 0)
    kosen_rows = counts.get("高等専門学校", 0)

    check.details["mext_target_total_rows"] = total_rows
    check.details["mext_target_university_rows"] = university_rows
    check.details["mext_target_specialty_rows"] = specialty_rows
    check.details["mext_target_short_college_rows"] = short_college_rows
    check.details["mext_target_kosen_rows"] = kosen_rows

    if total_rows < 3000:
        check.fail(f"{workbook_member} target-institution row count too low: {total_rows}")
    if university_rows < 700:
        check.fail(f"{workbook_member} university rows below 700-scope floor: {university_rows}")
    if specialty_rows < 1700:
        check.fail(f"{workbook_member} specialty-school rows below 1700-scope floor: {specialty_rows}")
    if short_college_rows < 200:
        check.fail(f"{workbook_member} short-college rows below expected floor: {short_college_rows}")
    if kosen_rows < 50:
        check.fail(f"{workbook_member} kosen rows below expected floor: {kosen_rows}")


def _read_zip_json(check: ZipCheck, member: str, *, label: str) -> dict[str, Any] | None:
    body = _read_zip_text(check, member)
    if body is None:
        return None
    try:
        payload = json.loads(body)
    except json.JSONDecodeError as exc:
        check.fail(f"{member} invalid {label} JSON: {exc}")
        return None
    if not isinstance(payload, dict):
        check.fail(f"{member} {label} JSON must contain an object")
        return None
    return payload


def _check_discovery_gold_set_contract(check: ZipCheck, names: set[str]) -> None:
    schema_member = "data/discovery-gold-set/schema.json"
    if schema_member not in names:
        return
    _read_zip_json(check, schema_member, label="discovery gold-set schema")

    entry_members = sorted(
        name
        for name in names
        if name.startswith("data/discovery-gold-set/entries/") and name.endswith(".json")
    )
    if not entry_members:
        return

    outcome_counts: dict[str, int] = {}
    pattern_type_counts: dict[str, int] = {}
    pattern_source_counts: dict[str, int] = {}
    expected_predictions: dict[str, dict[str, Any]] = {}
    for member in entry_members:
        payload = _read_zip_json(check, member, label="discovery gold-set")
        if payload is None:
            continue
        schema_version = payload.get("schema_version")
        if schema_version != "discovery-gold-set/v0.1":
            check.fail(f"{member} has unsupported discovery gold-set schema_version: {schema_version!r}")
        entry_id = payload.get("entry_id")
        if not isinstance(entry_id, str) or not entry_id:
            check.fail(f"{member} missing discovery gold-set entry_id")
        outcome = payload.get("outcome")
        if not isinstance(outcome, str) or not outcome:
            check.fail(f"{member} missing discovery gold-set outcome")
            continue
        outcome_counts[outcome] = outcome_counts.get(outcome, 0) + 1
        expected_result = payload.get("expected_result")
        if isinstance(expected_result, dict):
            pattern_type = str(expected_result.get("pattern_type") or "")
            if pattern_type:
                pattern_type_counts[pattern_type] = pattern_type_counts.get(pattern_type, 0) + 1
                pattern_source = _discovery_pattern_source(pattern_type)
                if pattern_source:
                    pattern_source_counts[pattern_source] = pattern_source_counts.get(pattern_source, 0) + 1
        _check_discovery_gold_entry_semantics(check, member, payload)
        if isinstance(entry_id, str) and entry_id:
            expected_predictions[entry_id] = _expected_prediction_from_gold_entry(payload)

    missing = sorted(DISCOVERY_GOLD_REQUIRED_OUTCOMES - set(outcome_counts))
    if missing:
        check.fail(f"missing discovery gold-set outcomes: {missing}")

    check.details["discovery_gold_set_entries"] = len(entry_members)
    check.details["discovery_gold_set_outcomes"] = dict(sorted(outcome_counts.items()))
    check.details["discovery_gold_pattern_types"] = dict(sorted(pattern_type_counts.items()))
    check.details["discovery_gold_pattern_sources"] = dict(sorted(pattern_source_counts.items()))
    check.details["discovery_gold_undemonstrated_pattern_sources"] = [
        source for source in DISCOVERY_GOLD_TRACKED_EXTRACTOR_SOURCES if source not in pattern_source_counts
    ]
    if check.details["discovery_gold_undemonstrated_pattern_sources"]:
        check.warn(
            "tracked PDF discovery extractor sources lack gold-set demonstrations: "
            f"{check.details['discovery_gold_undemonstrated_pattern_sources']}"
        )
    _check_discovery_gold_expected_predictions(check, names, expected_predictions)


def _check_discovery_gold_entry_semantics(check: ZipCheck, member: str, payload: dict[str, Any]) -> None:
    outcome = payload.get("outcome")
    school = payload.get("school")
    expected = payload.get("expected_result")
    if outcome not in DISCOVERY_GOLD_ALLOWED_OUTCOMES:
        check.fail(f"{member} has unsupported discovery gold-set outcome: {outcome!r}")
        return
    if not isinstance(school, dict):
        check.fail(f"{member} missing discovery gold-set school object")
        return
    if not isinstance(expected, dict):
        check.fail(f"{member} missing discovery gold-set expected_result object")
        return

    target_year = _strict_int(payload.get("target_fiscal_year"))
    school_id = _strict_int(school.get("school_id"))
    fiscal_year = _strict_int(expected.get("fiscal_year"))
    pdf_url = str(expected.get("pdf_url") or "")
    pdf_type = str(expected.get("pdf_type") or "")
    strict_success = expected.get("strict_target_year_success")

    if (
        target_year is None
        or target_year < MIN_SUPPORTED_TARGET_FISCAL_YEAR
        or target_year > MAX_SUPPORTED_TARGET_FISCAL_YEAR
    ):
        check.fail(
            f"{member} target_fiscal_year must be an integer in {SUPPORTED_TARGET_FISCAL_YEAR_RANGE_LABEL}"
        )
    if school_id is None or school_id <= 0:
        check.fail(f"{member} school.school_id must be a positive integer")

    if outcome == "accepted_target_pdf":
        if not pdf_url:
            check.fail(f"{member} accepted_target_pdf requires expected_result.pdf_url")
        if pdf_type != "target":
            check.fail(f"{member} accepted_target_pdf requires expected_result.pdf_type=target")
        if fiscal_year != target_year:
            check.fail(f"{member} accepted_target_pdf fiscal_year must equal target_fiscal_year")
        if strict_success is not True:
            check.fail(f"{member} accepted_target_pdf requires strict_target_year_success=true")
    elif outcome == "publication_lag_latest_public":
        if not pdf_url:
            check.fail(f"{member} publication_lag_latest_public requires expected_result.pdf_url")
        if pdf_type != "target":
            check.fail(f"{member} publication_lag_latest_public requires expected_result.pdf_type=target")
        if fiscal_year is None or target_year is None or fiscal_year >= target_year:
            check.fail(f"{member} publication_lag_latest_public fiscal_year must be older than target_fiscal_year")
        if strict_success is True:
            check.fail(f"{member} publication_lag_latest_public requires strict_target_year_success=false")
    elif outcome == "needs_operator_review":
        if strict_success is True:
            check.fail(f"{member} needs_operator_review requires strict_target_year_success=false")
    elif outcome in {"no_target_candidate_found", "site_fetch_error"}:
        if pdf_url:
            check.fail(f"{member} {outcome} must not carry expected_result.pdf_url")
        if fiscal_year is not None:
            check.fail(f"{member} {outcome} must not carry expected_result.fiscal_year")
        if strict_success is True:
            check.fail(f"{member} {outcome} requires strict_target_year_success=false")


def _strict_int(value: object) -> int | None:
    if isinstance(value, bool) or not isinstance(value, int):
        return None
    return value


def _expected_prediction_from_gold_entry(payload: dict[str, Any]) -> dict[str, Any]:
    expected = payload.get("expected_result")
    expected_result: dict[str, Any] = expected if isinstance(expected, dict) else {}
    prediction = {
        "entry_id": str(payload.get("entry_id") or ""),
        "outcome": str(payload.get("outcome") or ""),
        "pdf_url": str(expected_result.get("pdf_url") or ""),
        "fiscal_year": expected_result.get("fiscal_year"),
        "strict_target_year_success": bool(expected_result.get("strict_target_year_success", False)),
    }
    if pattern_type := str(expected_result.get("pattern_type") or ""):
        prediction["pattern_type"] = pattern_type
    return prediction


def _discovery_pattern_source(pattern_type: str) -> str:
    if pattern_type == "wordpress_download_manager":
        return pattern_type
    for suffix in ("_cache_busted", "_direct", "_wordpress"):
        if pattern_type.endswith(suffix):
            return pattern_type.removesuffix(suffix)
    return pattern_type


def _check_discovery_gold_expected_predictions(
    check: ZipCheck,
    names: set[str],
    expected_by_entry_id: dict[str, dict[str, Any]],
) -> None:
    member = "data/discovery-gold-set/expected-predictions.jsonl"
    if member not in names:
        check.fail(f"{member} missing from core ZIP")
        return

    body = _read_zip_text(check, member)
    if body is None:
        return

    seen: dict[str, dict[str, Any]] = {}
    for line_no, line in enumerate(body.splitlines(), start=1):
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            check.fail(f"{member}:{line_no} invalid JSONL prediction: {exc}")
            continue
        if not isinstance(payload, dict):
            check.fail(f"{member}:{line_no} prediction must be a JSON object")
            continue
        entry_id = payload.get("entry_id")
        if not isinstance(entry_id, str) or not entry_id:
            check.fail(f"{member}:{line_no} prediction missing entry_id")
            continue
        if entry_id in seen:
            check.fail(f"{member}:{line_no} duplicate prediction entry_id: {entry_id}")
        seen[entry_id] = payload

    for entry_id, expected in sorted(expected_by_entry_id.items()):
        prediction = seen.get(entry_id)
        if prediction is None:
            check.fail(f"{member} missing prediction for discovery gold-set entry: {entry_id}")
            continue
        for key, expected_value in expected.items():
            if prediction.get(key) != expected_value:
                check.fail(
                    f"{member} prediction mismatch for {entry_id}: "
                    f"{key}={prediction.get(key)!r}, expected {expected_value!r}"
                )

    for entry_id in sorted(set(seen) - set(expected_by_entry_id)):
        check.fail(f"{member} contains unexpected prediction entry_id: {entry_id}")

    check.details["discovery_gold_expected_predictions"] = len(seen)


def _require_text(check: ZipCheck, body: str, member: str, needle: str) -> None:
    if needle not in body:
        check.fail(f"{member} missing required token: {needle}")


def _reject_text(check: ZipCheck, body: str, member: str, needle: str) -> None:
    if needle in body:
        check.fail(f"{member} contains forbidden token: {needle}")


def _local_user_path_tokens() -> tuple[str, ...]:
    users = {getpass.getuser(), Path.home().name}
    users.update(
        user.strip()
        for user in os.environ.get("EIDP_FORBIDDEN_LOCAL_USERS", "").split(",")
        if user.strip()
    )
    tokens: list[str] = []
    for user in sorted(users - GENERIC_LOCAL_USERS):
        tokens.extend((user, f"/Users/{user}", f"C:\\Users\\{user}", f"C:/Users/{user}"))
    return tuple(tokens)


def _reject_local_user_path_tokens(check: ZipCheck, body: str, member: str) -> None:
    if WINDOWS_REAL_USER_PATH_RE.search(body) or MAC_REAL_USER_PATH_RE.search(body):
        check.fail(f"{member} contains local-user path token")
        return
    if any(token in body for token in _local_user_path_tokens()):
        check.fail(f"{member} contains local-user path token")


def _reject_bare_rc_assignment(check: ZipCheck, body: str, member: str) -> None:
    """Catch stale launcher lines like ``"RC=-1"`` that cmd tries to execute."""
    for lineno, line in enumerate(body.splitlines(), start=1):
        stripped = line.strip()
        if stripped.startswith('"RC='):
            check.fail(f"{member} line {lineno} has bare RC assignment; use set \"RC=%ERRORLEVEL%\"")


def _section_until_next_level2_heading(body: str, start_token: str) -> str | None:
    start = body.find(start_token)
    if start < 0:
        return None
    next_heading = re.search(r"\n##\s+\d+\.", body[start + 1 :])
    if next_heading is None:
        return body[start:]
    return body[start : start + 1 + next_heading.start()]


def _check_operator_e2e_template_version_neutral_fields(check: ZipCheck, body: str, member: str) -> None:
    """Keep the current package handoff fields reusable across ZIP versions."""
    guarded_sections = (
        ("current candidate", "現行投入候補（Mac / non-Windows gate 済み、Windows 未実証）:"),
        ("evidence collection commands", "## 3. 証跡採取コマンド"),
    )
    for label, start_token in guarded_sections:
        section = _section_until_next_level2_heading(body, start_token)
        if section is None:
            check.fail(f"{member} missing version-neutral {label} section: {start_token}")
            continue
        if match := WINDOWS_ZIP_VERSION_RE.search(section):
            check.fail(
                f"{member} {label} section has hard-coded Windows ZIP version; "
                f"use vXXX/<core-zip-file-name> placeholder instead: {match.group(0)}"
            )
        if match := SHA256_HEX_RE.search(section):
            check.fail(
                f"{member} {label} section has hard-coded SHA256; "
                f"copy the digest from the sidecar/current release status at execution time: {match.group(0)}"
            )


def _check_bat_common(check: ZipCheck, member: str, body: str) -> None:
    _require_text(check, body, member, 'cd /d "%~dp0\\.."')
    _require_text(check, body, member, 'set "EIDP_APP_ROOT=%CD%"')
    _require_text(check, body, member, "PYTHONIOENCODING=utf-8")
    _require_text(check, body, member, "PYTHONUTF8=1")


def _check_bat_contracts(check: ZipCheck, names: set[str]) -> None:
    """Validate packaged .bat launchers, not just their filenames.

    Unit tests already pin the source files. This verifier catches a stale
    or hand-edited ZIP before it reaches the Windows VM gate.
    """
    required_tokens: dict[str, tuple[str, ...]] = {
        "EIDP-setup.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\first_setup.bat"',
            'set "RC=%ERRORLEVEL%"',
            "EIDP-start.bat",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-start.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\launch.bat"',
            'set "RC=%ERRORLEVEL%"',
            "EIDP-setup.bat",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-diagnose.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\diagnose.bat"',
            'set "RC=%ERRORLEVEL%"',
            "Diagnostics collected",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-stage6-evidence.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\collect_stage6_evidence.bat"',
            'set "RC=%ERRORLEVEL%"',
            "Stage 6 evidence bundle created",
            "logs\\stage6-evidence-*.zip",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-stage6-verify-evidence.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\verify_stage6_evidence.bat"',
            'set "RC=%ERRORLEVEL%"',
            "Stage 6 evidence ZIP verified",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-stage6-recovery.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\stage6_recovery_check.bat" %*',
            'set "RC=%ERRORLEVEL%"',
            "Stage 6 recovery check passed",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "EIDP-repair-launcher.bat": (
            'cd /d "%~dp0"',
            'call "%~dp0scripts\\repair_streamlit_launcher.bat" %*',
            'set "RC=%ERRORLEVEL%"',
            "Launcher repair check completed",
            "pause",
            "endlocal & exit /b %RC%",
        ),
        "scripts/first_setup.bat": (
            "runtime\\python\\python.exe",
            "scripts\\offline_pip_install.py",
            ".setup.lock",
            "SETUP_LOCK_STALE_HOURS=2",
            "Removed stale setup lock",
            "setup is already running in this folder",
            "endlocal & exit /b %SETUP_RC%",
            "PYTHONPATH=%EIDP_APP_ROOT%\\src",
            "OFFLINE_PIP",
            "VENV_SITE_PACKAGES=%EIDP_APP_ROOT%\\.venv\\Lib\\site-packages",
            "venv",
            "-m venv --without-pip",
            ".venv\\Scripts\\python.exe",
            '"%RUNTIME_PY%" "%OFFLINE_PIP%" install',
            '--target "%VENV_SITE_PACKAGES%"',
            "--no-index",
            "--no-cache-dir",
            "--upgrade",
            "--force-reinstall",
            "--no-deps",
            "EIDP_WHEEL",
            "eidp-*.whl",
            "wheelhouse",
            "-m eidp.cli db-bootstrap --sqlite",
            "import-excel",
            "rebuild-school-year-tasks",
            "EIDP_REGISTER_WEEKLY_TASK",
            "schtasks",
            "New-ScheduledTaskSettingsSet",
            "-RestartCount 3",
            "-RestartInterval (New-TimeSpan -Minutes 30)",
            "Set-ScheduledTask -TaskName 'EIDP Weekly Run'",
            "retry-on-failure is not configured",
            "validate_install.bat",
            "--after-setup",
        ),
        "scripts/launch.bat": (
            ".venv\\Scripts\\python.exe",
            "PYTHONPATH=%EIDP_APP_ROOT%\\src",
            "Start-Process 'http://localhost:8501'",
            "-m streamlit run",
            "streamlit run",
            "--server.address 127.0.0.1",
            "--server.headless true",
            "--browser.gatherUsageStats false",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        ".streamlit/config.toml": (
            "[server]",
            'address = "127.0.0.1"',
            "headless = true",
            "runOnSave = false",
            "[browser]",
            "gatherUsageStats = false",
        ),
        "scripts/weekly_run.bat": (
            ".venv\\Scripts\\python.exe",
            "PYTHONPATH=%EIDP_APP_ROOT%\\src",
            "Get-Date -Format yyyyMMdd",
            "run_weekly_target_year_discovery.py",
            "EIDP_WEEKLY_LIMIT",
            "--limit %EIDP_WEEKLY_LIMIT%",
            "%*",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/bootstrap_pdfs.bat": (
            ".venv\\Scripts\\python.exe",
            "bootstrap_pdf_pipeline.py",
            "Get-Date -Format yyyyMMdd-HHmmss",
            "bootstrap-pdfs-%RUN_ID%.log",
            "bootstrap-pdfs-%RUN_ID%.json",
            '--progress-file "%PROGRESS_PATH%"',
            '> "%LOG_PATH%" 2>&1',
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/diagnose.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "diagnostics-%DIAG_STAMP%.txt",
            "validate_windows_install.py",
            "--after-setup",
            "--after-bootstrap",
            "validate_after_bootstrap_rc",
            'set "VALIDATE_BOOTSTRAP_RC=!ERRORLEVEL!"',
            "validate_after_bootstrap_ship_gate_rc",
            'set "VALIDATE_BOOTSTRAP_SHIP_GATE_RC=!ERRORLEVEL!"',
            "--after-weekly",
            "validate_after_weekly_rc",
            'set "VALIDATE_WEEKLY_RC=!ERRORLEVEL!"',
            "validate_after_weekly_ship_gate_rc",
            'set "VALIDATE_WEEKLY_SHIP_GATE_RC=!ERRORLEVEL!"',
            "--require-ship-gate",
            "BUILD_INFO.json",
            "last_run.json",
            "final objective ship readiness",
            "report ship-readiness --json --fail-on-missing-goal",
            "ship_readiness_rc",
            "retroactive fiscal-year ship readiness",
            "settings.target_fiscal_year",
            "'ship-readiness', '--fy'",
            "retroactive_fiscal_year",
            "retroactive_ship_readiness_rc",
            "stage6 recovery check",
            "stage6_recovery_check.py",
            "stage6_recovery_rc",
            "latest discovery RCA batch plan",
            "discovery-rca-batch-plan.json",
            "latest bootstrap progress",
            "latest bootstrap log tail",
            "endlocal & exit /b 0",
        ),
        "scripts/collect_stage6_evidence.bat": (
            "collect_stage6_evidence.py",
            "diagnose.bat",
            "--json",
            "stage6-evidence-*.zip",
            "does not copy",
            "live SQLite database",
            "downloaded PDFs",
            "endlocal & exit /b %BUNDLE_RC%",
        ),
        "scripts/collect_bug_report.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "collect_bug_report.py",
            '--root "%EIDP_APP_ROOT%"',
            "[collect_bug_report] ERROR: no Python found",
            "exit /b %ERRORLEVEL%",
        ),
        "scripts/verify_stage6_evidence.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "verify_stage6_evidence.py",
            "stage6-evidence-*.zip",
            "stage6-evidence-verify-%VERIFY_STAMP%.json",
            "--require-label last_run",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/validate_install.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "validate_windows_install.py",
            '"%EIDP_APP_ROOT%" %*',
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/stage6_recovery_check.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "stage6_recovery_check.py",
            "--expected-weekly-action",
            "--probe-weekly-dry-run",
            "--probe-lock",
            "EIDP_EXPECTED_WEEKLY_ACTION",
            "EIDP_RECOVERY_PROBE_WEEKLY_DRY_RUN",
            "EIDP_RECOVERY_PROBE_LOCK",
            "expected weekly action: skipped",
            "expected weekly action",
            "stage6-recovery-%RECOVERY_STAMP%.json",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/stage6_residual_cleanup.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            "stage6_residual_cleanup.py",
            "--apply",
            "dry-run unless --apply",
            "stage6_residual_cleanup",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/repair_streamlit_launcher.bat": (
            ".venv\\Scripts\\python.exe",
            "runtime\\python\\python.exe",
            'set "PYTHONPATH=%EIDP_APP_ROOT%\\src;%PYTHONPATH%"',
            "repair_streamlit_launcher.py",
            "dry-run unless --apply",
            'set "RC=%ERRORLEVEL%"',
            "endlocal & exit /b %RC%",
        ),
        "scripts/uninstall.bat": (
            "schtasks /Delete",
            "EIDP Weekly Run",
            "data\\",
        ),
    }
    forbidden_tokens: dict[str, tuple[str, ...]] = {
        "scripts/first_setup.bat": ("import-master",),
        "scripts/launch.bat": ("streamlit.main",),
        "scripts/weekly_run.bat": ("%date:~",),
        "scripts/uninstall.bat": ("rmdir", "del ", "erase ", "rd "),
    }

    for member, tokens in required_tokens.items():
        if member not in names:
            continue
        body = _read_zip_text(check, member)
        if body is None:
            continue
        _reject_bare_rc_assignment(check, body, member)
        if member.startswith("scripts/") and member != "scripts/uninstall.bat":
            _check_bat_common(check, member, body)
        for token in tokens:
            _require_text(check, body, member, token)
        lowered = body.lower()
        for token in forbidden_tokens.get(member, ()):
            _reject_text(check, lowered, member, token)


def _check_python_entrypoint_contracts(check: ZipCheck, names: set[str]) -> None:
    """Validate packaged Python helper scripts that the VM gate relies on."""
    required_tokens: dict[str, tuple[str, ...]] = {
        "scripts/offline_pip_install.py": (
            "platform._wmi_query",
            "raise OSError",
            "PIP_DISABLE_PIP_VERSION_CHECK",
            "PIP_NO_INPUT",
            "PIP_NO_CACHE_DIR",
            "pip._internal.cli.main",
        ),
        "src/eidp/windows_platform.py": (
            "disable_wmi_platform_queries",
            "platform._wmi_query",
            "raise OSError",
            "_eidp_wmi_disabled",
        ),
        "src/sitecustomize.py": (
            "disable_wmi_platform_queries",
        ),
        "scripts/validate_windows_install.py": (
            "CORE_FILES",
            "build_commit",
            "scripts/validate_install.bat",
            "--after-setup",
            "--after-bootstrap",
            "--after-weekly",
            "--require-ocr-addon",
            "--require-ocr-runtime",
            "--require-playwright-addon",
            "last_run.json status must be success",
            "support_recipient",
            "sqlite_integrity_check",
            "PRAGMA integrity_check",
            "department_change missing column",
            "uq_document_file_hash",
            "TARGET_FY_SCHOOL_TYPE",
            "sqlite_target_fy_yield_pct",
            "sqlite_target_fy_operator_reviewable_yield_pct",
            "bootstrap ship_gate_status pass does not match SQLite operator-reviewable coverage",
            "weekly summary after.coverage does not match SQLite target-FY coverage",
        ),
        "scripts/collect_stage6_evidence.py": (
            "build_evidence_bundle",
            "stage6-evidence-manifest.json",
            "BASE_EVIDENCE_PATTERNS",
            "EXCEL_EVIDENCE_PATTERN",
            "include_excel",
            "EXCLUDED_DATA_PARTS",
            "eidp.sqlite3",
            "pdfs",
            "last_run.json",
            "stage6-recovery-*.json",
            "discovery-rca-batch-plan.json",
        ),
        "scripts/collect_bug_report.py": (
            "build_bug_report_bundle",
            "scrub_json_value",
            "local bundle generation",
            "--note",
            "--json",
            "never uploads data",
        ),
        "src/eidp/bug_signals/detector.py": (
            "BugSignal",
            "scan_bug_signals",
            "scan_p0_bug_signals",
            "weekly_run_error",
            "weekly_run_timeout_no_last_run",
            "stale_lock",
            "weekly_run_log_error",
            "PRAGMA integrity_check",
        ),
        "src/eidp/bug_signals/bundle.py": (
            "build_bug_report_bundle",
            "scrub_text",
            "scrub_json_value",
            "bug-signals.json",
            "manifest.json",
            "local-only bundle",
            "data/eidp.sqlite3 and sidecars are never bundled",
            "data/pdfs is never bundled",
            "data/output/*.xlsx is never bundled",
        ),
        "src/eidp/review/_pages/bug_report.py": (
            "不具合レポート",
            "scan_bug_signals",
            "build_bug_report_bundle",
            "scrub_json_value",
            "ローカルレポートZIPを作成",
            "application/zip",
        ),
        "scripts/verify_stage6_evidence.py": (
            "verify_stage6_evidence_bundle",
            "stage6-evidence-manifest.json",
            "DEFAULT_REQUIRED_LABELS",
            "FORBIDDEN_EXACT_ENTRIES",
            "FORBIDDEN_PREFIXES",
            "EXCEL_EXPORT_PREFIX",
            "allow_excel",
            "eidp.sqlite3",
            "data/pdfs/",
            "runtime/",
            "wheelhouse/",
            "missing_required_labels",
        ),
        "scripts/verify_stage6_return.py": (
            "verify_stage6_return",
            "target_pdf_auto_yield_pct",
            "target_pdf_excel_ready_yield_pct",
            "operator_reviewable_yield_pct",
            "ship_gate_status",
            "KPI actual must match",
            "last_run finished_at must be ISO datetime",
            "SHIP_GATE_EXCEPTION_REASONS",
            "release_exception_reason",
            "mature_year_proof_json",
            "release_exception_record",
            "release-exception actual is blank",
            "release-exception row missing or malformed",
            "Stage 2-5c Windows VM gate 済み",
            "Runbook 修正反映済み",
            "release exception requires --mature-year-proof-json",
            "release exception requires --release-exception-record",
            "mature-year proof JSON ok must be true",
            "mature-year proof JSON must reference verifier proof JSON file",
            "mature-year proof years must match passing proof JSON years",
            "release exception cannot be combined with allow-unmeasured KPI mode",
            "RELEASE_CONCLUSIONS",
            "READY",
            "RC_ONLY",
            "NOT_READY",
            "release conclusion must be READY for release approval",
            "Decision must be READY for release approval",
            "must be YYYY-MM-DD",
            "Date must not be in the future",
            "Date must be on or after last_run finished_at date",
            "Date must be on or after release exception Approval date",
            "release exception record Date is required",
            "release exception record Date must be YYYY-MM-DD",
            "release exception record Date must not be in the future",
            "release exception record Date must match Approval date",
            "Approver must not be a placeholder",
            "Name must not be a placeholder",
            "Approval date must be on or after mature-year proof finished_at date",
            "Release scope",
            "Release scope must limit approval to v1.0 mature-year proof only",
            "FY2026/R8 status acknowledged",
            "release exception record FY2026/R8 status acknowledged must be yes",
            "Required follow-up",
            "Required follow-up must require FY2026/R8 strict-yield rerun",
            "Owner sign-off:",
            "業務員 sign-off:",
            "Excel ready 率",
            "Excel 整合性",
            "出力ファイル:",
            "Excel output file proof must include a generated data/output/*.xlsx workbook path",
            "manual_action_log 件数",
            "JSONL outbox 未送信件数",
            "JSONL action_id 重複",
            "min_target_pdf_auto_yield",
            "min_target_pdf_auto_denominator_count",
            "target_pdf_auto_denominator_count",
            "target_pdf_auto_denominator_scope",
            "max_manual_workload",
        ),
        "scripts/build_mature_year_acquisition_proof.py": (
            "build_mature_year_acquisition_proof",
            "MATURE_YEAR_SHIP_GATE_METRIC_BASIS",
            "MATURE_YEAR_PROOF_MIN_DENOMINATOR",
            "target_pdf_auto_denominator_count",
            "target_pdf_auto_denominator_scope",
            "target_pdf_auto_yield_pct",
            "operator_reviewable_yield_pct",
            "ship_gate_status",
            "estimated_manual_workload_pct",
            "--case",
            "FY=path",
        ),
        "scripts/evaluate_strict_yield_bound.py": (
            "evaluate_bound",
            "strict_yield_upper_bound",
            "no_go_upper_bound_below_required",
            "still_possible_below_gate",
            "max_possible_strict_yield_pct_if_all_remaining_pass",
            "--strict-gap-json",
            "--strict-successes",
        ),
        "scripts/repair_streamlit_launcher.py": (
            "repair_launcher",
            "streamlit.main",
            "-m streamlit run",
            "--apply",
            ".bak",
            "dry-run",
            "resolve(strict=True)",
            "_unique_backup_path",
            "backup already exists",
            "acquire_lock",
            "post-repair validation failed",
        ),
        "scripts/stage6_recovery_check.py": (
            "EIDP Weekly Run",
            "DEFAULT_INTERRUPTED_STAGE6_PATHS",
            "--expected-weekly-action",
            "--probe-weekly-dry-run",
            "--probe-lock",
            "Scheduled task action check skipped",
            "weekly_dry_run_probe",
            "lock_probe",
            "residual_paths",
            "action_matches_expected",
        ),
        "scripts/stage6_residual_cleanup.py": (
            "DEFAULT_INTERRUPTED_STAGE6_PATHS",
            "--apply",
            "dry_run",
            "stage6-residual-cleanup",
            "refusing to move symlink or junction",
            "refusing copy/delete fallback",
            "_archive_by_rename",
            "refusing to move path outside USERPROFILE",
        ),
        "scripts/run_weekly_target_year_discovery.py": (
            "acquire_lock",
            "last_run.json",
            "write_last_run",
            "write_progress",
            "--progress-file",
            "prune_run_logs",
            "prune_run_artifacts",
            "RUN_ARTIFACT_PATTERNS",
            "run_pdf_discovery",
            "run_ingestion",
            "write_text_atomic",
            "operator_reviewable_yield_pct",
            "ship_gate_operator_coverage_pct",
        ),
        "scripts/bootstrap_pdf_pipeline.py": (
            "Step 2b",
            "known URL / corporation fallback discovery",
            "discovered-urls-50.csv",
            "prefecture_aggregator,seed_csv,corporation_pattern,school_domain_override,scrapling_stealth",
            "--target-fiscal-year",
            "_effective_target_fiscal_year",
            "target_fiscal_year=effective_target_fiscal_year",
            "target_fiscal_year=target_fiscal_year",
            "progress_callback",
            "write_text_atomic",
            "operator_reviewable_yield_pct",
            "ship_gate_operator_coverage_pct",
        ),
        "scripts/atomic_write.py": (
            "write_text_atomic",
            "tmp_path.replace",
        ),
        "scripts/ship_gate_contract.py": (
            "BOOTSTRAP_SHIP_GATE_METRIC_BASIS",
            "WEEKLY_SHIP_GATE_METRIC_BASIS",
            "SHIP_GATE_AUTO_YIELD_PCT",
            "SHIP_GATE_OPERATOR_COVERAGE_PCT",
            "MATURE_YEAR_SHIP_GATE_METRIC_BASIS",
            "MATURE_YEAR_PROOF_MIN_DENOMINATOR",
            "WEEKLY_SHIP_GATE_DENOMINATOR_SCOPE",
            "SHIP_GATE_EXCEPTION_REASONS",
            "post_bootstrap_operator_reviewable_coverage",
            "weekly_strict_target_pdf_and_operator_reviewable_acquisition",
            "mature_year_retroactive_strict_target_pdf_and_operator_reviewable_acquisition",
            "target_missing_schools_before_run",
            "publication_lag",
            "ship_gate_status_from_operator_coverage",
            "ship_gate_status_from_yield",
            "is_ship_gate_exception_reason",
        ),
        "requirements-windows.txt": (
            "jsonschema>=4.0,<5.0",
        ),
        "src/eidp/cli.py": (
            "_echo_import_excel_results",
            "invalid_year",
            "想定外の年度",
            "--target-fiscal-year",
            "MIN_SUPPORTED_TARGET_FISCAL_YEAR",
            "MAX_SUPPORTED_TARGET_FISCAL_YEAR",
            "target_fiscal_year=target_fiscal_year",
            '_require_app_lock("cli_import_excel")',
            '_require_app_lock("cli_match_mext")',
            '_require_app_lock("cli_reconcile")',
            '_require_app_lock("cli_discover_urls")',
            '_require_app_lock("cli_crawl_school_urls")',
            '_require_app_lock("cli_discover_pdfs")',
            '_require_app_lock("cli_ingest_pdfs")',
            '_require_app_lock("cli_db_bootstrap")',
            '_require_app_lock("cli_rebuild_school_year_tasks")',
            '_require_app_lock("cli_prefecture_aggregate")',
            '_require_app_lock("cli_audit_flush")',
            '_require_app_lock("cli_populate_reviews")',
            '_require_app_lock("cli_weekly_update")',
            '_require_app_lock("cli_firecrawl_discover")',
            '_require_app_lock("cli_seed_discovery_gold_sites")',
        ),
        "src/eidp/cli_discovery.py": (
            "register_discovery_commands",
            'app.command("eval-discovery-gold")',
            'app.command("discovery-gold-expected-predictions")',
            "--fail-on-regression",
            "--fail-on-undemonstrated-pattern-sources",
            "_discovery_gold_gate_failed",
            "_reject_relative_path_traversal",
            "discovery-rca-outcome-validate",
        ),
        "src/eidp/cli_reports.py": (
            "report_app = typer.Typer",
            '@report_app.command("ship-readiness")',
            "--fail-on-missing-goal",
            "_exit_report_db_error",
            "database_not_ready",
            "report query failed; database is not initialized or the schema is incomplete",
            "ok_operator_review",
            "ok_strict",
            "configured_target_fiscal_year",
            "calendar_current_fiscal_year",
            "is_retroactive_fiscal_year",
        ),
        "src/eidp/cli_tools.py": (
            "register_tool_commands",
            "_echo_excel_export_results",
            "_exit_database_not_ready_error",
            "export_master_workbook",
            "export_competition_workbook",
            "diff_workbooks",
            "run_full_evaluation",
        ),
        "src/eidp/reports/ship_readiness.py": (
            "compute_ship_readiness",
            "strict_target_pdf_rate",
            'name="strict_target_pdf"',
            "estimated_manual_workload",
            "excel_ready",
            "strict_auto_target_pdf_min",
            "manual_workload_max",
        ),
        "src/eidp/config.py": (
            "target_fiscal_year",
            "MIN_SUPPORTED_TARGET_FISCAL_YEAR = 2019",
            "MAX_SUPPORTED_TARGET_FISCAL_YEAR = 2099",
            "SUPPORTED_TARGET_FISCAL_YEAR_RANGE_LABEL",
            "pdf_discovery_experimental_extractors: bool = False",
            "def _validate_target_fiscal_year",
            "target_fiscal_year outside supported range",
        ),
        "src/eidp/review/_pages/settings_page.py": (
            "MIN_SUPPORTED_TARGET_FISCAL_YEAR",
            "MAX_SUPPORTED_TARGET_FISCAL_YEAR",
        ),
        "src/eidp/fiscal_year.py": (
            "def has_fiscal_year_text",
            "active_japanese_eras",
            "era.initial",
            "era.romanized",
            "20\\d{2}",
        ),
        "src/eidp/review/school_scope.py": (
            'OPERATOR_SCHOOL_TYPE_SCOPE: str | None = "専門学校"',
            'OPERATOR_SCHOOL_SCOPE_LABEL = "専門学校"',
        ),
        "src/eidp/review/app.py": (
            "log_manual_action",
            'action_type="school_code_approved"',
            'action_type="school_code_corrected"',
            'action_type="school_code_rejected"',
            'action_type="school_code_skipped"',
            "PAGE_BUG_REPORT",
            "_render_bug_signal_banner",
            "異常を",
            "レポート作成",
            "render_bug_report",
        ),
        "src/eidp/review/_pages/url_candidate_review.py": (
            "log_manual_action",
            'action_type="url_candidate_approved"',
            'action_type="url_candidate_rejected"',
            'target_table="school_site"',
            'target_table="review_item"',
        ),
        "src/eidp/review/_pages/school_year_tasks.py": (
            "target_pdf_auto_yield_pct",
            "operator_reviewable_yield_pct",
            "レビュー目安",
            "レビュー判定",
            "画像PDF/OCR待ちが",
            "OCR add-on 未導入",
        ),
        "src/eidp/review/operator_pages.py": (
            "log_manual_action",
            "apply_dept_alias_proposal",
            'action_type="dept_alias_approved"',
            "void_department_change",
            'action_type="dept_change_void"',
            "Excelを閉じてから再実行してください",
        ),
        "src/eidp/excel/exporter.py": (
            "thresholds_from_env",
            "def excel_confidence_thresholds()",
            "def excel_min_extraction_confidence()",
            "def excel_auto_flag_extraction_confidence()",
            'LOW_CONFIDENCE_EXCLUSION_SHEET = "出力除外_低信頼"',
            "_exportable_confidence_sql",
            "excel_min_extraction_confidence()",
            "export_quality_warnings",
            "_low_confidence_reason",
        ),
        "src/eidp/excel/competition_exporter.py": (
            "TargetFiscalYearDataMissingError",
            "business_target_export = fiscal_year is None",
            "settings.target_fiscal_year",
            'gap_report_for_export(session, fiscal_year=fiscal_year, school_type="専門学校")',
            "export_gap.has_target_year_data",
            "run target-year acquisition/review before business export",
            '"target_yearly_rows": export_gap.target_yearly_rows',
            '"excel_ready_schools": export_gap.excel_ready_schools',
        ),
        "src/eidp/reports/coverage.py": (
            "from eidp.config import settings",
            "settings.target_fiscal_year",
            "schools_with_target_pdf_current_fy",
            "target_pdf_current_fy_rate",
            "d_fy == fy",
            "Document.fiscal_year == fiscal_year",
            "Document.fiscal_year < fiscal_year",
            "stale_fallback_schools",
            "target_pdf = int(coverage.schools_with_target_pdf_current_fy)",
        ),
        "src/eidp/reports/gaps.py": (
            "from eidp.config import settings",
            "settings.target_fiscal_year",
        ),
        "src/eidp/db/audit.py": (
            "ManualActionLog",
            "def log_manual_action",
            "action_id=str(uuid.uuid4())",
            "session.flush()",
        ),
        "src/eidp/db/audit_outbox.py": (
            'DEFAULT_OUTBOX_PATH = Path("data/audit/manual-actions.jsonl")',
            'OUTBOX_ARCHIVE_GLOB = "manual-actions-*.jsonl"',
            "def flush_audit_outbox",
            "def _candidate_outbox_paths",
            "def _is_matching_outbox_archive",
            "ManualActionLog",
            "jsonl_exported_at",
            "jsonl_export_error",
            "os.fsync",
        ),
        "src/eidp/db/sqlite_bootstrap.py": (
            "_refuse_orphaned_sqlite_sidecars",
            "main SQLite database file is missing but SQLite sidecar files exist",
            "verify_sqlite_integrity",
            "PRAGMA integrity_check",
            "ensure_sqlite_additive_columns",
            "_DEPARTMENT_CHANGE_VOID_COLUMNS",
            "void_reason",
            "ensure_sqlite_document_file_hash_index",
            "uq_document_file_hash",
        ),
        "src/eidp/scraper/pdf_discovery.py": (
            "strict_target_fiscal_year",
            "MIN_SUPPORTED_FISCAL_YEAR = MIN_SUPPORTED_TARGET_FISCAL_YEAR",
            "MAX_SUPPORTED_FISCAL_YEAR = MAX_SUPPORTED_TARGET_FISCAL_YEAR",
            "MIN_SUPPORTED_FISCAL_YEAR <= fiscal_year <= MAX_SUPPORTED_FISCAL_YEAR",
            "max(MIN_SUPPORTED_FISCAL_YEAR, target_year - 8)",
            "min(MAX_SUPPORTED_FISCAL_YEAR + 1, target_year + 3)",
            "target_fiscal_year_not_detected",
            "fiscal_year_mismatch:",
            "target_application_not_detected",
            "prefecture_index_current_year",
            'trusted_year_evidence if strict_target_fiscal_year else ""',
            '"target_fiscal_year" not in evidence.extra',
            "replace(evidence",
            "candidate.detected_fiscal_year >= target_year",
            "has_fiscal_year_text",
            "_candidate_download_year_rank",
            "PDF_LINK_ATTRIBUTE_NAMES",
            "PDF_DATA_ATTRIBUTE_TAG_PATTERN",
            "PDF_SCRIPT_URL_PATTERN",
            "PDF_META_REFRESH_PATTERN",
            "PDF_OPTION_VALUE_PATTERN",
            "PDF_FORM_ACTION_PATTERN",
            "PDF_INPUT_TAG_PATTERN",
            "PDF_EMBED_TAG_NAMES",
            "PDF_EMBED_ATTRIBUTE_NAMES",
            "experimental_extractors: bool | None = None",
            "settings.pdf_discovery_experimental_extractors",
            "if experimental_enabled:",
            "_pdf_delivery_pattern",
            'source="meta_refresh"',
            'source="select_option"',
            'source="form_action"',
            'source="data_attribute"',
            'source="onclick"',
            'source="input_control"',
            'endswith("_direct")',
            "_pdf_url_from_meta_refresh_content",
            '"http-equiv"',
            "_pdf_element_context_text",
            'pattern_type="embed"',
            '"value"',
            '"action"',
            "_pdf_urls_from_script_attribute",
            '"onclick"',
            "button|span|div",
            '"data-href"',
            '"data-url"',
            '"data-file"',
            '"data-pdf"',
            '"data-src"',
            "_candidate_dedupe_preference",
            "_candidate_dedupe_year_preference",
            "candidate_year == target_year",
            "target_fiscal_year=target_year",
            "(?:[?#][^\"\\']*)?",
            "_without_url_fragment",
            "_append_or_upgrade_candidate",
            "candidate_budget_dropped",
            "max_general_candidate_scan=",
            "SchoolSite.school_id.asc()",
            "SchoolSite.id.asc()",
        ),
        "src/eidp/scraper/url_normalization.py": (
            "TRACKING_QUERY_PARAMS",
            '"utm_source"',
            '"utm_medium"',
            '"gclid"',
            "key.lower() not in TRACKING_QUERY_PARAMS",
            'key.lower() == "wpdmdl"',
        ),
        "src/eidp/scraper/url_discovery.py": (
            "School.prefecture.asc()",
            "School.id.asc()",
            "SchoolSite.school_id.asc()",
            "SchoolSite.id.asc()",
        ),
        "src/eidp/scraper/discovery_gold_set.py": (
            "Draft202012Validator",
            "_load_and_validate_discovery_gold_payload",
            "discovery gold-set schema validation failed",
            "entries_by_key",
            "(entry.school_id, entry.target_fiscal_year)",
            "_target_fiscal_year_from_evidence_payload",
            "school_id_counts",
            'extra.get("detected_fiscal_year")) or _int_or_none(extra.get("target_fiscal_year")',
            "DISCOVERY_GOLD_NO_TARGET_EVIDENCE_REASONS",
            '"pre_filtered_non_target_hint"',
            '"classified_non_target"',
            '"all_negative_score"',
            '"needs_operator_review": 3',
            '"site_fetch_error": 2',
            "_is_better_tie_break_prediction",
            'candidate.outcome == current.outcome == "publication_lag_latest_public"',
            "(candidate.fiscal_year or 0) > (current.fiscal_year or 0)",
            "pattern_type_mismatch",
            'pattern_type=str(payload.get("pattern_type") or "")',
            "json.JSONDecodeError",
            "line_number",
        ),
        "src/eidp/scraper/discovery_evidence_summary.py": (
            "top_reasons=_sorted_counter_items",
            "def _sorted_counter_items",
            "key=lambda item: (-item[1], item[0])",
            "json.JSONDecodeError",
            "line_number",
        ),
        "src/eidp/pdf/extractor.py": (
            "def _extract_fiscal_year",
            "from eidp.config import",
            "MIN_SUPPORTED_TARGET_FISCAL_YEAR",
            "MAX_SUPPORTED_TARGET_FISCAL_YEAR",
            "MIN_SUPPORTED_FISCAL_YEAR = MIN_SUPPORTED_TARGET_FISCAL_YEAR",
            "MAX_SUPPORTED_FISCAL_YEAR = MAX_SUPPORTED_TARGET_FISCAL_YEAR",
            "fiscal_year < MIN_SUPPORTED_FISCAL_YEAR",
            "fiscal_year > MAX_SUPPORTED_FISCAL_YEAR",
            "fiscal_year_from_japanese_era_text",
            "settings.target_fiscal_year if max_fiscal_year is None",
            'filing_dates = re.findall(r"(20\\d{2})[./]\\d{1,2}[./]\\d{1,2}"',
            'all_years = re.findall(r"(20\\d{2})[\\.\\s年/]"',
            "MIN_SUPPORTED_FISCAL_YEAR <= int(y) <= max_valid_year",
            "max_valid_year",
            "format_fiscal_year_as_japanese_era(fiscal_year)",
        ),
        "src/eidp/pipeline/ingest.py": (
            "DepartmentYearly",
            "SupportRecipient",
            "from eidp.config import settings",
            "target_fiscal_year: int | None = None",
            "fiscal_year_cap = target_fiscal_year if target_fiscal_year is not None else settings.target_fiscal_year",
            "compute_pdf_parse_breakdown",
            "breakdown_to_json",
            "verdict = classify(breakdown.composite, thresholds_from_env())",
            'is_current_row = verdict in ("auto", "auto_flag")',
            "revision=next_revision",
            "if is_current_row:\n                dy = DepartmentYearly(",
            "is_current=True",
            'stats["yearly_review_pending"] += 1',
            'sr_is_current = sr_verdict in ("auto", "auto_flag")',
            "if sr_is_current:\n            sr = SupportRecipient(",
            "support_recipient_review_pending",
            'stats["support_recipient_review_pending"] = 1',
            "if yearly_review > 0 or sr_review > 0:",
            'doc.ingest_status = "review_pending"',
            "doc.is_current_year = fiscal_year >= fiscal_year_cap",
            "target_fiscal_year=target_fiscal_year",
            "max_fiscal_year=fiscal_year_cap",
            "has_fiscal_year_text",
        ),
        "src/eidp/pipeline/manual_entry.py": (
            "log_manual_action",
            'action_type="manual_entry"',
            'target_table="department_yearly"',
            'target_table="department_change"',
            'target_table="document"',
        ),
        "src/eidp/pipeline/fiscal_year_override.py": (
            "log_manual_action",
            'action_type="fiscal_year_override"',
            'target_table="department_yearly"',
            'target_table="support_recipient"',
            'target_table="school_year_status"',
            'target_table="document"',
        ),
        "src/eidp/ocr/tesseract.py": (
            "locate_tesseract",
            "locate_tessdata",
            "run_tesseract_on_image",
            "parse_tesseract_tsv",
            "EIDP_TESSERACT_BIN",
            "tesseract.exe",
            "jpn.traineddata",
            "--tessdata-dir",
        ),
        "src/eidp/ocr/availability.py": (
            "detect_ocr_availability",
            "availability_banner_text",
            "availability_banner_severity",
            "has_jpn_traineddata",
            "auto_can_run",
            "OCR add-on 未インストール",
        ),
    }
    current_fy_literal_tokens = (
        "2026",
        "令和8",
        "令和８",
        "R8",
    )
    forbidden_tokens: dict[str, tuple[str, ...]] = {
        "scripts/bootstrap_pdf_pipeline.py": current_fy_literal_tokens,
        "scripts/run_weekly_target_year_discovery.py": ("export_excel", *current_fy_literal_tokens),
        "scripts/validate_windows_install.py": current_fy_literal_tokens,
        "src/eidp/pdf/extractor.py": current_fy_literal_tokens,
        "src/eidp/reports/coverage.py": current_fy_literal_tokens,
        "src/eidp/reports/gaps.py": current_fy_literal_tokens,
        "src/eidp/reports/ship_readiness.py": current_fy_literal_tokens,
        "src/eidp/scraper/pdf_discovery.py": (
            *current_fy_literal_tokens,
            "r8",
            "renewalconfirmationapplication",
            "renewal-confirmation-application",
            "renewal confirmation application",
        ),
    }

    for member, tokens in required_tokens.items():
        if member not in names:
            continue
        body = _read_zip_text(check, member)
        if body is None:
            continue
        for token in tokens:
            _require_text(check, body, member, token)
        for token in forbidden_tokens.get(member, ()):
            _reject_text(check, body, member, token)


def _check_operator_runbook_contract(check: ZipCheck, names: set[str]) -> None:
    """Catch stale operator-facing instructions in the packaged runbook.

    The Windows ZIP can pass structural checks while still shipping old
    navigation guidance. That is a release blocker because the target user
    launches from the runbook, not from the source tree.
    """
    member = "docs/runbooks/eidp-windows.md"
    if member not in names:
        return
    body = _read_zip_text(check, member)
    if body is None:
        return
    _reject_local_user_path_tokens(check, body, member)
    for token in (
        "業務員クイック",
        "学校キュー",
        "実行中のパッケージ",
        "詳細 operator",
        "週次URL/PDF再取得",
        "対象年度を変更して保存すると、学校キューも同時に再計算されます",
        "scripts\\weekly_run.bat` は管理者向けの復旧入口",
        "logs\\stage6-recovery-*.json",
        "stage6_residual_cleanup.bat",
        "logs\\stage6-residual-cleanup-*.json",
        "logs\\stage6-residual-archive\\<timestamp>",
        "ClearAllForwardings=no",
        "ExitOnForwardFailure=yes",
        "127.0.0.1:18501:127.0.0.1:8501",
        "127.0.0.1:18501/_stcore/health",
        "127.0.0.1:18502:127.0.0.1:8502",
        "logs\\stage6-evidence-*.zip",
        "EIDP-stage6-evidence.bat",
        "EIDP-stage6-verify-evidence.bat",
        "EIDP-stage6-recovery.bat",
        "verify_stage6_evidence.py",
        "stage6-evidence-verify-*.json",
        "--require-label last_run",
        "scripts\\stage6_recovery_check.bat",
        "アンチウイルスにより隔離された",
        "新しい ZIP へ更新する場合",
        "data\\eidp.sqlite3-wal",
        "data\\eidp.sqlite3-shm",
        "PRAGMA wal_checkpoint(TRUNCATE)",
        "VACUUM INTO",
        "db-backup --output $dbBackup",
        "eidp-backup-$ts.sqlite3",
        'Get-ChildItem "$old\\data" -Force',
        '$_.Name -notlike "eidp-backup-*.sqlite3"',
        "data\\.lock",
        '/XF ".lock" "eidp.sqlite3-wal" "eidp.sqlite3-shm" "eidp-backup-*.sqlite3"',
    ):
        _require_text(check, body, member, token)
    for token in (
        "画面左のサイドバーに 12 ページ",
    ):
        _reject_text(check, body, member, token)

    e2e_member = "docs/runbooks/eidp-operator-e2e-template.md"
    if e2e_member not in names:
        return
    e2e_body = _read_zip_text(check, e2e_member)
    if e2e_body is None:
        return
    _reject_local_user_path_tokens(check, e2e_body, e2e_member)
    for token in (
        "ship_readiness_rc",
        "retroactive_fiscal_year",
        "is_retroactive_fiscal_year",
        "retroactive_ship_readiness_rc",
        "stage6_recovery_rc",
        "stage6 recovery check",
        "task.execute",
        "task.expected_action",
        "task.action_matches_expected",
        "residual_paths",
        "recommendations",
        "strict target PDF 自動取得率",
        "推定手作業率",
        "release exception reason",
        "mature-year proof JSON",
        "mature-year proof years",
        "Excel ready 率",
        "logs\\diagnostics-*.txt",
        "127.0.0.1:18501:127.0.0.1:8501",
        "127.0.0.1:18501/_stcore/health",
        "Mac tunnel health",
        "logs\\stage6-evidence-*.zip",
        "logs\\stage6-evidence-verify-*.json",
        "logs\\stage6-recovery-*.json",
        "logs\\stage6-residual-cleanup-*.json",
        "Get-Volume C",
        "ipconfig /flushdns",
        "data\\.lock.meta",
        "Get-ScheduledTaskInfo -TaskName",
        "Task Scheduler retry",
        "現行投入候補（Mac / non-Windows gate 済み、Windows 未実証）:",
        "dist/eidp-windows-vXXX.zip",
        "core ZIP sha256 sidecar note",
        "logs/release-gate-vXXX-retroactive.json",
        '$zip = "C:\\EIDP-staging\\<core-zip-file-name>"',
        '$expected = "<copy SHA256 from .sha256 sidecar or current-release-status>"',
        "v1.0-rc",
        "FY2026/R8 の current-year yield gate",
        "Date: YYYY-MM-DD",
        "実際に生成・確認した `.xlsx` のパスを記入する。",
    ):
        _require_text(check, e2e_body, e2e_member, token)
    _check_operator_e2e_template_version_neutral_fields(check, e2e_body, e2e_member)


def _check_build_info(check: ZipCheck, names: set[str]) -> None:
    if "BUILD_INFO.json" not in names:
        return
    if not zipfile.is_zipfile(check.path):
        return
    with zipfile.ZipFile(check.path) as zf:
        try:
            payload = json.loads(zf.read("BUILD_INFO.json").decode("utf-8"))
        except (KeyError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            check.fail(f"BUILD_INFO.json is not valid JSON: {exc}")
            return
    if not isinstance(payload, dict):
        check.fail("BUILD_INFO.json must contain an object")
        return
    for key in ("app", "built_at_utc", "git_commit", "git_branch", "git_dirty"):
        value = payload.get(key)
        if not isinstance(value, str) or not value:
            check.fail(f"BUILD_INFO.json missing string field: {key}")
    if payload.get("app") != "EIDP":
        check.fail("BUILD_INFO.json app must be EIDP")
    commit = payload.get("git_commit")
    if not isinstance(commit, str) or len(commit) != 40:
        check.fail("BUILD_INFO.json git_commit must be a full 40-character commit hash")
    if payload.get("git_dirty") != "false":
        check.fail("BUILD_INFO.json git_dirty must be false")
    check.details["build_info"] = payload


def _read_manifest(check: ZipCheck, member: str) -> dict[str, Any] | None:
    if not zipfile.is_zipfile(check.path):
        return None
    with zipfile.ZipFile(check.path) as zf:
        try:
            raw = zf.read(member)
        except KeyError:
            return None
    try:
        manifest = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        check.fail(f"manifest is not valid UTF-8 JSON: {member}: {exc}")
        return None
    if not isinstance(manifest, dict):
        check.fail(f"manifest is not a JSON object: {member}")
        return None
    return manifest


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _manifest_file_entries(check: ZipCheck, manifest: dict[str, Any], *, manifest_path: str) -> list[dict[str, Any]]:
    files = manifest.get("files")
    if not isinstance(files, list):
        check.fail(f"{manifest_path} missing files list")
        return []
    entries: list[dict[str, Any]] = []
    seen_paths: set[str] = set()
    for index, entry in enumerate(files):
        if not isinstance(entry, dict):
            check.fail(f"{manifest_path} files[{index}] is not an object")
            continue
        path = entry.get("path")
        size = entry.get("size")
        sha256 = entry.get("sha256")
        if not isinstance(path, str) or not path:
            check.fail(f"{manifest_path} files[{index}] missing path")
            continue
        if path in seen_paths:
            check.fail(f"{manifest_path} contains duplicate file path: {path}")
        seen_paths.add(path)
        if not isinstance(size, int) or size < 0:
            check.fail(f"{manifest_path} file has invalid size: {path}")
        if not isinstance(sha256, str) or len(sha256) != 64:
            check.fail(f"{manifest_path} file has invalid sha256: {path}")
        entries.append(entry)
    return entries


def _check_manifest_integrity(
    check: ZipCheck,
    names: set[str],
    *,
    manifest_path: str,
    required_paths: tuple[str, ...],
) -> None:
    """Validate add-on MANIFEST.json against actual ZIP member bytes."""
    manifest = _read_manifest(check, manifest_path)
    if manifest is None:
        return

    entries = _manifest_file_entries(check, manifest, manifest_path=manifest_path)
    by_path = {entry["path"]: entry for entry in entries if isinstance(entry.get("path"), str)}

    for required in required_paths:
        if required not in by_path:
            check.fail(f"manifest missing required file path: {required}")

    zip_payload_names = sorted(name for name in names if name != manifest_path and not name.endswith("/"))
    missing_from_manifest = [name for name in zip_payload_names if name not in by_path]
    if missing_from_manifest:
        check.fail(f"manifest missing ZIP payload entries: {missing_from_manifest[:5]}")

    manifest_only = sorted(path for path in by_path if path not in names)
    if manifest_only:
        check.fail(f"manifest references missing ZIP entries: {manifest_only[:5]}")

    with zipfile.ZipFile(check.path) as zf:
        for path, entry in sorted(by_path.items()):
            if path not in names:
                continue
            payload = zf.read(path)
            expected_size = entry.get("size")
            expected_sha = entry.get("sha256")
            if isinstance(expected_size, int) and len(payload) != expected_size:
                check.fail(
                    f"manifest size mismatch for {path}: "
                    f"expected={expected_size} actual={len(payload)}"
                )
            if isinstance(expected_sha, str) and len(expected_sha) == 64:
                actual_sha = _sha256_bytes(payload)
                if actual_sha != expected_sha:
                    check.fail(
                        f"manifest sha256 mismatch for {path}: "
                        f"expected={expected_sha} actual={actual_sha}"
                    )

    check.details["manifest_files"] = len(entries)


def verify_core_zip(path: Path) -> ZipCheck:
    check = ZipCheck(name="core", path=path)
    names = _read_zip_names(check)
    if names is None:
        return check

    _check_exact(check, names, CORE_REQUIRED_EXACT)
    _check_prefixes(check, names, CORE_REQUIRED_PREFIXES)
    _check_no_pycache(check, names)
    _check_windows_safe_paths(check, names)
    _check_no_runtime_data(check, names)
    _check_no_secret_files(check, names)
    _check_no_historical_runbooks(check, names)
    _check_no_demo_prototype_runtime(check, names)
    _check_wheelhouse(check, names, require_project=True)
    _check_bat_contracts(check, names)
    _check_python_entrypoint_contracts(check, names)
    _check_operator_runbook_contract(check, names)
    _check_prefecture_seed_contract(check, names)
    _check_mext_authority_index_contract(check, names)
    _check_discovery_gold_set_contract(check, names)
    _check_build_info(check, names)

    check.details["entry_count"] = len(names)
    check.details["has_runtime"] = (
        "runtime/python/python.exe" in names and "runtime/uv.exe" in names
    )
    return check


def verify_ocr_addon_zip(path: Path) -> ZipCheck:
    check = ZipCheck(name="ocr-addon", path=path)
    names = _read_zip_names(check)
    if names is None:
        return check

    _check_exact(check, names, OCR_REQUIRED_EXACT)
    _check_no_pycache(check, names)
    _check_windows_safe_paths(check, names)
    _check_manifest_integrity(
        check,
        names,
        manifest_path="ocr-addon/MANIFEST.json",
        required_paths=OCR_REQUIRED_EXACT[:-1],
    )

    check.details["entry_count"] = len(names)
    return check


def verify_playwright_addon_zip(path: Path) -> ZipCheck:
    check = ZipCheck(name="playwright-addon", path=path)
    names = _read_zip_names(check)
    if names is None:
        return check

    _check_exact(check, names, PLAYWRIGHT_REQUIRED_EXACT)
    _check_prefixes(check, names, PLAYWRIGHT_REQUIRED_PREFIXES)
    _check_no_pycache(check, names)
    _check_windows_safe_paths(check, names)

    if not any(name.startswith("playwright-addon/ms-playwright/") and name.endswith("chrome.exe") for name in names):
        check.fail("playwright add-on missing Chromium chrome.exe")
    if not any(
        name.startswith("playwright-addon/wheelhouse/")
        and Path(name).name.startswith("playwright-")
        and name.endswith(".whl")
        for name in names
    ):
        check.fail("playwright add-on missing playwright-*.whl")
    if not any(
        name.startswith("playwright-addon/wheelhouse/")
        and Path(name).name.startswith("scrapling-")
        and name.endswith(".whl")
        for name in names
    ):
        check.fail("playwright add-on missing scrapling-*.whl")

    _check_manifest_integrity(
        check,
        names,
        manifest_path="playwright-addon/MANIFEST.json",
        required_paths=(),
    )

    check.details["entry_count"] = len(names)
    return check


def render_text(checks: list[ZipCheck]) -> str:
    lines: list[str] = []
    for check in checks:
        state = "OK" if check.ok else "FAIL"
        lines.append(f"{state} {check.name}: {check.path}")
        for key, value in sorted(check.details.items()):
            lines.append(f"  {key}: {value}")
        for warning in check.warnings:
            lines.append(f"  warning: {warning}")
        for error in check.errors:
            lines.append(f"  error: {error}")
    return "\n".join(lines)


def checks_to_json(checks: list[ZipCheck]) -> str:
    payload = [
        {
            "name": check.name,
            "path": str(check.path),
            "ok": check.ok,
            "errors": check.errors,
            "warnings": check.warnings,
            "details": check.details,
        }
        for check in checks
    ]
    return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)


def _require_demonstrated_discovery_patterns(check: ZipCheck) -> None:
    """Fail core ZIPs whose tracked discovery extractors lack gold-set demonstrations."""

    missing = check.details.get("discovery_gold_undemonstrated_pattern_sources")
    if isinstance(missing, list) and missing:
        check.fail(f"undemonstrated discovery extractor sources: {missing}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Verify EIDP Windows distribution ZIPs.")
    parser.add_argument("core_zip", type=Path, help="Path to eidp-windows.zip")
    parser.add_argument("--ocr-addon", type=Path, default=None)
    parser.add_argument("--playwright-addon", type=Path, default=None)
    parser.add_argument("--json", action="store_true", help="Emit machine-readable JSON")
    parser.add_argument(
        "--require-demonstrated-discovery-patterns",
        action="store_true",
        help="Fail when tracked PDF discovery extractor sources lack discovery gold-set demonstrations",
    )
    args = parser.parse_args(argv)

    core_check = verify_core_zip(args.core_zip)
    if args.require_demonstrated_discovery_patterns:
        _require_demonstrated_discovery_patterns(core_check)
    checks = [core_check]
    if args.ocr_addon is not None:
        checks.append(verify_ocr_addon_zip(args.ocr_addon))
    if args.playwright_addon is not None:
        checks.append(verify_playwright_addon_zip(args.playwright_addon))

    print(checks_to_json(checks) if args.json else render_text(checks))
    return 0 if all(check.ok for check in checks) else 1


if __name__ == "__main__":
    sys.exit(main())
