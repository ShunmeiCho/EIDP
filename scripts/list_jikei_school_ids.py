"""Resolve 滋慶 group competitor schools (from the 競合校 template) to school.id values.

Output: ordered, deduped list printed one-per-line, plus
`--school-id N` formatted args ready to paste into discover-pdfs.

Usage:
    python scripts/list_jikei_school_ids.py
    python scripts/list_jikei_school_ids.py --include-corp-groups
    python scripts/list_jikei_school_ids.py --template path/to/template.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import or_

from eidp.db.models import School, SchoolAlias
from eidp.db.session import SessionLocal
from eidp.excel.competition_exporter import _norm

DEFAULT_TEMPLATE = Path("sample/20250826更新版_競合校の在校生数.xlsx")
JIKEI_SHEET = "滋慶"
JIKEI_CORP_GROUPS = (
    "滋慶学園",
    "大阪滋慶学園",
    "東京滋慶学園",
    "滋慶コミュニケーションアート",
)


def _read_template_school_names(template_path: Path) -> list[str]:
    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb[JIKEI_SHEET]
    names: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=6, max_row=300, max_col=1, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        s = str(cell).replace("\n", "").strip()
        if not s:
            continue
        key = _norm(s)
        if key in seen:
            continue
        seen.add(key)
        names.append(s)
    return names


def _lookup_school_ids(session, template_names: list[str]) -> dict[str, list[School]]:
    """For each template name, return matching School rows (best-effort)."""
    out: dict[str, list[School]] = {}
    for name in template_names:
        norm_name = _norm(name)
        # Try canonical school_name first, then SchoolAlias.alias.
        rows = session.query(School).all()
        # Substring match in normalized space — competitor names often differ
        # from the official school_name (短縮形, 地名 separation, etc.)
        matches: list[School] = [
            s for s in rows
            if norm_name in _norm(s.school_name) or _norm(s.school_name) in norm_name
        ]
        if not matches:
            alias_rows = (
                session.query(SchoolAlias)
                .filter(SchoolAlias.alias.isnot(None))
                .all()
            )
            alias_school_ids = {
                a.school_id for a in alias_rows
                if norm_name in _norm(a.alias) or _norm(a.alias) in norm_name
            }
            if alias_school_ids:
                matches = (
                    session.query(School)
                    .filter(School.id.in_(alias_school_ids))
                    .all()
                )
        out[name] = matches
    return out


def _corp_group_school_ids(session) -> list[School]:
    return (
        session.query(School)
        .filter(or_(*[School.corporation_name == c for c in JIKEI_CORP_GROUPS]))
        .order_by(School.corporation_name, School.school_name)
        .all()
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument(
        "--include-corp-groups", action="store_true",
        help="Also include all schools from 滋慶 corporation groups.",
    )
    args = parser.parse_args()

    template_names = _read_template_school_names(args.template)
    print(f"# Template 滋慶 sheet: {len(template_names)} schools")
    for n in template_names:
        print(f"#   {n}")
    print()

    session = SessionLocal()
    try:
        name_matches = _lookup_school_ids(session, template_names)
        all_ids: list[int] = []
        all_resolved: list[tuple[str, School]] = []
        for name, schools in name_matches.items():
            print(f"# Match for: {name}")
            if not schools:
                print("    (no match)")
                continue
            for s in schools:
                print(f"    id={s.id} corp={s.corporation_name} name={s.school_name} pref={s.prefecture}")
                if s.id not in all_ids:
                    all_ids.append(s.id)
                    all_resolved.append((name, s))

        if args.include_corp_groups:
            print()
            print(f"# Corp-group expansion ({JIKEI_CORP_GROUPS}):")
            extras = _corp_group_school_ids(session)
            for s in extras:
                if s.id in all_ids:
                    continue
                print(f"    id={s.id} corp={s.corporation_name} name={s.school_name}")
                all_ids.append(s.id)

        print()
        print(f"# Total unique school_ids: {len(all_ids)}")
        print()
        print("# discover-pdfs args:")
        print("uv run eidp discover-pdfs \\")
        for sid in all_ids:
            print(f"  --school-id {sid} \\")
        print("  --evidence-log output/discovery_rejections_jikei.jsonl \\")
        print("  --batch-size 200 --rate-limit 1.0")
    finally:
        session.close()


if __name__ == "__main__":
    main()
