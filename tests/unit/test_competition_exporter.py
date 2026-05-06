from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

import eidp.excel.competition_exporter as competition_exporter
from eidp.db.models import Base
from eidp.excel.competition_exporter import (
    YearColumns,
    _append_year_columns_to_block,
    _group_triplets_into_blocks,
    _norm,
    export_competition_workbook,
    parse_sheet_schema,
    parse_template,
)

SAMPLE_TEMPLATE = Path(__file__).resolve().parents[2] / "sample" / "20250826更新版_競合校の在校生数.xlsx"


def _empty_session() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return Session(engine)


def _empty_template(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "empty"
    wb.save(path)


def _build_category_sheet() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "ゲーム"
    # Row 2: year markers
    ws.cell(2, 4, value=2024)
    ws.cell(2, 6, value=2025)
    # Row 3: 在籍数 / 留学生 headers
    ws.cell(3, 4, value="在籍数")
    ws.cell(3, 5, value="留学生")
    ws.cell(3, 6, value="在籍数")
    ws.cell(3, 7, value="留学生")
    # Row 4: data
    ws.cell(4, 1, value="日本工学院（蒲田）")
    ws.cell(4, 2, value="ゲームクリエイター科")
    ws.cell(4, 3, value="2年制")
    ws.cell(4, 4, value=292)
    ws.cell(4, 5, value=50)
    return wb


def test_parse_sheet_schema_finds_single_block_for_category() -> None:
    wb = _build_category_sheet()
    schema = parse_sheet_schema(wb["ゲーム"])
    assert schema is not None
    assert len(schema.blocks) == 1
    block = schema.blocks[0]
    assert block.school_col == 1
    assert block.dept_col == 2
    assert [yc.fiscal_year for yc in block.year_cols] == [2024, 2025]
    assert len(block.data_rows) == 1
    row = block.data_rows[0]
    assert row.school_name == "日本工学院(蒲田)"
    assert row.dept_name == "ゲームクリエイター科"
    assert row.duration_label == "2年制"
    assert row.block_id == 0


def test_norm_strips_whitespace_and_normalizes_fullwidth() -> None:
    assert _norm("日本工学院 （蒲田）") == "日本工学院(蒲田)"
    assert _norm(None) == ""
    assert _norm("ｸﾞﾗﾌｨｯｸ") == "グラフィック"


def test_group_triplets_splits_on_column_gap() -> None:
    """学校単位 sheet has a multi-column gap between left and right blocks."""
    triplets = [
        YearColumns(2024, 3, 4),
        YearColumns(2025, 5, 6),
        # Gap of 12 columns -> new block
        YearColumns(2024, 19, 20),
        YearColumns(2025, 21, 22),
    ]
    groups = _group_triplets_into_blocks(triplets)
    assert len(groups) == 2
    assert [yc.zaiseki_col for yc in groups[0]] == [3, 5]
    assert [yc.zaiseki_col for yc in groups[1]] == [19, 21]


def test_group_triplets_keeps_consecutive_years_in_one_block() -> None:
    triplets = [
        YearColumns(2024, 3, 4),
        YearColumns(2025, 5, 6),
    ]
    groups = _group_triplets_into_blocks(triplets)
    assert len(groups) == 1


def test_competition_export_defaults_to_configured_target_fiscal_year(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Business exports must not silently choose the old year with more DB rows."""
    session = _empty_session()
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _empty_template(template)
    monkeypatch.setattr(competition_exporter.settings, "target_fiscal_year", 2027)
    monkeypatch.setattr(
        competition_exporter,
        "auto_select_fiscal_year",
        lambda _session: pytest.fail("business export must not auto-select fiscal year"),
    )

    try:
        result = export_competition_workbook(session, template, output)
    finally:
        session.close()

    assert result["fiscal_year"] == 2027


def test_competition_export_explicit_year_is_admin_backcompat(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _empty_session()
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _empty_template(template)
    monkeypatch.setattr(competition_exporter.settings, "target_fiscal_year", 2027)

    try:
        result = export_competition_workbook(session, template, output, fiscal_year=2025)
    finally:
        session.close()

    assert result["fiscal_year"] == 2025


@pytest.mark.skipif(not SAMPLE_TEMPLATE.exists(), reason="sample template absent")
def test_parse_template_recognises_all_16_sheets() -> None:
    schemas = parse_template(SAMPLE_TEMPLATE)
    assert len(schemas) == 16
    rollup = schemas["学校単位での比較"]
    assert rollup.is_rollup is True
    # 学校単位 has TWO side-by-side comparison blocks
    assert len(rollup.blocks) == 2
    # Both blocks must contain 2025
    for b in rollup.blocks:
        years = {yc.fiscal_year for yc in b.year_cols}
        assert 2025 in years
    # Right block school column is well past col 16 (left block ends at 16)
    assert rollup.blocks[1].school_col > 16


@pytest.mark.skipif(not SAMPLE_TEMPLATE.exists(), reason="sample template absent")
def test_rollup_sheet_extracts_both_block_data_rows() -> None:
    """Regression: parser previously dropped right-block schools entirely."""
    schemas = parse_template(SAMPLE_TEMPLATE)
    rollup = schemas["学校単位での比較"]
    left_rows = rollup.blocks[0].data_rows
    right_rows = rollup.blocks[1].data_rows
    assert len(left_rows) > 0
    assert len(right_rows) > 0
    # Verify a known left-block school
    left_schools = {r.school_name for r in left_rows}
    assert "HAL東京" in left_schools
    # Verify right-block schools are different (independent comparison set)
    right_schools = {r.school_name for r in right_rows}
    assert right_schools - left_schools, (
        "right block should have schools not present in left block"
    )


def test_append_year_to_left_rollup_block_preserves_right_block_identity() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "学校単位での比較"
    # Left block years: C:F. Right block years: I:L. H is the right school-name col.
    ws.cell(3, 3, value=2024)
    ws.cell(3, 5, value=2025)
    ws.cell(3, 9, value=2024)
    ws.cell(3, 11, value=2025)
    for col in (3, 5, 9, 11):
        ws.cell(4, col, value="在籍数")
        ws.cell(4, col + 1, value="留学生")
    ws.cell(6, 2, value="左学校")
    ws.cell(6, 8, value="右学校")

    schema = parse_sheet_schema(ws)
    assert schema is not None
    assert len(schema.blocks) == 2

    new_cols = _append_year_columns_to_block(
        ws, schema, schema.blocks[0], 2026, schema.blocks[1:]
    )

    assert new_cols.zaiseki_col == 7
    assert new_cols.intl_col == 8
    assert schema.blocks[1].school_col == 10
    assert schema.blocks[1].year_cols[0].zaiseki_col == 11
    assert ws.cell(6, schema.blocks[1].school_col).value == "右学校"


@pytest.mark.skipif(not SAMPLE_TEMPLATE.exists(), reason="sample template absent")
def test_parse_template_extracts_data_rows_for_every_sheet() -> None:
    schemas = parse_template(SAMPLE_TEMPLATE)
    for name, schema in schemas.items():
        total_rows = sum(len(b.data_rows) for b in schema.blocks)
        assert total_rows > 0, f"{name} has no data rows across any block"
