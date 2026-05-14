from __future__ import annotations

import importlib

import openpyxl
import pytest
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

from eidp.db.models import Department, DepartmentYearly, School, SupportRecipient
from eidp.db.sqlite_bootstrap import bootstrap_sqlite
from eidp.excel.exporter import (
    _write_sairoku,
    diff_workbook_business_values,
    diff_workbook_values,
    export_master_workbook,
)


def _write_cells(path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = openpyxl.Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=col_index).value = value
    workbook.save(path)
    workbook.close()


def test_diff_workbook_values_reports_cell_and_sheet_differences(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cells(
        exported,
        {
            "Common": [["学校", "定員"], ["A専門学校", 101]],
            "ExportedOnly": [["extra"]],
        },
    )
    _write_cells(
        original,
        {
            "Common": [["学校", "定員"], ["A専門学校", 100]],
            "OriginalOnly": [["missing"]],
        },
    )

    result = diff_workbook_values(exported, original, max_diffs=5)

    assert result["ok"] is False
    assert result["missing_sheets"] == ["OriginalOnly"]
    assert result["extra_sheets"] == ["ExportedOnly"]
    assert result["differing_cells"] == 1
    assert result["samples"] == [
        {
            "sheet": "Common",
            "cell": "B2",
            "exported": 101,
            "original": 100,
        }
    ]


def test_diff_workbook_values_honors_numeric_tolerance(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cells(exported, {"Common": [["value"], [100.0001]]})
    _write_cells(original, {"Common": [["value"], [100.0]]})

    result = diff_workbook_values(exported, original, numeric_tolerance=0.001)

    assert result["ok"] is True
    assert result["differing_cells"] == 0


def test_diff_workbook_values_handles_sparse_dimensions(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cells(exported, {"Common": [["value"], [1]]})

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Common"
    worksheet.cell(row=1, column=1).value = "value"
    worksheet.cell(row=500, column=1).value = "tail"
    workbook.save(original)
    workbook.close()

    result = diff_workbook_values(exported, original, max_diffs=2)

    assert result["ok"] is False
    assert result["differing_cells"] == 2
    assert result["samples"] == [
        {"sheet": "Common", "cell": "A2", "exported": 1, "original": None},
        {"sheet": "Common", "cell": "A500", "exported": None, "original": "tail"},
    ]


def test_diff_workbook_business_values_aligns_zaiseki_year_fields_by_key(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    key = ["東京都", "片柳学園", "日本工学院専門学校", "工業", "ITスペシャリスト科", "昼", 4]
    _write_cells(
        original,
        {
            "在籍のみ抜粋": [
                [None, None, None, None, None, None, None, "在籍者数", None, "留学生数", None],
                [
                    "都道府県",
                    "法人名",
                    "学校名",
                    "課程名",
                    "学科名",
                    "昼夜",
                    "年限",
                    "2019年度",
                    "2020年度",
                    "2019年度",
                    "2020年度",
                ],
                [*key, 355, 407, 1, 4],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "在籍のみ抜粋": [
                [None, None, None, None, None, None, None, "在籍者数", None, None, "留学生数", None, None],
                [
                    "都道府県",
                    "法人名",
                    "学校名",
                    "課程名",
                    "学科名",
                    "昼夜",
                    "年限",
                    "2019年度",
                    "2020年度",
                    "2021年度",
                    "2019年度",
                    "2020年度",
                    "2021年度",
                ],
                [*key, 355, 407, 428, 1, 4, 2],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["在籍のみ抜粋"])

    assert result["differing_fields"] == 0
    assert result["extra_fields"] == {"在籍のみ抜粋": ["在籍者数:2021年度", "留学生数:2021年度"]}
    assert result["samples"] == []


def test_diff_workbook_business_values_reports_taisho_metric_diff(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    header = [
        "番号",
        "年度",
        "学校番号",
        "都道府県",
        "法人名",
        "学校名",
        "前年在籍",
        "前半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "後半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "年間",
        "家計急変多子世帯",
        "総計",
        "備考",
        "受給比率",
    ]
    original_row = [
        None,
        "2025年度",
        None,
        "東京都",
        "片柳学園",
        "日本工学院専門学校",
        6319,
        0,
        None,
        None,
        None,
        None,
        0,
        None,
        None,
        None,
        None,
        100,
        0,
        100,
        None,
        0.0158,
    ]
    exported_row = [
        1,
        "2025年度",
        None,
        "東京都",
        "片柳学園",
        "日本工学院専門学校",
        6319,
        0,
        None,
        None,
        None,
        None,
        0,
        None,
        None,
        None,
        None,
        101,
        0,
        101,
        None,
        0.0160,
    ]
    _write_cells(
        original,
        {"対象比率": [header, original_row]},
    )
    _write_cells(
        exported,
        {"対象比率": [header, exported_row]},
    )

    result = diff_workbook_business_values(exported, original, sheets=["対象比率"], max_diffs=2)

    assert result["missing_rows"] == 0
    assert result["extra_rows"] == 0
    assert result["differing_fields"] == 3
    assert result["sheet_summaries"]["対象比率"]["category_counts"] == {"numeric_mismatch": 3}
    assert result["sheet_summaries"]["対象比率"]["field_counts"] == {
        "受給比率": 1,
        "年間": 1,
        "総計": 1,
    }
    assert result["samples"] == [
        {
            "sheet": "対象比率",
            "key": "2025年度 | 東京都 | 片柳学園 | 日本工学院専門学校",
            "field": "年間",
            "exported": 101,
            "original": 100,
        },
        {
            "sheet": "対象比率",
            "key": "2025年度 | 東京都 | 片柳学園 | 日本工学院専門学校",
            "field": "総計",
            "exported": 101,
            "original": 100,
        },
    ]


def test_diff_workbook_business_values_categorizes_blank_vs_error_or_unknown(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    header = [
        "番号",
        "年度",
        "学校番号",
        "都道府県",
        "法人名",
        "学校名",
        "前年在籍",
        "前半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "後半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "年間",
        "家計急変多子世帯",
        "総計",
        "備考",
        "受給比率",
    ]
    _write_cells(
        original,
        {
            "対象比率": [
                header,
                [None, "2025年度", None, "東京都", "片柳学園", "日本工学院専門学校", "不明", *([None] * 14), "#DIV/0!"],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "対象比率": [
                header,
                [1, "2025年度", None, "東京都", "片柳学園", "日本工学院専門学校", None, *([None] * 14), None],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["対象比率"])

    assert result["differing_fields"] == 2
    assert result["sheet_summaries"]["対象比率"]["category_counts"] == {
        "export_blank_vs_original_error_or_unknown": 2
    }
    assert result["sheet_summaries"]["対象比率"]["field_counts"] == {"前年在籍": 1, "受給比率": 1}


def test_diff_workbook_business_values_reports_row_gap_samples_and_soft_matches(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    header = [
        "番号",
        "年度",
        "学校番号",
        "都道府県",
        "法人名",
        "学校名",
        "前年在籍",
        "前半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "後半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "年間",
        "家計急変多子世帯",
        "総計",
        "備考",
        "受給比率",
    ]
    _write_cells(
        original,
        {
            "対象比率": [
                header,
                [None, "2025年度", None, "東京都", "旧法人", "日本工学院専門学校", 100, *([None] * 15)],
                [None, "2025年度", None, "東京都", "別法人", "別学校", 50, *([None] * 15)],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "対象比率": [
                header,
                [1, "2025年度", None, "東京都", "新法人", "日本工学院専門学校", 100, *([None] * 15)],
                [2, "2025年度", None, "東京都", "追加法人", "追加学校", 10, *([None] * 15)],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["対象比率"], max_diffs=1)
    summary = result["sheet_summaries"]["対象比率"]

    assert summary["missing_rows"] == 2
    assert summary["extra_rows"] == 2
    assert summary["missing_rows_soft_matched"] == 1
    assert summary["extra_rows_soft_matched"] == 1
    assert summary["missing_row_samples"] == [
        "2025年度 | 東京都 | 旧法人 | 日本工学院専門学校",
    ]
    assert summary["extra_row_samples"] == [
        "2025年度 | 東京都 | 新法人 | 日本工学院専門学校",
    ]


def test_diff_workbook_business_values_reports_gakka_year_field_diff(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    key = ["東京都", "片柳学園", "日本工学院専門学校", "工業", "ITスペシャリスト科", "昼", 4]
    _write_cells(
        original,
        {
            "学科別": [
                [None, None, None, None, None, None, None, "2024年度", None, None],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "収定", "在籍", "中退率"],
                [*key, 100, 95, "#DIV/0!"],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "学科別": [
                [None, None, None, None, None, None, None, "2024年度", None, None],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "収定", "在籍", "中退率"],
                [*key, 100, 96, None],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["学科別"])
    summary = result["sheet_summaries"]["学科別"]

    assert result["missing_rows"] == 0
    assert result["extra_rows"] == 0
    assert result["differing_fields"] == 2
    assert summary["category_counts"] == {
        "numeric_mismatch": 1,
        "export_blank_vs_original_error_or_unknown": 1,
    }
    assert summary["field_counts"] == {"2024年度:中退率": 1, "2024年度:在籍": 1}


def test_diff_workbook_business_values_soft_matches_gakka_corporation_drift(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cells(
        original,
        {
            "学科別": [
                [None, None, None, None, None, None, None, "2024年度"],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "在籍"],
                ["東京都", "旧法人", "日本工学院専門学校", "工業", "ITスペシャリスト科", "昼", 4, 95],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "学科別": [
                [None, None, None, None, None, None, None, "2024年度"],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "在籍"],
                ["東京都", "新法人", "日本工学院専門学校", "工業", "ITスペシャリスト科", "昼", 4, 95],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["学科別"])
    summary = result["sheet_summaries"]["学科別"]

    assert summary["missing_rows"] == 1
    assert summary["extra_rows"] == 1
    assert summary["missing_rows_soft_matched"] == 1
    assert summary["extra_rows_soft_matched"] == 1


def test_diff_workbook_business_values_soft_matches_nfkc_width_variants(tmp_path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cells(
        original,
        {
            "在籍のみ抜粋": [
                [None, None, None, None, None, None, None, "在籍者数"],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "2024年度"],
                ["兵庫県", "大原学園", "姫路情報ITクリエイター法律専門学校", "工業", "情報IT学科", "昼", 2, 80],
            ]
        },
    )
    _write_cells(
        exported,
        {
            "在籍のみ抜粋": [
                [None, None, None, None, None, None, None, "在籍者数"],
                ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限", "2024年度"],
                ["兵庫県", "大原学園", "姫路情報ＩＴクリエイター法律専門学校", "工業", "情報IT学科", "昼", 2, 80],
            ]
        },
    )

    result = diff_workbook_business_values(exported, original, sheets=["在籍のみ抜粋"])
    summary = result["sheet_summaries"]["在籍のみ抜粋"]

    assert summary["missing_rows"] == 1
    assert summary["extra_rows"] == 1
    assert summary["missing_rows_soft_matched"] == 1
    assert summary["extra_rows_soft_matched"] == 1


def test_excel_exporter_confidence_thresholds_follow_central_env(monkeypatch) -> None:
    import eidp.excel.exporter as exporter_module

    monkeypatch.setenv("EIDP_CONFIDENCE_REVIEW", "0.76")
    monkeypatch.setenv("EIDP_CONFIDENCE_AUTO", "0.91")
    reloaded = importlib.reload(exporter_module)
    try:
        assert reloaded.EXCEL_MIN_EXTRACTION_CONFIDENCE == 0.76
        assert reloaded.EXCEL_AUTO_FLAG_EXTRACTION_CONFIDENCE == 0.91
        assert reloaded._low_confidence_reason() == "confidence<0.76"
    finally:
        monkeypatch.delenv("EIDP_CONFIDENCE_REVIEW", raising=False)
        monkeypatch.delenv("EIDP_CONFIDENCE_AUTO", raising=False)
        importlib.reload(exporter_module)


def test_excel_exporter_year_windows_follow_target_fiscal_year(monkeypatch) -> None:
    import eidp.config as config_module
    import eidp.excel.exporter as exporter_module

    monkeypatch.setenv("EIDP_TARGET_FISCAL_YEAR", "2025")
    importlib.reload(config_module)
    reloaded = importlib.reload(exporter_module)
    try:
        assert reloaded.FISCAL_YEARS[-1] == 2025
        assert reloaded.ENROLLMENT_YEARS[-1] == 2024
    finally:
        monkeypatch.delenv("EIDP_TARGET_FISCAL_YEAR", raising=False)
        importlib.reload(config_module)
        importlib.reload(exporter_module)


@pytest.fixture()
def sqlite_engine(tmp_path):
    db_path = tmp_path / "exporter.sqlite3"
    engine = create_engine(f"sqlite:///{db_path}", future=True)
    bootstrap_sqlite(engine)
    yield engine
    engine.dispose()


def test_sairoku_export_preserves_schools_without_year_status() -> None:
    engine = create_engine("sqlite:///:memory:")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE school (
                    id INTEGER PRIMARY KEY,
                    prefecture TEXT,
                    corporation_name TEXT,
                    school_name TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE school_year_status (
                    school_id INTEGER NOT NULL,
                    fiscal_year INTEGER NOT NULL,
                    legacy_status TEXT,
                    status TEXT,
                    -- Sprint 8.2.a added is_current as the discriminator for
                    -- the partial unique index. The exporter (Sprint 8.2.1)
                    -- now joins on it so demoted revisions don't shadow
                    -- current ones.
                    is_current INTEGER NOT NULL DEFAULT 1
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO school (id, prefecture, corporation_name, school_name)
                VALUES
                    (1, '東京都', '学校法人A', '採録済み学校'),
                    (2, '大阪府', '学校法人B', '未採録学校')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO school_year_status (school_id, fiscal_year, legacy_status, status, is_current)
                VALUES (1, 2025, '○', 'verified', 1)
                """
            )
        )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    with Session(engine) as session:
        count = _write_sairoku(worksheet, session)

    assert count == 2
    assert worksheet.cell(row=2, column=3).value == "採録済み学校"
    assert worksheet.cell(row=3, column=3).value == "未採録学校"
    assert all(cell.value is None for cell in worksheet[3][3:])


def test_export_master_filters_low_confidence_rows_and_reports_auto_flag(sqlite_engine, tmp_path) -> None:
    with Session(sqlite_engine) as session:
        low_school = School(
            prefecture="東京都",
            corporation_name="低信頼法人",
            school_name="低信頼専門学校",
            school_type="専門学校",
            status="active",
        )
        auto_school = School(
            prefecture="東京都",
            corporation_name="要確認法人",
            school_name="要確認専門学校",
            school_type="専門学校",
            status="active",
        )
        session.add_all([low_school, auto_school])
        session.flush()

        low_dept = Department(school_id=low_school.id, canonical_name="低信頼学科")
        auto_dept = Department(school_id=auto_school.id, canonical_name="要確認学科")
        session.add_all([low_dept, auto_dept])
        session.flush()

        session.add_all(
            [
                DepartmentYearly(
                    department_id=low_dept.id,
                    fiscal_year=2026,
                    revision=1,
                    is_current=True,
                    capacity=10,
                    enrollment=20,
                    extraction_confidence=0.64,
                    extraction_method="pdf_parse",
                ),
                DepartmentYearly(
                    department_id=auto_dept.id,
                    fiscal_year=2026,
                    revision=1,
                    is_current=True,
                    capacity=30,
                    enrollment=40,
                    extraction_confidence=0.80,
                    extraction_method="pdf_parse",
                ),
                SupportRecipient(
                    school_id=low_school.id,
                    fiscal_year=2026,
                    revision=1,
                    is_current=True,
                    annual_total=11,
                    grand_total=11,
                    extraction_confidence=0.64,
                ),
                SupportRecipient(
                    school_id=auto_school.id,
                    fiscal_year=2026,
                    revision=1,
                    is_current=True,
                    annual_total=33,
                    grand_total=33,
                    extraction_confidence=0.80,
                ),
            ]
        )
        session.commit()

        output = tmp_path / "master.xlsx"
        stats = export_master_workbook(session, output)

    assert stats["quality_department_yearly_low_confidence_current"] == 1
    assert stats["quality_department_yearly_auto_flag_current"] == 1
    assert stats["quality_support_recipient_low_confidence_current"] == 1
    assert stats["quality_support_recipient_auto_flag_current"] == 1
    assert stats["出力除外_低信頼"] == 2

    wb = openpyxl.load_workbook(output, data_only=True)
    try:
        assert "出力除外_低信頼" in wb.sheetnames
        excluded = wb["出力除外_低信頼"]
        assert [excluded.cell(row=1, column=col).value for col in range(1, 9)] == [
            "種別",
            "行ID",
            "学校名",
            "学科名",
            "年度",
            "confidence",
            "理由",
            "転記先",
        ]
        excluded_rows = [
            [excluded.cell(row=row, column=col).value for col in range(1, 9)]
            for row in range(2, excluded.max_row + 1)
        ]
        assert [
            [
                "department_yearly",
                1,
                "低信頼専門学校",
                "低信頼学科",
                2026,
                0.64,
                "confidence<0.70",
                "学科別/在籍のみ抜粋",
            ],
            ["support_recipient", 1, "低信頼専門学校", None, 2026, 0.64, "confidence<0.70", "対象比率"],
        ] == excluded_rows

        taisho = wb["対象比率"]
        taisho_school_names = [taisho.cell(row=row, column=6).value for row in range(2, taisho.max_row + 1)]
        assert taisho_school_names == ["要確認専門学校"]

        gakka = wb["学科別"]
        rows_by_school = {
            gakka.cell(row=row, column=3).value: row
            for row in range(3, gakka.max_row + 1)
        }
        assert "低信頼専門学校" in rows_by_school
        assert "要確認専門学校" in rows_by_school

        year_col = next(
            col
            for col in range(1, gakka.max_column + 1)
            if gakka.cell(row=1, column=col).value == "2026年度"
        )
        assert gakka.cell(row=rows_by_school["低信頼専門学校"], column=year_col).value is None
        assert gakka.cell(row=rows_by_school["要確認専門学校"], column=year_col).value == 30
    finally:
        wb.close()
