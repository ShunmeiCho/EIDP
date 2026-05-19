"""Sprint 8.4.c.3 — Excel preview page helper regression."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from eidp.db.models import (
    Department,
    DepartmentYearly,
    Document,
    ManualActionLog,
    School,
    SchoolYearStatus,
    SupportRecipient,
)
from eidp.db.sqlite_bootstrap import bootstrap_sqlite
from eidp.reports.coverage import ExportGapReport
from eidp.review._pages import excel_preview as excel_preview_mod
from eidp.review._pages.excel_preview import (
    SHEET_ORDER,
    audit_excel_preview_generated,
    build_preview_workbook,
    count_unmatched_and_gap,
    format_sheet_preview,
    format_sheet_preview_from_bytes,
    store_preview_session_state,
)


@pytest.fixture()
def engine(tmp_path: Path):
    db_path = tmp_path / "preview.sqlite3"
    engine = create_engine(f"sqlite:///{db_path}", future=True)
    bootstrap_sqlite(engine)
    yield engine
    engine.dispose()


def _seed(session: Session) -> None:
    """Two schools: A has data, B has none. Used for unmatched count."""
    a = School(prefecture="東京都", corporation_name="法人A", school_name="A学校",
               school_type="専門学校", status="active")
    b = School(prefecture="東京都", corporation_name="法人B", school_name="B学校",
               school_type="専門学校", status="active")
    session.add_all([a, b])
    session.flush()

    doc = Document(
        school_id=a.id,
        source_url="https://example.com/a.pdf",
        file_hash=("a" * 64),
        pdf_type="target", content_type="text",
        fiscal_year=2026, ingest_status="ingested",
        downloaded_at=datetime.now(UTC),
    )
    session.add(doc)
    session.flush()

    dept = Department(school_id=a.id, canonical_name="A学科")
    session.add(dept)
    session.flush()
    session.add(DepartmentYearly(
        department_id=dept.id, document_id=doc.id,
        fiscal_year=2026, revision=1, is_current=True,
        enrollment=10, capacity=20, graduates=2,
        extraction_method="pdf_parse",
    ))
    # SR: collected status
    session.add(SupportRecipient(
        school_id=a.id, document_id=doc.id, fiscal_year=2026,
        revision=1, is_current=True,
        annual_total=100, grand_total=100,
    ))
    # SchoolYearStatus: A has 'collected', B will have 'partial' (gap)
    session.add(SchoolYearStatus(
        school_id=a.id, document_id=doc.id, fiscal_year=2026,
        revision=1, is_current=True, status="collected",
    ))
    session.add(SchoolYearStatus(
        school_id=b.id, fiscal_year=2026,
        revision=1, is_current=True, status="partial",
    ))
    session.flush()


# ---------------------------------------------------------------------------
# build_preview_workbook
# ---------------------------------------------------------------------------


def test_build_preview_workbook_produces_4_sheets(engine):
    with Session(engine) as session:
        _seed(session)
        session.commit()

        preview = build_preview_workbook(session)
        assert preview.workbook.sheetnames == list(SHEET_ORDER)
        assert preview.counts.keys() == set(SHEET_ORDER)
        assert preview.quality_warnings == {
            "department_yearly_low_confidence_current": 0,
            "department_yearly_auto_flag_current": 0,
            "support_recipient_low_confidence_current": 0,
            "support_recipient_auto_flag_current": 0,
        }


def test_build_preview_workbook_reports_quality_warnings(engine):
    with Session(engine) as session:
        _seed(session)
        school = School(
            prefecture="東京都",
            corporation_name="法人C",
            school_name="C学校",
            school_type="専門学校",
            status="active",
        )
        session.add(school)
        session.flush()
        dept = Department(school_id=school.id, canonical_name="C学科")
        session.add(dept)
        session.flush()
        session.add(
            DepartmentYearly(
                department_id=dept.id,
                fiscal_year=2026,
                revision=1,
                is_current=True,
                enrollment=10,
                capacity=20,
                extraction_confidence=0.64,
                extraction_method="pdf_parse",
            )
        )
        session.add(
            SupportRecipient(
                school_id=school.id,
                fiscal_year=2026,
                revision=1,
                is_current=True,
                annual_total=10,
                grand_total=10,
                extraction_confidence=0.80,
            )
        )
        session.commit()

        preview = build_preview_workbook(session)

    assert preview.quality_warnings["department_yearly_low_confidence_current"] == 1
    assert preview.quality_warnings["support_recipient_auto_flag_current"] == 1


def test_quality_warning_messages_follow_configured_thresholds(monkeypatch):
    monkeypatch.setenv("EIDP_CONFIDENCE_REVIEW", "0.76")
    monkeypatch.setenv("EIDP_CONFIDENCE_AUTO", "0.91")

    assert "confidence 0.76 未満の current 行が 3 件" in excel_preview_mod._low_confidence_message(3)
    assert "confidence 0.76以上0.91未満の要確認行が 4 件" in excel_preview_mod._auto_flag_confidence_message(4)


def test_build_preview_workbook_does_not_touch_filesystem(engine, tmp_path):
    """The page contract is 'no disk write until download click'.
    build_preview_workbook MUST NOT create any file in tmp_path even
    when given a session — it's purely in-memory."""
    import os
    snapshot_before = set(os.listdir(tmp_path))
    with Session(engine) as session:
        _seed(session)
        session.commit()
        preview = build_preview_workbook(session)
        # to_bytes() also stays in-memory.
        b = preview.to_bytes()
        assert isinstance(b, (bytes, bytearray))
        assert len(b) > 0

    snapshot_after = set(os.listdir(tmp_path))
    # Only the SQLite file from the engine fixture should have appeared.
    new_files = snapshot_after - snapshot_before
    assert all(f.endswith(".sqlite3") or f.endswith(".sqlite3-wal") or f.endswith(".sqlite3-shm")
               for f in new_files), f"unexpected files: {new_files}"


def test_to_bytes_round_trip_via_openpyxl(engine):
    """The bytes returned must be a valid xlsx — load them back with
    openpyxl and confirm sheet names round-trip."""
    import io

    import openpyxl

    with Session(engine) as session:
        _seed(session)
        session.commit()
        preview = build_preview_workbook(session)
        b = preview.to_bytes()

    wb2 = openpyxl.load_workbook(io.BytesIO(b))
    assert set(wb2.sheetnames) == set(SHEET_ORDER)


# ---------------------------------------------------------------------------
# format_sheet_preview
# ---------------------------------------------------------------------------


def test_format_sheet_preview_caps_at_max_rows(engine):
    with Session(engine) as session:
        _seed(session)
        session.commit()
        preview = build_preview_workbook(session)
        rows = format_sheet_preview(preview.workbook, "採録状況", max_rows=2)
        assert len(rows) <= 2


def test_format_sheet_preview_unknown_sheet_raises(engine):
    with Session(engine) as session:
        preview = build_preview_workbook(session)
    with pytest.raises(ValueError, match="not in workbook"):
        format_sheet_preview(preview.workbook, "未知シート")


def test_format_sheet_preview_from_bytes_closes_workbook(monkeypatch):
    calls: list[str] = []

    class FakeWorkbook:
        sheetnames = ["採録状況"]

        def __getitem__(self, sheet_name: str):
            assert sheet_name == "採録状況"
            return self

        def iter_rows(self, *, values_only: bool):
            assert values_only is True
            return iter([("header",), ("row",)])

        def close(self) -> None:
            calls.append("closed")

    monkeypatch.setattr(
        excel_preview_mod.openpyxl,
        "load_workbook",
        lambda *args, **kwargs: FakeWorkbook(),
    )

    rows = format_sheet_preview_from_bytes(b"fake-xlsx", "採録状況", max_rows=1)

    assert rows == [["header"]]
    assert calls == ["closed"]


def test_store_preview_session_state_serializes_bytes_and_drops_workbook_handle() -> None:
    calls: list[str] = []
    state: dict[str, object] = {
        "excel_preview_workbook": object(),
    }

    class FakePreview:
        counts = {"採録状況": 1}
        quality_warnings = {"department_yearly_low_confidence_current": 0}

        def to_bytes(self) -> bytes:
            calls.append("to_bytes")
            return b"fake-xlsx"

        def close(self) -> None:
            calls.append("close")

    export_gap = object()

    store_preview_session_state(state, FakePreview(), export_gap=export_gap)  # type: ignore[arg-type]

    assert state["excel_preview_bytes"] == b"fake-xlsx"
    assert state["excel_preview_counts"] == {"採録状況": 1}
    assert state["excel_preview_gap"] is export_gap
    assert state["excel_preview_quality_warnings"] == {"department_yearly_low_confidence_current": 0}
    assert "excel_preview_workbook" not in state
    assert calls == ["to_bytes", "close"]


def test_audit_excel_preview_generated_writes_manual_action_log(engine) -> None:
    with Session(engine) as session:
        report = ExportGapReport(
            fiscal_year=2026,
            school_type="専門学校",
            total_schools=2,
            schools_with_url=1,
            no_url_schools=1,
            target_pdf_schools=1,
            stale_fallback_schools=0,
            missing_target_pdf_schools=1,
            extracted_schools=1,
            excel_ready_schools=1,
            target_yearly_rows=3,
        )

        audit_excel_preview_generated(
            session,
            counts={"採録状況": 2, "対象比率": 3},
            quality_warnings={"department_yearly_low_confidence_current": 1},
            export_gap=report,
        )
        session.commit()

        audit = session.query(ManualActionLog).one()
        assert audit.action_type == "excel_preview_generated"
        assert audit.target_table == "excel_export"
        assert audit.target_id is None
        assert '"fiscal_year": 2026' in (audit.new_value or "")
        assert '"excel_ready_schools": 1' in (audit.new_value or "")
        assert '"採録状況": 2' in (audit.new_value or "")


def test_render_records_excel_preview_generation_in_manual_action_log_contract() -> None:
    import inspect

    source = inspect.getsource(excel_preview_mod.render)

    assert "audit_excel_preview_generated(" in source
    assert "ui_excel_preview" in source


# ---------------------------------------------------------------------------
# count_unmatched_and_gap
# ---------------------------------------------------------------------------


def test_count_unmatched_and_gap_basic(engine):
    with Session(engine) as session:
        _seed(session)
        session.commit()
        counts = count_unmatched_and_gap(session)

    assert counts.total_schools == 2
    assert counts.target_pdf_schools == 1
    assert counts.missing_target_pdf_schools == 1
    assert counts.extracted_schools == 1
    assert counts.target_yearly_rows == 1
    assert counts.has_target_year_data is True


def test_count_unmatched_and_gap_can_cover_universities(engine):
    with Session(engine) as session:
        _seed(session)
        university = School(
            prefecture="東京都",
            corporation_name="法人U",
            school_name="U大学",
            school_type="大学",
            status="active",
        )
        session.add(university)
        session.flush()
        doc = Document(
            school_id=university.id,
            source_url="https://example.com/u.pdf",
            file_hash=("u" * 64),
            pdf_type="target",
            content_type="text",
            fiscal_year=2026,
            ingest_status="ingested",
            downloaded_at=datetime.now(UTC),
        )
        session.add(doc)
        session.flush()
        dept = Department(school_id=university.id, canonical_name="U学部")
        session.add(dept)
        session.flush()
        session.add(
            DepartmentYearly(
                department_id=dept.id,
                document_id=doc.id,
                fiscal_year=2026,
                revision=1,
                is_current=True,
                enrollment=30,
                capacity=40,
                graduates=5,
                extraction_method="pdf_parse",
            )
        )
        session.commit()

        specialty = count_unmatched_and_gap(session, school_type="専門学校")
        all_schools = count_unmatched_and_gap(session, school_type=None)

    assert specialty.total_schools == 2
    assert specialty.target_yearly_rows == 1
    assert all_schools.total_schools == 3
    assert all_schools.target_pdf_schools == 2
    assert all_schools.target_yearly_rows == 2


def test_count_excludes_inactive_schools(engine):
    with Session(engine) as session:
        _seed(session)
        # Mark B inactive — it should drop out of schools_total.
        b = session.query(School).filter(School.school_name == "B学校").one()
        b.status = "inactive"
        session.commit()

        counts = count_unmatched_and_gap(session)
        assert counts.total_schools == 1
        assert counts.missing_target_pdf_schools == 0


def test_count_uses_current_target_yearly_rows_only(engine):
    """Export readiness ignores non-current yearly rows."""
    with Session(engine) as session:
        _seed(session)
        b = session.query(School).filter(School.school_name == "B学校").one()
        dept = Department(school_id=b.id, canonical_name="B学科")
        session.add(dept)
        session.flush()
        session.add(
            DepartmentYearly(
                department_id=dept.id,
                fiscal_year=2026,
                revision=1,
                is_current=False,
                enrollment=999,
                capacity=999,
            )
        )
        session.commit()

        counts = count_unmatched_and_gap(session)
        assert counts.target_yearly_rows == 1
        assert counts.extracted_schools == 1
