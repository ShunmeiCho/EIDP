from __future__ import annotations

from pathlib import Path

import openpyxl
from typer.testing import CliRunner

from eidp.cli import app
from eidp.db.locking import acquire_lock


class FakeSession:
    def __init__(self) -> None:
        self.commits = 0
        self.rollbacks = 0
        self.closed = False

    def commit(self) -> None:
        self.commits += 1

    def rollback(self) -> None:
        self.rollbacks += 1

    def close(self) -> None:
        self.closed = True


def _write_cli_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Common"
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index).value = value
    workbook.save(path)
    workbook.close()


def test_diff_excel_values_mode_exits_nonzero_when_cells_differ(tmp_path: Path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    _write_cli_workbook(exported, [["学校", "定員"], ["A専門学校", 101]])
    _write_cli_workbook(original, [["学校", "定員"], ["A専門学校", 100]])

    result = CliRunner().invoke(
        app,
        [
            "diff-excel",
            str(exported),
            "--original",
            str(original),
            "--values",
            "--fail-on-diff",
        ],
    )

    assert result.exit_code == 1
    assert "Workbook value comparison" in result.output
    assert "differing_cells: 1" in result.output
    assert "Common!B2" in result.output


def test_diff_excel_business_values_mode_exits_nonzero_when_fields_differ(tmp_path: Path) -> None:
    exported = tmp_path / "exported.xlsx"
    original = tmp_path / "original.xlsx"
    header = [
        "番号",
        "年度",
        "学校番号",
        "都道府県",
        "法人名",
        "学校名",
        "前年在籍",
        "前半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "後半期",
        "第Ⅰ区分",
        "第Ⅱ区分",
        "第Ⅲ区分",
        "第Ⅳ区分",
        "年間",
        "家計急変多子世帯",
        "総計",
        "備考",
        "受給比率",
    ]
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "対象比率"
    worksheet.append(header)
    exported_row = [
        1,
        "2025年度",
        None,
        "東京都",
        "片柳学園",
        "日本工学院専門学校",
        6319,
        0,
        None,
        None,
        None,
        None,
        0,
        None,
        None,
        None,
        None,
        101,
        0,
        101,
        None,
        0.0160,
    ]
    worksheet.append(exported_row)
    workbook.save(exported)
    workbook.close()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "対象比率"
    worksheet.append(header)
    original_row = [
        None,
        "2025年度",
        None,
        "東京都",
        "片柳学園",
        "日本工学院専門学校",
        6319,
        0,
        None,
        None,
        None,
        None,
        0,
        None,
        None,
        None,
        None,
        100,
        0,
        100,
        None,
        0.0158,
    ]
    worksheet.append(original_row)
    workbook.save(original)
    workbook.close()

    result = CliRunner().invoke(
        app,
        [
            "diff-excel",
            str(exported),
            "--original",
            str(original),
            "--business-values",
            "--fail-on-diff",
        ],
    )

    assert result.exit_code == 1
    assert "Workbook business-value comparison" in result.output
    assert "differing_fields: 3" in result.output
    assert "対象比率 | 2025年度 | 東京都 | 片柳学園 | 日本工学院専門学校 | 年間" in result.output


def test_discover_pdfs_cli_uses_strict_target_fiscal_year(monkeypatch, tmp_path: Path) -> None:
    calls: dict[str, object] = {}
    fake_session = FakeSession()

    import eidp.config as config_mod
    import eidp.db.session as db_session
    import eidp.scraper.pdf_discovery as pdf_discovery

    def fake_run_pdf_discovery(session, storage_dir, **kwargs):  # noqa: ANN001, ANN003
        calls["session"] = session
        calls["storage_dir"] = storage_dir
        calls["kwargs"] = kwargs
        return {"downloaded": 0}

    monkeypatch.setattr(config_mod.settings, "target_fiscal_year", 2026)
    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(pdf_discovery, "run_pdf_discovery", fake_run_pdf_discovery)

    result = CliRunner().invoke(
        app,
        [
            "discover-pdfs",
            "--storage-dir",
            str(tmp_path / "pdfs"),
            "--batch-size",
            "3",
            "--evidence-log",
            str(tmp_path / "evidence.jsonl"),
        ],
    )

    assert result.exit_code == 0, result.output
    assert calls["session"] is fake_session
    assert calls["kwargs"]["target_fiscal_year"] == 2026
    assert calls["kwargs"]["strict_target_fiscal_year"] is True
    assert fake_session.commits == 1
    assert fake_session.rollbacks == 0
    assert fake_session.closed is True


def test_weekly_update_uses_strict_target_fiscal_year(monkeypatch, tmp_path: Path) -> None:
    calls: dict[str, object] = {}
    fake_session = FakeSession()

    import eidp.config as config_mod
    import eidp.db.session as db_session
    import eidp.excel.exporter as exporter
    import eidp.pipeline.ingest as ingest
    import eidp.scraper.pdf_discovery as pdf_discovery
    import eidp.scraper.url_discovery as url_discovery

    def fake_run_pdf_discovery(session, storage_dir, **kwargs):  # noqa: ANN001, ANN003
        calls["session"] = session
        calls["storage_dir"] = storage_dir
        calls["kwargs"] = kwargs
        return {"downloaded": 0}

    monkeypatch.setattr(config_mod.settings, "target_fiscal_year", 2026)
    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(url_discovery, "verify_urls_sync", lambda *_args, **_kwargs: {"verified": 0})
    monkeypatch.setattr(
        url_discovery,
        "get_discovery_stats",
        lambda _session: {"verified_disclosure": 0, "coverage_verified": "0%"},
    )
    monkeypatch.setattr(pdf_discovery, "run_pdf_discovery", fake_run_pdf_discovery)
    monkeypatch.setattr(ingest, "run_ingestion", lambda *_args, **_kwargs: {"processed": 0})
    monkeypatch.setattr(exporter, "export_master_workbook", lambda *_args, **_kwargs: {"exported": 0})

    result = CliRunner().invoke(
        app,
        [
            "weekly-update",
            "--storage-dir",
            str(tmp_path / "pdfs"),
            "--export-path",
            str(tmp_path / "weekly.xlsx"),
        ],
    )

    assert result.exit_code == 0, result.output
    assert calls["session"] is fake_session
    assert calls["kwargs"]["target_fiscal_year"] == 2026
    assert calls["kwargs"]["strict_target_fiscal_year"] is True
    assert fake_session.commits == 3
    assert fake_session.rollbacks == 0
    assert fake_session.closed is True


def test_weekly_update_refuses_when_app_lock_is_held(monkeypatch, tmp_path: Path) -> None:
    fake_session = FakeSession()
    data_dir = tmp_path / "data"
    lock_path = data_dir / ".lock"

    import eidp.config as config_mod
    import eidp.db.session as db_session
    import eidp.excel.exporter as exporter
    import eidp.pipeline.ingest as ingest
    import eidp.scraper.pdf_discovery as pdf_discovery
    import eidp.scraper.url_discovery as url_discovery

    monkeypatch.setattr(config_mod.settings, "data_dir", data_dir)
    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(url_discovery, "verify_urls_sync", lambda *_args, **_kwargs: {"verified": 0})
    monkeypatch.setattr(
        url_discovery,
        "get_discovery_stats",
        lambda _session: {"verified_disclosure": 0, "coverage_verified": "0%"},
    )
    monkeypatch.setattr(pdf_discovery, "run_pdf_discovery", lambda *_args, **_kwargs: {"downloaded": 0})
    monkeypatch.setattr(ingest, "run_ingestion", lambda *_args, **_kwargs: {"processed": 0})
    monkeypatch.setattr(exporter, "export_master_workbook", lambda *_args, **_kwargs: {"exported": 0})

    with acquire_lock(lock_path, owner="ui"):
        result = CliRunner().invoke(
            app,
            [
                "weekly-update",
                "--storage-dir",
                str(tmp_path / "pdfs"),
                "--export-path",
                str(tmp_path / "weekly.xlsx"),
            ],
        )

    assert result.exit_code == 5
    assert "another EIDP process is running" in result.output
    assert fake_session.commits == 0
    assert fake_session.rollbacks == 0
    assert fake_session.closed is False


def test_import_excel_refuses_when_app_lock_is_held(monkeypatch, tmp_path: Path) -> None:
    fake_session = FakeSession()
    data_dir = tmp_path / "data"
    lock_path = data_dir / ".lock"

    import eidp.config as config_mod
    import eidp.db.session as db_session
    import eidp.excel.importer as importer

    monkeypatch.setattr(config_mod.settings, "data_dir", data_dir)
    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(importer, "import_all", lambda *_args, **_kwargs: {"sheet": {"rows": 0}})

    with acquire_lock(lock_path, owner="ui"):
        result = CliRunner().invoke(app, ["import-excel", str(tmp_path / "master.xlsx")])

    assert result.exit_code == 5
    assert "another EIDP process is running" in result.output
    assert fake_session.commits == 0
    assert fake_session.rollbacks == 0
    assert fake_session.closed is False


def test_import_excel_surfaces_invalid_year_warning(monkeypatch, tmp_path: Path) -> None:
    fake_session = FakeSession()
    data_dir = tmp_path / "data"

    import eidp.config as config_mod
    import eidp.db.session as db_session
    import eidp.excel.importer as importer

    monkeypatch.setattr(config_mod.settings, "data_dir", data_dir)
    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(
        importer,
        "import_all",
        lambda *_args, **_kwargs: {
            "採録状況": {"rows": 1},
            "対象比率": {"rows": 0, "invalid_year": 2},
        },
    )

    result = CliRunner().invoke(app, ["import-excel", str(tmp_path / "master.xlsx")])

    assert result.exit_code == 0, result.output
    assert "対象比率: {'rows': 0, 'invalid_year': 2}" in result.output
    assert "WARNING: 対象比率 で想定外の年度の行を 2 件スキップしました。" in result.output
    assert "Import complete." in result.output
    assert fake_session.commits == 1
    assert fake_session.rollbacks == 0
    assert fake_session.closed is True


def test_export_excel_prints_quality_warnings_separately(monkeypatch, tmp_path: Path) -> None:
    fake_session = FakeSession()

    import eidp.db.session as db_session
    import eidp.excel.exporter as exporter

    monkeypatch.setattr(db_session, "SessionLocal", lambda: fake_session)
    monkeypatch.setattr(
        exporter,
        "export_master_workbook",
        lambda *_args, **_kwargs: {
            "採録状況": 3,
            "対象比率": 2,
            "出力除外_低信頼": 1,
            "quality_department_yearly_low_confidence_current": 1,
            "quality_support_recipient_auto_flag_current": 4,
        },
    )

    result = CliRunner().invoke(app, ["export-excel", "--output", str(tmp_path / "out.xlsx")])

    assert result.exit_code == 0, result.output
    assert "採録状況: 3 rows" in result.output
    assert "対象比率: 2 rows" in result.output
    assert "出力除外_低信頼: 1 rows" in result.output
    assert "Quality warnings:" in result.output
    assert "department_yearly_low_confidence_current: 1" in result.output
    assert "support_recipient_auto_flag_current: 4" in result.output
    assert "quality_department_yearly_low_confidence_current: 1 rows" not in result.output
    assert fake_session.closed is True
