from __future__ import annotations

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from eidp.db.models import (
    Base,
    Department,
    DepartmentYearly,
    Document,
    School,
)
from eidp.excel.competition_exporter import (
    CompetitionMatcher,
    MatchResult,
    TemplateRow,
    _diagnose_gap,
    _norm_dept_kana,
    _norm_school_key,
    export_competition_workbook,
)


def _session() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return Session(engine)


def _template_row(school_name: str = "東京X", dept_name: str | None = "X科") -> TemplateRow:
    return TemplateRow(
        row_index=1,
        school_name=school_name,
        dept_name=dept_name,
        duration_label="2年制",
    )


def _match(school_id: int | None, dept_ids: list[int] | None = None) -> MatchResult:
    return MatchResult(
        template_row=_template_row(),
        sheet_name="ゲーム",
        school_id=school_id,
        department_ids=dept_ids or [],
        matched_via="exact" if school_id else "unmatched",
    )


def test_school_missing_when_school_id_is_none() -> None:
    session = _session()
    reason, _ = _diagnose_gap(session, _match(None), 2026)
    assert reason == "school_missing"


def test_school_no_document_when_school_has_no_docs() -> None:
    session = _session()
    session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="X"))
    session.flush()
    reason, _ = _diagnose_gap(session, _match(1), 2026)
    assert reason == "school_no_document"


def test_school_mismatch_doc_rejected() -> None:
    session = _session()
    session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="X"))
    session.add(Document(
        id=10, school_id=1,
        source_url="https://x.ac.jp/pdf/y.pdf",
        pdf_type="target",
        ingest_status="school_mismatch",
    ))
    session.flush()
    reason, detail = _diagnose_gap(session, _match(1), 2026)
    assert reason == "school_mismatch_doc_rejected"
    assert "x.ac.jp" in detail


def test_school_doc_old_year_only_when_fy_coverage_is_stale() -> None:
    session = _session()
    session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="X"))
    session.add(Document(
        id=10, school_id=1,
        source_url="https://x.ac.jp/pdf/y.pdf",
        pdf_type="target",
        ingest_status="ingested",
        fiscal_year=2025,
    ))
    session.flush()
    reason, detail = _diagnose_gap(session, _match(1), 2026)
    assert reason == "school_doc_old_year_only"
    assert "2025" in detail


def test_dept_unmatched_when_school_fy_ok_but_no_dept_match() -> None:
    session = _session()
    session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="X"))
    session.add(Department(id=5, school_id=1, canonical_name="他の科"))
    session.add(Document(
        id=10, school_id=1,
        source_url="u", pdf_type="target",
        ingest_status="ingested", fiscal_year=2026,
    ))
    session.flush()
    # dept_ids empty → dept_unmatched
    reason, detail = _diagnose_gap(session, _match(1, dept_ids=[]), 2026)
    assert reason == "dept_unmatched"
    assert "db_dept_count=1" in detail


def test_norm_school_key_strips_common_suffixes() -> None:
    assert _norm_school_key("東京コミュニケーションアート専門学校") == "東京コミュニケーションアート"
    assert _norm_school_key("東京デザインテクノロジーセンター専門学校") == "東京デザインテクノロジーセンター"
    # Already short — no change
    assert _norm_school_key("東京コミュニケーションアート") == "東京コミュニケーションアート"


def test_matcher_suffix_strip_recovers_template_abbreviation() -> None:
    session = _session()
    session.add(School(
        id=1, prefecture="東京", corporation_name="滋慶学園",
        school_name="東京コミュニケーションアート専門学校",
    ))
    session.flush()

    matcher = CompetitionMatcher(session)
    row = TemplateRow(
        row_index=6,
        school_name="東京コミュニケーションアート",  # template uses abbreviated form
        dept_name=None,
        duration_label=None,
    )
    result = matcher.match("滋慶", row)
    assert result.school_id == 1
    assert result.matched_via == "suffix_strip"


def test_alias_pointing_to_multiple_schools_is_marked_ambiguous() -> None:
    """If SchoolAlias has the same alias_name pointing to two different
    school_id (data-integrity leak), the matcher must refuse rather than
    silently picking the first one.
    """
    from eidp.db.models import SchoolAlias

    session = _session()
    try:
        session.add(School(id=1, prefecture="東京", corporation_name="A", school_name="AAA"))
        session.add(School(id=2, prefecture="東京", corporation_name="B", school_name="BBB"))
        session.add(SchoolAlias(school_id=1, alias_name="X", alias_type="t", source="s"))
        session.add(SchoolAlias(school_id=2, alias_name="X", alias_type="t", source="s"))
        session.flush()

        matcher = CompetitionMatcher(session)
        row = TemplateRow(row_index=6, school_name="X", dept_name=None, duration_label=None)
        result = matcher.match("ゲーム", row)
        assert result.school_id is None
        assert result.matched_via == "school_name_ambiguous"
    finally:
        session.close()


def test_suffix_strip_collision_blocks_fuzzy_match() -> None:
    """Two schools with identical short keys must not silently collapse.

    Regression for Codex-flagged risk: first-wins indexing would write
    competition-row data to the wrong school.
    """
    session = _session()
    try:
        session.add(School(
            id=1, prefecture="東京", corporation_name="A",
            school_name="東京デザイン専門学校",
        ))
        session.add(School(
            id=2, prefecture="東京", corporation_name="B",
            school_name="東京デザイン学校",  # different school, same short key
        ))
        session.flush()

        matcher = CompetitionMatcher(session)
        # Template abbreviates either as "東京デザイン" — ambiguous
        row = TemplateRow(
            row_index=6, school_name="東京デザイン",
            dept_name=None, duration_label=None,
        )
        result = matcher.match("ゲーム", row)
        assert result.school_id is None
        assert result.matched_via == "school_name_ambiguous"

        reason, _ = _diagnose_gap(session, result, 2026)
        assert reason == "school_name_ambiguous"
    finally:
        session.close()


def test_norm_dept_kana_folds_long_vowel_and_small_i() -> None:
    # 'クリエーター' (ー) and 'クリエイター' (ィ) collapse to same key
    assert _norm_dept_kana("スーパークリエーター科(昼一)") == _norm_dept_kana("スーパークリエイター科(昼一)")
    # Different depts must NOT collapse
    assert _norm_dept_kana("看護学科") != _norm_dept_kana("情報学科")


def test_matcher_consumes_department_change_alias() -> None:
    """HIGH fix: DepartmentChange(change_type='alias') must flow into match().

    Previously the approve button in Proposals Review's Dept tab wrote a
    DepartmentChange row but matcher never read it, so approvals were
    audit-only. This test locks that behaviour in.
    """
    from eidp.db.models import DepartmentChange

    session = _session()
    try:
        session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="TSM"))
        # Canonical DB name: 'プロミュージシャン科'; template uses '学科'
        session.add(Department(id=521, school_id=1, canonical_name="プロミュージシャン科"))
        session.add(DepartmentChange(
            department_id=521,
            change_type="alias",
            fiscal_year=2026,
            old_name="プロミュージシャン学科",  # template form
            new_name="プロミュージシャン科",    # DB form
            verified=False,
        ))
        session.flush()

        matcher = CompetitionMatcher(session)
        row = TemplateRow(
            row_index=6, school_name="TSM",
            dept_name="プロミュージシャン学科",  # template form
            duration_label=None,
        )
        result = matcher.match("滋慶", row)
        assert result.school_id == 1
        assert result.department_ids == [521]
        assert "dept_alias" in result.matched_via
    finally:
        session.close()


def test_matcher_dept_kana_matches_across_transliteration_variants() -> None:
    """Kana fold alone should bridge the ー↔イ drift when the rest of the
    dept name is identical. Structural differences like '昼間部一' vs '昼一'
    are NOT this rule's concern (would need a separate shift-normalizer)."""
    session = _session()
    try:
        session.add(School(
            id=1, prefecture="東京", corporation_name="滋慶",
            school_name="東京コミュニケーションアート専門学校",
        ))
        session.add(Department(
            id=5, school_id=1, canonical_name="スーパークリエーター科",
        ))
        session.flush()

        matcher = CompetitionMatcher(session)
        row = TemplateRow(
            row_index=6,
            school_name="東京コミュニケーションアート",
            dept_name="スーパークリエイター科",  # イ instead of ー
            duration_label=None,
        )
        result = matcher.match("滋慶", row)
        assert result.school_id == 1
        assert result.department_ids == [5]
        assert "dept_kana" in result.matched_via
    finally:
        session.close()


def test_no_fy_data_when_dept_matched_but_yearly_missing() -> None:
    session = _session()
    session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="X"))
    session.add(Department(id=5, school_id=1, canonical_name="X科"))
    session.add(Document(
        id=10, school_id=1,
        source_url="u", pdf_type="target",
        ingest_status="ingested", fiscal_year=2026,
    ))
    session.flush()
    reason, _ = _diagnose_gap(session, _match(1, dept_ids=[5]), 2026)
    assert reason == "no_fy_data"


def test_export_overwrites_stale_gap_report_when_no_gaps(tmp_path) -> None:
    session = _session()
    try:
        session.add(School(id=1, prefecture="東京", corporation_name="C", school_name="学校A"))
        session.add(Department(id=5, school_id=1, canonical_name="ゲーム科"))
        session.add(
            DepartmentYearly(
                department_id=5,
                fiscal_year=2025,
                revision=1,
                is_current=True,
                enrollment=10,
                intl_students=2,
            )
        )
        session.flush()

        template = tmp_path / "template.xlsx"
        output = tmp_path / "out.xlsx"
        gap = tmp_path / "gap.csv"
        gap.write_text("stale gap that must be removed\n", encoding="utf-8")

        wb = Workbook()
        ws = wb.active
        ws.title = "ゲーム"
        ws.cell(2, 4, value=2025)
        ws.cell(3, 4, value="在籍数")
        ws.cell(3, 5, value="留学生")
        ws.cell(4, 1, value="学校A")
        ws.cell(4, 2, value="ゲーム科")
        ws.cell(4, 3, value="2年制")
        wb.save(template)

        stats = export_competition_workbook(session, template, output, 2025, gap)

        assert stats["unmatched"] == 0
        contents = gap.read_text(encoding="utf-8")
        assert "stale" not in contents
        assert contents.startswith("gap_reason,gap_detail,sheet,row")
    finally:
        session.close()
