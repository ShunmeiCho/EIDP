"""Sprint 8.2.1 — current-read-path regression.

When ``support_recipient`` and ``school_year_status`` carry both an
``is_current=False`` historical revision AND an ``is_current=True`` current
revision for the same (school, fiscal_year), every read path must report
the current value:

  * Excel exporter — 採録状況 sheet must use current revision's status, not
    a demoted older one. 対象比率 sheet must show one row per
    (school, fiscal_year), not the sum of all revisions.
  * Streamlit review dashboard — excluded count must reflect current
    revision only.
  * Review populate — excluded set used for skipping schools must reflect
    current revision only.
  * pdf_discovery — excluded_school_ids must reflect current revision only.
  * Reconciler — latest_status fetched in the loop must respect current
    flag and not pull a demoted row.

Without these filters a recent ``fiscal_year_override`` rewrite would leave
Excel and the discovery loop reading mid-state from before the override.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from eidp.db.models import (
    Document,
    School,
    SchoolYearStatus,
    SupportRecipient,
)
from eidp.db.sqlite_bootstrap import bootstrap_sqlite
from eidp.excel.exporter import _write_sairoku, _write_taisho_hiritu


@pytest.fixture()
def engine(tmp_path: Path):
    db_path = tmp_path / "current_reads.sqlite3"
    engine = create_engine(f"sqlite:///{db_path}", future=True)
    bootstrap_sqlite(engine)
    yield engine
    engine.dispose()


def _seed_school_with_two_revisions(session: Session) -> School:
    """Set up a school with both a demoted (is_current=False) and current
    (is_current=True) revision in SR and SYS for fiscal_year 2026."""
    school = School(
        prefecture="東京都",
        corporation_name="テスト法人",
        school_name="テスト専門学校",
        school_type="専門学校",
        status="active",
    )
    session.add(school)
    session.flush()

    doc = Document(
        school_id=school.id,
        source_url="https://example.com/test.pdf",
        file_hash=("a" * 64),
        pdf_type="target",
        content_type="text",
        fiscal_year=2026,
    )
    session.add(doc)
    session.flush()

    # SchoolYearStatus: rev 1 demoted = "partial", rev 2 current = "collected".
    # If a read path doesn't filter is_current, exporter would report the
    # OLDER value (partial) by joining without is_current and seeing the
    # row with the lower id.
    session.add(
        SchoolYearStatus(
            school_id=school.id,
            document_id=doc.id,
            fiscal_year=2026,
            revision=1,
            is_current=False,
            status="partial",
        )
    )
    session.add(
        SchoolYearStatus(
            school_id=school.id,
            document_id=doc.id,
            fiscal_year=2026,
            revision=2,
            is_current=True,
            status="collected",
        )
    )

    # SupportRecipient: rev 1 demoted = annual_total=80, rev 2 current = 110.
    # If exporter doesn't filter, it would emit BOTH rows.
    session.add(
        SupportRecipient(
            school_id=school.id,
            document_id=doc.id,
            fiscal_year=2026,
            revision=1,
            is_current=False,
            annual_total=80,
            grand_total=80,
        )
    )
    session.add(
        SupportRecipient(
            school_id=school.id,
            document_id=doc.id,
            fiscal_year=2026,
            revision=2,
            is_current=True,
            annual_total=110,
            grand_total=110,
        )
    )

    session.flush()
    return school


# ---------------------------------------------------------------------------
# Excel exporter
# ---------------------------------------------------------------------------


def test_taisho_hiritu_emits_only_current_support_recipient_row(engine):
    with Session(engine) as session:
        _seed_school_with_two_revisions(session)
        session.commit()

        wb = openpyxl.Workbook()
        ws = wb.active
        count = _write_taisho_hiritu(ws, session)

        assert count == 1, "対象比率 must show exactly one row, not both revisions"

        # Header row + one data row
        annual_total_idx = 17  # zero-indexed: column index of annual_total
        data_row = list(ws.iter_rows(values_only=True))[1]
        # annual_total is column "年間" — count to it from the header structure
        # (number=0, fiscal_year=1, school_number=2, prefecture=3, corp=4,
        # school=5, prev_enrollment=6, first_half_total=7, cat1..4=8..11,
        # second_half_total=12, cat1..4=13..16, annual_total=17, ...)
        assert data_row[annual_total_idx] == 110, (
            f"annual_total must be the current revision (110), got {data_row[annual_total_idx]}"
        )


def test_sairoku_uses_current_school_year_status_revision(engine):
    with Session(engine) as session:
        _seed_school_with_two_revisions(session)
        session.commit()

        wb = openpyxl.Workbook()
        ws = wb.active
        count = _write_sairoku(ws, session)

        assert count == 1
        rows = list(ws.iter_rows(values_only=True))
        # Header row, then one data row: prefecture, corp, school, then years
        data = rows[1]
        # The 2026 column for this school must be 'collected', NOT 'partial'.
        # FISCAL_YEARS is dynamic; locate the 2026 column by header index.
        header = rows[0]
        idx_2026 = header.index("2026年度") if "2026年度" in header else None
        if idx_2026 is None:
            pytest.skip("2026年度 not in computed FISCAL_YEARS for this run")
        assert data[idx_2026] == "collected", (
            f"採録状況 must show the current revision's status. got: {data[idx_2026]!r}"
        )


# ---------------------------------------------------------------------------
# Review dashboard & populate
# ---------------------------------------------------------------------------


def test_review_dashboard_excluded_filter_uses_current_revision(engine):
    """Demoted ``excluded_reason`` must NOT keep a school out of the
    dashboard's review queue."""
    from eidp.review.app import _load_dashboard_stats

    with Session(engine) as session:
        # Seed a school whose latest fiscal_year has TWO revisions:
        #   rev 1 demoted: excluded_reason='閉校'  (would have hidden the school)
        #   rev 2 current: excluded_reason=None
        school = School(
            prefecture="東京都",
            corporation_name="テスト法人",
            school_name="テスト専門学校",
            school_type="専門学校",
            status="active",
            school_code=None,  # so it lands in unresolved
        )
        session.add(school)
        session.flush()
        session.add(
            SchoolYearStatus(
                school_id=school.id, fiscal_year=2026,
                revision=1, is_current=False,
                status="excluded", excluded_reason="閉校",
            )
        )
        session.add(
            SchoolYearStatus(
                school_id=school.id, fiscal_year=2026,
                revision=2, is_current=True,
                status="collected", excluded_reason=None,
            )
        )
        session.commit()

        stats = _load_dashboard_stats(session)
        # The school must NOT count as excluded.
        assert stats["excluded"] == 0, (
            "demoted excluded_reason must not count under current-revision filter"
        )
        assert stats["unresolved"] == 1


def test_latest_excluded_school_ids_helper_uses_current_revision(engine):
    """The shared helper used by pdf_discovery / review / populate /
    reconciler must include only schools whose CURRENT revision in the
    LATEST fiscal year carries an excluded_reason."""
    from eidp.db.current_helpers import latest_excluded_school_ids

    with Session(engine) as session:
        s_active = School(
            prefecture="東京都", corporation_name="A法人",
            school_name="現役校", status="active",
        )
        s_excluded = School(
            prefecture="東京都", corporation_name="B法人",
            school_name="閉校校", status="active",
        )
        session.add_all([s_active, s_excluded])
        session.flush()

        # Active school: rev 1 demoted excluded, rev 2 current active.
        session.add_all([
            SchoolYearStatus(school_id=s_active.id, fiscal_year=2026,
                             revision=1, is_current=False,
                             status="excluded", excluded_reason="閉校"),
            SchoolYearStatus(school_id=s_active.id, fiscal_year=2026,
                             revision=2, is_current=True,
                             status="collected", excluded_reason=None),
            # Currently excluded school: just one current excluded revision.
            SchoolYearStatus(school_id=s_excluded.id, fiscal_year=2026,
                             revision=1, is_current=True,
                             status="excluded", excluded_reason="閉校"),
        ])
        session.commit()

        excluded = {row[0] for row in latest_excluded_school_ids(session)}
        assert s_excluded.id in excluded
        assert s_active.id not in excluded, (
            "school whose current revision is active must NOT be flagged as "
            "excluded just because a demoted revision says 閉校"
        )


def test_verify_identity_does_not_count_historic_year_excluded(engine):
    """Sprint 8.2.2 — verify_identity must scope excluded counting to the
    LATEST fiscal year's current revision. A school whose FY2025 current
    revision is 閉校 but whose FY2026 current revision is collected must
    NOT count as excluded (the school is back in scope).

    Without this, verify_identity would return excluded_no_code_needed=1
    for an active school, inflating ``pass`` and producing false positives
    on acceptance gates.
    """
    from pathlib import Path
    from eidp.matcher.reconciler import verify_identity

    with Session(engine) as session:
        # Active school: latest year is 2026 with no exclusion, but 2025
        # carried 閉校.
        active = School(
            prefecture="東京都", corporation_name="A法人",
            school_name="現役校", status="active", school_code=None,
        )
        session.add(active)
        session.flush()
        session.add_all([
            # FY2025 current = excluded (historical state)
            SchoolYearStatus(school_id=active.id, fiscal_year=2025,
                             revision=1, is_current=True,
                             status="excluded", excluded_reason="閉校"),
            # FY2026 current = active again (school re-opened or re-found)
            SchoolYearStatus(school_id=active.id, fiscal_year=2026,
                             revision=1, is_current=True,
                             status="collected", excluded_reason=None),
        ])
        session.commit()

        report = verify_identity(session, Path("/nonexistent"))

        assert report["excluded_no_code_needed"] == 0, (
            "FY2025 'historical' exclusion must NOT count when FY2026 "
            f"current revision is active. report={report}"
        )
        assert report["truly_unresolved"] == 1, report


def test_reconciler_verify_identity_excluded_count_uses_current_revision(engine):
    """``verify_identity`` counts excluded schools to subtract from the
    'still missing school_code' tally. After 8.2.1 it must scope that count
    to the current revision."""
    from pathlib import Path

    from eidp.matcher.reconciler import verify_identity

    with Session(engine) as session:
        # School with rev1 demoted exclusion, rev2 current active, no code.
        active = School(
            prefecture="東京都", corporation_name="A法人",
            school_name="現役校", status="active", school_code=None,
        )
        # School with current excluded — should count as excluded.
        excluded = School(
            prefecture="東京都", corporation_name="B法人",
            school_name="閉校校", status="active", school_code=None,
        )
        session.add_all([active, excluded])
        session.flush()

        session.add_all([
            SchoolYearStatus(school_id=active.id, fiscal_year=2025,
                             revision=1, is_current=False,
                             status="excluded", excluded_reason="閉校"),
            SchoolYearStatus(school_id=active.id, fiscal_year=2025,
                             revision=2, is_current=True,
                             status="collected", excluded_reason=None),
            SchoolYearStatus(school_id=excluded.id, fiscal_year=2025,
                             revision=1, is_current=True,
                             status="excluded", excluded_reason="閉校"),
        ])
        session.commit()

        # data_dir of "/nonexistent" makes verify_identity skip the optional
        # target_institutions.xlsx coverage check; the path we're asserting
        # (excluded_no_code_needed count) is unaffected.
        report = verify_identity(session, Path("/nonexistent"))

        # Only the school whose current revision is excluded should count
        # toward "no code needed because school is excluded". The active
        # school must remain in truly_unresolved despite its demoted 閉校
        # revision.
        assert report["excluded_no_code_needed"] == 1, report
        assert report["truly_unresolved"] == 1, report
