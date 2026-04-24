"""Template-preserving exporter for 競合校の在校生数.xlsx.

Reads the担当者 template workbook and overlays new fiscal-year columns
without disturbing the existing 16-sheet structure or row order.

Output:
- Filled workbook (template clone + new fiscal-year columns + ratios)
- Gap report CSV (sorted), listing template rows that could not be matched
"""

from __future__ import annotations

import csv
import re
import shutil
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func as sql_func
from sqlalchemy.orm import Session

from eidp.db.models import Department, DepartmentYearly, Document, School, SchoolAlias

log = structlog.get_logger(__name__)

_ENROLLMENT_HEADER_TEXT = "在籍数"
_INTL_HEADER_TEXT = "留学生"
_PREV_RATIO_HEADER = "前年比"
_INTL_RATIO_HEADER_PARTS = ("留学生", "比率")

# Column gap between two side-by-side comparison blocks. Sample has gap=1
# (cols 17-18 empty between left 3-16 and right 19-32). Use >=2 as a block
# boundary heuristic.
_BLOCK_GAP_THRESHOLD = 2


def _norm(s: object) -> str:
    """NFKC normalize and strip ALL whitespace for matching."""
    if s is None:
        return ""
    text = str(s)
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", "", text)


# Suffix strip for fuzzy school-name matching. 競合校 templates often use
# abbreviated forms ("東京コミュニケーションアート") while DB stores the
# formal name ("東京コミュニケーションアート専門学校"). Strip so both
# collapse onto the same canonical key.
_SCHOOL_NAME_SUFFIXES = ("専門学校", "高等専門学校", "専修学校", "学校", "専門", "大学", "短期大学")


def _norm_school_key(s: object) -> str:
    """NFKC + whitespace strip + common suffix strip for school-name match."""
    key = _norm(s)
    for suffix in _SCHOOL_NAME_SUFFIXES:
        if key.endswith(suffix) and len(key) > len(suffix) + 1:
            return key[: -len(suffix)]
    return key


@dataclass(frozen=True)
class YearColumns:
    """Where to read/write a fiscal-year's pair of cells."""

    fiscal_year: int
    zaiseki_col: int  # 在籍数 column (1-indexed)
    intl_col: int  # 留学生 column (1-indexed)


@dataclass(frozen=True)
class TemplateRow:
    """A single school+dept entry in the template."""

    row_index: int
    school_name: str
    dept_name: str | None
    duration_label: str | None
    block_id: int = 0  # 0 = left/only block, 1 = right block (rollup sheet)


@dataclass
class SheetBlock:
    """One comparison block within a sheet.

    Category sheets have 1 block. 学校単位での比較 has 2 side-by-side blocks.
    """

    school_col: int
    dept_col: int | None
    duration_col: int | None
    year_cols: list[YearColumns]
    data_rows: list[TemplateRow]


@dataclass
class SheetSchema:
    """Parsed structure of one template sheet."""

    name: str
    header_row: int
    blocks: list[SheetBlock]
    is_rollup: bool


def _find_header_row_and_year_triplets(
    ws: Worksheet,
) -> tuple[int, list[YearColumns]]:
    """Scan first few rows for the 在籍数 header and infer all year columns."""
    for r in range(1, min(8, ws.max_row + 1)):
        triplets: list[YearColumns] = []
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value != _ENROLLMENT_HEADER_TEXT:
                continue
            year: int | None = None
            for offset in (1, 2):
                if r - offset < 1:
                    continue
                cand = ws.cell(r - offset, c).value
                if isinstance(cand, int) and 2010 <= cand <= 2100:
                    year = cand
                    break
                if isinstance(cand, str):
                    m = re.search(r"(20\d{2})", cand)
                    if m:
                        year = int(m.group(1))
                        break
            if year is None:
                continue
            intl_col = c + 1
            if ws.cell(r, intl_col).value != _INTL_HEADER_TEXT:
                continue
            triplets.append(
                YearColumns(fiscal_year=year, zaiseki_col=c, intl_col=intl_col)
            )
        if triplets:
            return r, sorted(triplets, key=lambda y: y.zaiseki_col)
    return -1, []


def _group_triplets_into_blocks(
    triplets: list[YearColumns],
) -> list[list[YearColumns]]:
    """Split sorted year triplets into blocks based on column-gap boundaries."""
    if not triplets:
        return []
    blocks: list[list[YearColumns]] = [[triplets[0]]]
    for prev, cur in zip(triplets, triplets[1:]):
        # Each triplet spans 2 cols (zaiseki, intl). Gap > 2 → new block.
        gap = cur.zaiseki_col - prev.intl_col
        if gap > _BLOCK_GAP_THRESHOLD:
            blocks.append([cur])
        else:
            blocks[-1].append(cur)
    return blocks


def _detect_block_id_cols(
    ws: Worksheet, first_year_col: int, is_rollup: bool
) -> tuple[int, int | None, int | None]:
    """Return (school_col, dept_col, duration_col) for a block.

    Category sheets: columns immediately before first_year_col hold
    school/dept/duration. Rollup sheet: only school column.
    """
    if is_rollup:
        # school sits at first_year_col - 1 (e.g. 2 for left block, 18 for right)
        return max(1, first_year_col - 1), None, None
    # Category: school=1, dept=2, duration=3 regardless of first_year_col
    return 1, 2, 3


def parse_sheet_schema(ws: Worksheet) -> SheetSchema | None:
    """Parse one sheet into a SheetSchema (multi-block aware)."""
    header_row, triplets = _find_header_row_and_year_triplets(ws)
    if header_row < 0:
        return None
    year_groups = _group_triplets_into_blocks(triplets)
    is_rollup = ws.title == "学校単位での比較"

    blocks: list[SheetBlock] = []
    for block_id, year_cols in enumerate(year_groups):
        first_year_col = year_cols[0].zaiseki_col
        school_col, dept_col, duration_col = _detect_block_id_cols(
            ws, first_year_col, is_rollup
        )

        data_rows: list[TemplateRow] = []
        last_school = ""
        for r in range(header_row + 1, ws.max_row + 1):
            school_raw = ws.cell(r, school_col).value
            dept_raw = ws.cell(r, dept_col).value if dept_col else None
            duration_raw = ws.cell(r, duration_col).value if duration_col else None

            school = _norm(school_raw)
            if school:
                last_school = school
            dept = _norm(dept_raw) if dept_raw else None

            if not school and not dept:
                continue
            effective_school = school if school else last_school
            if not effective_school:
                continue
            if is_rollup and not school:
                # rollup has no dept continuation; each data row has a school
                continue

            data_rows.append(
                TemplateRow(
                    row_index=r,
                    school_name=effective_school,
                    dept_name=dept,
                    duration_label=str(duration_raw) if duration_raw else None,
                    block_id=block_id,
                )
            )

        blocks.append(
            SheetBlock(
                school_col=school_col,
                dept_col=dept_col,
                duration_col=duration_col,
                year_cols=year_cols,
                data_rows=data_rows,
            )
        )

    return SheetSchema(
        name=ws.title, header_row=header_row, blocks=blocks, is_rollup=is_rollup
    )


def parse_template(template_path: Path) -> dict[str, SheetSchema]:
    wb = load_workbook(str(template_path), data_only=True)
    schemas: dict[str, SheetSchema] = {}
    for name in wb.sheetnames:
        schema = parse_sheet_schema(wb[name])
        if schema is not None:
            schemas[name] = schema
    return schemas


@dataclass
class MatchResult:
    """Outcome of matching a TemplateRow to DB entities."""

    template_row: TemplateRow
    sheet_name: str
    school_id: int | None
    department_ids: list[int] = field(default_factory=list)
    matched_via: str = "unmatched"
    gap_reason: str = ""
    gap_detail: str = ""


class CompetitionMatcher:
    """Match template (school, dept) rows to DB entities."""

    def __init__(self, session: Session) -> None:
        self.session = session
        self._school_index: dict[str, int] = {}
        self._alias_index: dict[str, int] = {}
        self._dept_cache: dict[int, list[Department]] = {}
        self._build_indices()

    def _build_indices(self) -> None:
        # Primary: exact normalised name lookup.
        # Secondary: suffix-stripped key so '東京コミュニケーションアート'
        # (template) matches '東京コミュニケーションアート専門学校' (DB).
        self._school_key_index: dict[str, int] = {}
        for s in self.session.query(School).all():
            full = _norm(s.school_name)
            if full and full not in self._school_index:
                self._school_index[full] = s.id
            short = _norm_school_key(s.school_name)
            if short and short not in self._school_key_index:
                self._school_key_index[short] = s.id
        for a in self.session.query(SchoolAlias).all():
            key = _norm(a.alias_name)
            if key and key not in self._alias_index:
                self._alias_index[key] = a.school_id
            short = _norm_school_key(a.alias_name)
            if short and short not in self._alias_index:
                self._alias_index[short] = a.school_id

    def _depts_for_school(self, school_id: int) -> list[Department]:
        if school_id not in self._dept_cache:
            self._dept_cache[school_id] = (
                self.session.query(Department)
                .filter(Department.school_id == school_id)
                .all()
            )
        return self._dept_cache[school_id]

    def match(self, sheet_name: str, row: TemplateRow) -> MatchResult:
        school_key = _norm(row.school_name)
        school_id = self._school_index.get(school_key)
        matched_via = "exact" if school_id is not None else ""
        if school_id is None:
            school_id = self._alias_index.get(school_key)
            matched_via = "alias" if school_id is not None else ""
        if school_id is None:
            short_key = _norm_school_key(row.school_name)
            school_id = self._school_key_index.get(short_key)
            matched_via = "suffix_strip" if school_id is not None else "unmatched"

        if school_id is None:
            return MatchResult(
                template_row=row, sheet_name=sheet_name,
                school_id=None, matched_via="unmatched",
            )

        if row.dept_name is None:
            return MatchResult(
                template_row=row, sheet_name=sheet_name,
                school_id=school_id, matched_via=matched_via,
            )

        dept_key = _norm(row.dept_name)
        depts = self._depts_for_school(school_id)
        matching = [d.id for d in depts if _norm(d.canonical_name) == dept_key]
        if matching:
            return MatchResult(
                template_row=row, sheet_name=sheet_name,
                school_id=school_id, department_ids=matching,
                matched_via=matched_via + "+dept",
            )
        for d in depts:
            cn = _norm(d.canonical_name)
            if dept_key and cn and (dept_key in cn or cn in dept_key):
                matching.append(d.id)
        return MatchResult(
            template_row=row, sheet_name=sheet_name,
            school_id=school_id, department_ids=matching,
            matched_via=(matched_via + "+dept_substr") if matching
            else (matched_via + "+dept_unmatched"),
        )


@dataclass
class YearlyAggregate:
    enrollment: int | None
    intl_students: int | None


def _aggregate_yearly(
    session: Session, dept_ids: list[int], fiscal_year: int
) -> YearlyAggregate:
    if not dept_ids:
        return YearlyAggregate(enrollment=None, intl_students=None)
    rows = (
        session.query(DepartmentYearly)
        .filter(
            DepartmentYearly.department_id.in_(dept_ids),
            DepartmentYearly.fiscal_year == fiscal_year,
            DepartmentYearly.is_current.is_(True),
        )
        .all()
    )
    if not rows:
        return YearlyAggregate(enrollment=None, intl_students=None)
    enroll = sum((r.enrollment or 0) for r in rows) if any(
        r.enrollment is not None for r in rows
    ) else None
    intl = sum((r.intl_students or 0) for r in rows) if any(
        r.intl_students is not None for r in rows
    ) else None
    return YearlyAggregate(enrollment=enroll, intl_students=intl)


def _aggregate_school_yearly(
    session: Session, school_id: int, fiscal_year: int
) -> YearlyAggregate:
    dept_ids = [
        d.id for d in session.query(Department)
        .filter(Department.school_id == school_id).all()
    ]
    return _aggregate_yearly(session, dept_ids, fiscal_year)


def _diagnose_gap(
    session: Session, result: MatchResult, fiscal_year: int
) -> tuple[str, str]:
    """Categorise why this row ended up in the gap report.

    Returns (reason, detail). Reasons business-operators can act on:
      school_missing                  — school not in School table at all
      school_mismatch_doc_rejected    — school exists, target PDF downloaded
                                         but ingest marked school_mismatch
      school_no_document              — school exists, no PDF ever downloaded
      school_doc_old_year_only        — school has docs, but none for this FY
      dept_unmatched                  — school ingested, dept name diverges
      no_fy_data                      — dept matched, no yearly row this FY
    """
    if result.school_id is None:
        return "school_missing", ""

    docs = (
        session.query(Document)
        .filter(Document.school_id == result.school_id)
        .all()
    )
    if not docs:
        return "school_no_document", ""

    mismatched = [d for d in docs if d.ingest_status == "school_mismatch"]
    ingested_fys = {d.fiscal_year for d in docs if d.ingest_status == "ingested"}
    if fiscal_year not in ingested_fys:
        if mismatched:
            urls = ", ".join(sorted({d.source_url for d in mismatched})[:2])
            return "school_mismatch_doc_rejected", urls
        if ingested_fys:
            return "school_doc_old_year_only", f"have_fys={sorted(y for y in ingested_fys if y)}"
        return "school_no_document", f"doc_statuses={sorted({d.ingest_status or 'none' for d in docs})}"

    # School has data for this FY but dept-level aggregation failed
    if result.template_row.dept_name and not result.department_ids:
        return "dept_unmatched", f"db_dept_count={len(session.query(Department).filter(Department.school_id == result.school_id).all())}"

    return "no_fy_data", ""


def auto_select_fiscal_year(session: Session) -> int:
    """Pick the fiscal year with the most DepartmentYearly coverage.

    Preferred signal for 担当者 reports where 最新 really means 'the most
    populated year in DB', not a calendar projection.
    """
    rows = (
        session.query(
            DepartmentYearly.fiscal_year,
            sql_func.count(DepartmentYearly.id),
        )
        .filter(
            DepartmentYearly.document_id.isnot(None),
            DepartmentYearly.is_current.is_(True),
        )
        .group_by(DepartmentYearly.fiscal_year)
        .order_by(sql_func.count(DepartmentYearly.id).desc())
        .all()
    )
    if rows:
        return int(rows[0][0])
    # Fallback to calendar year if DB is empty
    from datetime import datetime
    return datetime.now().year


def _append_year_columns_to_block(
    ws: Worksheet,
    schema: SheetSchema,
    block: SheetBlock,
    fiscal_year: int,
    following_blocks: list[SheetBlock] | None = None,
) -> YearColumns:
    """Append fiscal-year header cells at the end of a block."""
    following_blocks = following_blocks or []
    new_zaiseki = block.year_cols[-1].intl_col + 1
    new_intl = new_zaiseki + 1
    if following_blocks:
        ws.insert_cols(new_zaiseki, amount=2)
        for shifted in following_blocks:
            shifted.school_col += 2
            if shifted.dept_col is not None:
                shifted.dept_col += 2
            if shifted.duration_col is not None:
                shifted.duration_col += 2
            shifted.year_cols = [
                YearColumns(
                    fiscal_year=yc.fiscal_year,
                    zaiseki_col=yc.zaiseki_col + 2,
                    intl_col=yc.intl_col + 2,
                )
                for yc in shifted.year_cols
            ]
    if schema.header_row > 1:
        ws.cell(schema.header_row - 1, new_zaiseki, value=fiscal_year)
    ws.cell(schema.header_row, new_zaiseki, value=_ENROLLMENT_HEADER_TEXT)
    ws.cell(schema.header_row, new_intl, value=_INTL_HEADER_TEXT)
    # Ratio headers one row below header_row (R5 in most sheets)
    ratio_row = schema.header_row + 1
    if ws.cell(ratio_row, new_zaiseki).value is None:
        ws.cell(ratio_row, new_zaiseki, value=_PREV_RATIO_HEADER)
    if ws.cell(ratio_row, new_intl).value is None:
        ws.cell(ratio_row, new_intl, value="留学生\n比率")
    return YearColumns(
        fiscal_year=fiscal_year, zaiseki_col=new_zaiseki, intl_col=new_intl
    )


def _find_year_cols(
    block: SheetBlock, fiscal_year: int
) -> YearColumns | None:
    for yc in block.year_cols:
        if yc.fiscal_year == fiscal_year:
            return yc
    return None


def _prev_year_enrollment(
    ws: Worksheet, row_index: int, block: SheetBlock, fiscal_year: int
) -> int | None:
    """Read prior-year 在籍数 from template row for ratio computation."""
    prev_year = fiscal_year - 1
    prev_cols = _find_year_cols(block, prev_year)
    if prev_cols is None:
        return None
    val = ws.cell(row_index, prev_cols.zaiseki_col).value
    if isinstance(val, (int, float)) and val:
        return int(val)
    return None


def export_competition_workbook(
    session: Session,
    template_path: Path,
    output_path: Path,
    fiscal_year: int | None = None,
    gap_report_path: Path | None = None,
) -> dict[str, int]:
    """Generate the 競合校の在校生数 workbook for the given fiscal year.

    fiscal_year=None → pick the year with the most DB coverage.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"template not found: {template_path}")

    if fiscal_year is None:
        fiscal_year = auto_select_fiscal_year(session)
        log.info("auto_fiscal_year_selected", fiscal_year=fiscal_year)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = load_workbook(str(output_path))
    matcher = CompetitionMatcher(session)
    matched = 0
    unmatched_rows: list[MatchResult] = []
    cells_written = 0
    ratio_cells_written = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        schema = parse_sheet_schema(ws)
        if schema is None:
            continue

        for block_index, block in enumerate(schema.blocks):
            existing = _find_year_cols(block, fiscal_year)
            target_cols = existing or _append_year_columns_to_block(
                ws, schema, block, fiscal_year, schema.blocks[block_index + 1:]
            )
            ratio_row_offset = 1  # ratio is one row below data

            for row in block.data_rows:
                result = matcher.match(sheet_name, row)
                if schema.is_rollup:
                    if result.school_id is None:
                        unmatched_rows.append(result)
                        continue
                    agg = _aggregate_school_yearly(
                        session, result.school_id, fiscal_year
                    )
                else:
                    if not result.department_ids:
                        unmatched_rows.append(result)
                        continue
                    agg = _aggregate_yearly(
                        session, result.department_ids, fiscal_year
                    )

                if agg.enrollment is None and agg.intl_students is None:
                    unmatched_rows.append(result)
                    continue

                matched += 1

                if agg.enrollment is not None:
                    ws.cell(row.row_index, target_cols.zaiseki_col,
                            value=agg.enrollment)
                    cells_written += 1
                if agg.intl_students is not None:
                    ws.cell(row.row_index, target_cols.intl_col,
                            value=agg.intl_students)
                    cells_written += 1

                # 前年比 = this_year_enrollment / prev_year_enrollment
                prev_enroll = _prev_year_enrollment(
                    ws, row.row_index, block, fiscal_year
                )
                if agg.enrollment and prev_enroll:
                    ratio = agg.enrollment / prev_enroll
                    ws.cell(row.row_index + ratio_row_offset,
                            target_cols.zaiseki_col, value=ratio)
                    ratio_cells_written += 1
                # 留学生比率 = intl_students / enrollment  (legitimate ratio row)
                if agg.enrollment and agg.intl_students is not None:
                    intl_ratio = agg.intl_students / agg.enrollment
                    ws.cell(row.row_index + ratio_row_offset,
                            target_cols.intl_col, value=intl_ratio)
                    ratio_cells_written += 1

    wb.save(str(output_path))

    if gap_report_path is not None and unmatched_rows:
        for u in unmatched_rows:
            u.gap_reason, u.gap_detail = _diagnose_gap(session, u, fiscal_year)

        gap_report_path.parent.mkdir(parents=True, exist_ok=True)
        # Sort for deterministic output: reason → sheet → school → dept → row
        unmatched_rows.sort(
            key=lambda u: (
                u.gap_reason,
                u.sheet_name,
                u.template_row.school_name,
                u.template_row.dept_name or "",
                u.template_row.row_index,
            )
        )
        with gap_report_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(
                ["gap_reason", "gap_detail", "sheet", "row", "block_id",
                 "school_name", "dept_name", "duration", "school_id",
                 "matched_via"]
            )
            for u in unmatched_rows:
                writer.writerow([
                    u.gap_reason,
                    u.gap_detail,
                    u.sheet_name,
                    u.template_row.row_index,
                    u.template_row.block_id,
                    u.template_row.school_name,
                    u.template_row.dept_name or "",
                    u.template_row.duration_label or "",
                    u.school_id or "",
                    u.matched_via,
                ])

    log.info(
        "competition_export_complete",
        output=str(output_path),
        fiscal_year=fiscal_year,
        matched=matched,
        unmatched=len(unmatched_rows),
        cells_written=cells_written,
        ratio_cells_written=ratio_cells_written,
    )

    return {
        "matched": matched,
        "unmatched": len(unmatched_rows),
        "cells_written": cells_written,
        "ratio_cells_written": ratio_cells_written,
        "fiscal_year": fiscal_year,
    }
