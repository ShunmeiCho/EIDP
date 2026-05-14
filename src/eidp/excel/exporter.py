"""Excel exporter -- generates master workbook from PostgreSQL database.

Produces 4 sheets matching the legacy format:
  Sheet 1: 採録状況 (school columns + fiscal-year status columns)
  Sheet 2: 対象比率 (22 cols)
  Sheet 3: 学科別 (dynamic fiscal-year blocks, multi-row header)
  Sheet 4: 在籍のみ抜粋 (one-year-lag enrollment blocks, multi-row header)
"""

import re
import unicodedata
from collections import Counter
from collections.abc import Iterator
from itertools import zip_longest
from pathlib import Path
from typing import Any, TypedDict, TypeGuard

import openpyxl  # type: ignore[import-untyped]
import structlog
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from sqlalchemy import text
from sqlalchemy.orm import Session

from eidp.config import settings
from eidp.db.current_helpers import IS_CURRENT_TRUE_SQL
from eidp.extraction_confidence import thresholds_from_env

log = structlog.get_logger()

def _compute_fiscal_years() -> list[int]:
    """Compute Excel fiscal-year columns from the configured target FY."""
    return list(range(2019, int(settings.target_fiscal_year) + 1))

FISCAL_YEARS = _compute_fiscal_years()
ENROLLMENT_YEARS = FISCAL_YEARS[:-1]  # enrollment data lags by 1 year

# 学科別: year block fields (DB column order)
YEAR_BLOCK_FIELDS = [
    "capacity", "enrollment", "intl_students", "graduates",
    "advanced", "employed", "other", "prev_enrollment",
    "dropouts", "dropout_rate",
]
# Excel headers for year block columns
YEAR_BLOCK_HEADERS = [
    "収定", "在籍", "留学生", "卒業", "進学", "就職",
    "その他", "前年在籍", "中退", "中退率",
]
_EXCEL_CONFIDENCE_THRESHOLDS = thresholds_from_env()
EXCEL_MIN_EXTRACTION_CONFIDENCE = _EXCEL_CONFIDENCE_THRESHOLDS.review
EXCEL_AUTO_FLAG_EXTRACTION_CONFIDENCE = _EXCEL_CONFIDENCE_THRESHOLDS.auto
LOW_CONFIDENCE_EXCLUSION_SHEET = "出力除外_低信頼"

ExcelCell = object
ExcelRow = list[ExcelCell]
YearlyData = tuple[ExcelCell, ...]


class WorkbookValueDiffSample(TypedDict):
    sheet: str
    cell: str
    exported: object
    original: object


class WorkbookValueDiff(TypedDict):
    ok: bool
    missing_sheets: list[str]
    extra_sheets: list[str]
    differing_cells: int
    samples: list[WorkbookValueDiffSample]


class WorkbookBusinessDiffSample(TypedDict):
    sheet: str
    key: str
    field: str
    exported: object
    original: object


class WorkbookBusinessDiffSheetSummary(TypedDict):
    missing_rows: int
    extra_rows: int
    missing_rows_soft_matched: int
    extra_rows_soft_matched: int
    category_counts: dict[str, int]
    field_counts: dict[str, int]
    missing_row_samples: list[str]
    extra_row_samples: list[str]


class WorkbookBusinessDiff(TypedDict):
    ok: bool
    missing_sheets: list[str]
    extra_sheets: list[str]
    missing_rows: int
    extra_rows: int
    differing_fields: int
    missing_fields: dict[str, list[str]]
    extra_fields: dict[str, list[str]]
    sheet_summaries: dict[str, WorkbookBusinessDiffSheetSummary]
    samples: list[WorkbookBusinessDiffSample]


type BusinessKey = tuple[object, ...]
type BusinessRow = dict[str, object]
type BusinessTable = dict[BusinessKey, BusinessRow]


def _exportable_confidence_sql(alias: str) -> str:
    return (
        f"({alias}.extraction_confidence IS NULL "
        f"OR {alias}.extraction_confidence >= {EXCEL_MIN_EXTRACTION_CONFIDENCE})"
    )


def _low_confidence_reason() -> str:
    return f"confidence<{EXCEL_MIN_EXTRACTION_CONFIDENCE:.2f}"


def export_quality_warnings(session: Session) -> dict[str, int]:
    """Count current rows that need Excel export quality attention."""
    params = {
        "min_confidence": EXCEL_MIN_EXTRACTION_CONFIDENCE,
        "auto_flag_confidence": EXCEL_AUTO_FLAG_EXTRACTION_CONFIDENCE,
    }
    checks = {
        "department_yearly_low_confidence_current": text(f"""
            SELECT COUNT(*)
            FROM department_yearly dy
            WHERE dy.is_current = {IS_CURRENT_TRUE_SQL}
              AND dy.extraction_confidence IS NOT NULL
              AND dy.extraction_confidence < :min_confidence
        """),
        "department_yearly_auto_flag_current": text(f"""
            SELECT COUNT(*)
            FROM department_yearly dy
            WHERE dy.is_current = {IS_CURRENT_TRUE_SQL}
              AND dy.extraction_confidence IS NOT NULL
              AND dy.extraction_confidence >= :min_confidence
              AND dy.extraction_confidence < :auto_flag_confidence
        """),
        "support_recipient_low_confidence_current": text(f"""
            SELECT COUNT(*)
            FROM support_recipient sr
            WHERE sr.is_current = {IS_CURRENT_TRUE_SQL}
              AND sr.extraction_confidence IS NOT NULL
              AND sr.extraction_confidence < :min_confidence
        """),
        "support_recipient_auto_flag_current": text(f"""
            SELECT COUNT(*)
            FROM support_recipient sr
            WHERE sr.is_current = {IS_CURRENT_TRUE_SQL}
              AND sr.extraction_confidence IS NOT NULL
              AND sr.extraction_confidence >= :min_confidence
              AND sr.extraction_confidence < :auto_flag_confidence
        """),
    }
    return {
        key: int(session.execute(query, params).scalar() or 0)
        for key, query in checks.items()
    }


def _write_sairoku(ws: Worksheet, session: Session) -> int:
    """Sheet 1: 採録状況 -- school collection status per year.

    Returns the number of data rows written.
    """
    headers = ["都道府県", "法人名", "学校名"] + [f"{y}年度" for y in FISCAL_YEARS]
    ws.append(headers)

    # Generate CASE WHEN clauses dynamically from FISCAL_YEARS
    year_cols = ",\n            ".join(
        f"MAX(CASE WHEN sys.fiscal_year = {y} THEN COALESCE(sys.legacy_status, sys.status) END) AS y{y}"
        for y in FISCAL_YEARS
    )
    # Sprint 8.2.1: school_year_status is now append-only with revision support.
    # Filter the JOIN to is_current=true so Excel reflects only the latest
    # revision per (school, fiscal_year), never a stale 'partial' shadowing
    # a current 'collected'.
    query = text(f"""
        SELECT
            s.prefecture,
            s.corporation_name,
            s.school_name,
            {year_cols}
        FROM school s
        LEFT JOIN school_year_status sys
            ON sys.school_id = s.id
            AND sys.is_current = {IS_CURRENT_TRUE_SQL}
        GROUP BY s.id, s.prefecture, s.corporation_name, s.school_name
        ORDER BY s.id
    """)

    rows = session.execute(query).fetchall()
    for row in rows:
        ws.append(list(row))

    count = len(rows)
    log.info("sairoku_exported", rows=count)
    return count


def _write_taisho_hiritu(ws: Worksheet, session: Session) -> int:
    """Sheet 2: 対象比率 -- support recipient data.

    Returns the number of data rows written.
    """
    headers = [
        "番号", "年度", "学校番号", "都道府県", "法人名", "学校名",
        "前年在籍", "前半期", "第\u2160区分", "第\u2161区分", "第\u2162区分", "第\u2163区分",
        "後半期", "第\u2160区分", "第\u2161区分", "第\u2162区分", "第\u2163区分",
        "年間", "家計急変多子世帯", "総計", "備考", "受給比率",
    ]
    ws.append(headers)

    # Sprint 8.2.1: support_recipient is now append-only — filter to
    # is_current=true so the 対象比率 sheet shows exactly one row per
    # (school, fiscal_year), never an old + new pair side by side.
    query = text(f"""
        SELECT
            sr.id,
            sr.fiscal_year,
            sr.school_number,
            s.prefecture,
            s.corporation_name,
            s.school_name,
            sr.prev_enrollment,
            sr.first_half_total,
            sr.first_half_cat1,
            sr.first_half_cat2,
            sr.first_half_cat3,
            sr.first_half_cat4,
            sr.second_half_total,
            sr.second_half_cat1,
            sr.second_half_cat2,
            sr.second_half_cat3,
            sr.second_half_cat4,
            sr.annual_total,
            sr.household_change,
            sr.grand_total,
            sr.notes,
            sr.recipient_rate
        FROM support_recipient sr
        JOIN school s ON s.id = sr.school_id
        WHERE sr.is_current = {IS_CURRENT_TRUE_SQL}
          AND {_exportable_confidence_sql("sr")}
        ORDER BY sr.id
    """)

    rows = session.execute(query).fetchall()
    for row in rows:
        raw = list(row)
        # Format fiscal_year as "XXXX年度"
        raw[1] = f"{raw[1]}年度" if raw[1] else raw[1]
        ws.append(raw)

    count = len(rows)
    log.info("taisho_hiritu_exported", rows=count)
    return count


def _write_gakka(ws: Worksheet, session: Session) -> int:
    """Sheet 3: 学科別 -- department yearly data with multi-row header.

    Row 1: year group labels (merged spans in original, None-padded here)
    Row 2: field name headers
    Data rows start at row 3.

    Returns the number of data rows written.
    """
    # Row 1: year group header
    row1: ExcelRow = [None] * 7  # key columns have no year label
    for year in FISCAL_YEARS:
        label = f"{year}年度"
        if year == 2019:
            block_size = 10
        else:
            block_size = 11
        row1.append(label)
        row1.extend([None] * (block_size - 1))
    ws.append(row1)

    # Row 2: field name header
    row2 = ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限"]
    for year in FISCAL_YEARS:
        row2.extend(YEAR_BLOCK_HEADERS)
        if year >= 2020:
            row2.append("備考")
    ws.append(row2)

    # Data query: join department + department_yearly (pivoted)
    query = text("""
        SELECT
            s.prefecture,
            s.corporation_name,
            s.school_name,
            d.course_name,
            d.canonical_name,
            d.course_type,
            d.duration_years,
            d.id AS dept_id
        FROM department d
        JOIN school s ON s.id = d.school_id
        ORDER BY s.id, d.id
    """)

    depts = session.execute(query).fetchall()

    # Pre-fetch all yearly data keyed by (department_id, fiscal_year)
    yearly_query = text(f"""
        SELECT
            department_id, fiscal_year,
            capacity, enrollment, intl_students, graduates,
            advanced, employed, other, prev_enrollment,
            dropouts, dropout_rate, notes
        FROM department_yearly
        WHERE is_current = {IS_CURRENT_TRUE_SQL}
          AND {_exportable_confidence_sql("department_yearly")}
        ORDER BY department_id, fiscal_year
    """)
    yearly_rows = session.execute(yearly_query).fetchall()

    yearly_map: dict[tuple[int, int], YearlyData] = {}
    for yr in yearly_rows:
        yearly_map[(yr[0], yr[1])] = tuple(yr[2:])  # skip dept_id and fiscal_year

    count = 0
    for dept in depts:
        prefecture, corp, school, course, dept_name, day_night, duration, dept_id = dept
        row: ExcelRow = [prefecture, corp, school, course, dept_name, day_night, duration]

        for year in FISCAL_YEARS:
            yd = yearly_map.get((dept_id, year))
            if yd is not None:
                # yd = (capacity, enrollment, intl, graduates, advanced, employed,
                #        other, prev_enrollment, dropouts, dropout_rate, notes)
                row.extend(list(yd[:10]))  # 10 numeric fields
                if year >= 2020:
                    row.append(yd[10])  # notes
            else:
                if year == 2019:
                    row.extend([None] * 10)
                else:
                    row.extend([None] * 11)

        ws.append(row)
        count += 1

    log.info("gakka_exported", rows=count)
    return count


def _write_zaiseki(ws: Worksheet, session: Session) -> int:
    """Sheet 4: 在籍のみ抜粋 -- enrollment-only extract.

    Multi-row header:
      Row 1: group labels (在籍者数, 留学生数)
      Row 2: key + year column headers
    Data starts row 3. Uses years from 2019 through one year before the current fiscal year.

    Returns the number of data rows written.
    """
    # Row 1: group header (dynamic year count)
    n_years = len(ENROLLMENT_YEARS)
    row1: ExcelRow = [None] * 7  # key columns
    row1.append("在籍者数")
    row1.extend([None] * (n_years - 1))
    row1.append("留学生数")
    row1.extend([None] * (n_years - 1))
    ws.append(row1)

    # Row 2: field names
    row2 = ["都道府県", "法人名", "学校名", "課程名", "学科名", "昼夜", "年限"]
    for year in ENROLLMENT_YEARS:
        row2.append(f"{year}年度")
    for year in ENROLLMENT_YEARS:
        row2.append(f"{year}年度")
    ws.append(row2)

    # Data query: department joined with yearly enrollment/intl_students
    query = text("""
        SELECT
            s.prefecture,
            s.corporation_name,
            s.school_name,
            d.course_name,
            d.canonical_name,
            d.course_type,
            d.duration_years,
            d.id AS dept_id
        FROM department d
        JOIN school s ON s.id = d.school_id
        ORDER BY s.id, d.id
    """)
    depts = session.execute(query).fetchall()

    # Pre-fetch enrollment data for dynamic year range
    min_ey = min(ENROLLMENT_YEARS)
    max_ey = max(ENROLLMENT_YEARS)
    yearly_query = text(f"""
        SELECT department_id, fiscal_year, enrollment, intl_students
        FROM department_yearly
        WHERE is_current = {IS_CURRENT_TRUE_SQL}
          AND fiscal_year BETWEEN {min_ey} AND {max_ey}
          AND {_exportable_confidence_sql("department_yearly")}
        ORDER BY department_id, fiscal_year
    """)
    yearly_rows = session.execute(yearly_query).fetchall()

    enroll_map: dict[tuple[int, int], tuple[int | None, int | None]] = {}
    for yr in yearly_rows:
        enroll_map[(yr[0], yr[1])] = (yr[2], yr[3])

    count = 0
    for dept in depts:
        prefecture, corp, school, course, dept_name, day_night, duration, dept_id = dept
        row: ExcelRow = [prefecture, corp, school, course, dept_name, day_night, duration]

        # Enrollment values for the one-year-lag range.
        for year in ENROLLMENT_YEARS:
            data = enroll_map.get((dept_id, year))
            row.append(data[0] if data else None)

        # International student values for the one-year-lag range.
        for year in ENROLLMENT_YEARS:
            data = enroll_map.get((dept_id, year))
            row.append(data[1] if data else None)

        ws.append(row)
        count += 1

    log.info("zaiseki_exported", rows=count)
    return count


def _low_confidence_exclusion_rows(session: Session) -> list[ExcelRow]:
    params = {
        "min_confidence": EXCEL_MIN_EXTRACTION_CONFIDENCE,
        "low_confidence_reason": _low_confidence_reason(),
    }
    department_rows = session.execute(
        text(f"""
            SELECT
                'department_yearly' AS row_type,
                dy.id AS row_id,
                s.school_name,
                d.canonical_name AS department_name,
                dy.fiscal_year,
                dy.extraction_confidence,
                :low_confidence_reason AS reason,
                '学科別/在籍のみ抜粋' AS export_target
            FROM department_yearly dy
            JOIN department d ON d.id = dy.department_id
            JOIN school s ON s.id = d.school_id
            WHERE dy.is_current = {IS_CURRENT_TRUE_SQL}
              AND dy.extraction_confidence IS NOT NULL
              AND dy.extraction_confidence < :min_confidence
            ORDER BY dy.id
        """),
        params,
    ).fetchall()
    support_rows = session.execute(
        text(f"""
            SELECT
                'support_recipient' AS row_type,
                sr.id AS row_id,
                s.school_name,
                NULL AS department_name,
                sr.fiscal_year,
                sr.extraction_confidence,
                :low_confidence_reason AS reason,
                '対象比率' AS export_target
            FROM support_recipient sr
            JOIN school s ON s.id = sr.school_id
            WHERE sr.is_current = {IS_CURRENT_TRUE_SQL}
              AND sr.extraction_confidence IS NOT NULL
              AND sr.extraction_confidence < :min_confidence
            ORDER BY sr.id
        """),
        params,
    ).fetchall()
    rows: list[ExcelRow] = []
    for row in [*department_rows, *support_rows]:
        raw = list(row)
        if raw[5] is not None:
            raw[5] = float(raw[5])
        rows.append(raw)
    return rows


def _write_low_confidence_exclusions(ws: Worksheet, rows: list[ExcelRow]) -> int:
    ws.append(["種別", "行ID", "学校名", "学科名", "年度", "confidence", "理由", "転記先"])
    for row in rows:
        ws.append(row)
    return len(rows)


def export_master_workbook(session: Session, output_path: Path) -> dict[str, int]:
    """Generate the master Excel workbook from database.

    Args:
        session: SQLAlchemy session connected to the EIDP database.
        output_path: Where to write the .xlsx file.

    Returns:
        Dict mapping sheet name to number of data rows exported.
    """
    log.info("export_start", output=str(output_path))

    wb = openpyxl.Workbook()

    # Sheet 1: 採録状況
    ws_sairoku = wb.active
    ws_sairoku.title = "採録状況"
    sairoku_count = _write_sairoku(ws_sairoku, session)

    # Sheet 2: 対象比率
    ws_taisho = wb.create_sheet("対象比率")
    taisho_count = _write_taisho_hiritu(ws_taisho, session)

    # Sheet 3: 学科別
    ws_gakka = wb.create_sheet("学科別")
    gakka_count = _write_gakka(ws_gakka, session)

    # Sheet 4: 在籍のみ抜粋
    ws_zaiseki = wb.create_sheet("在籍のみ抜粋")
    zaiseki_count = _write_zaiseki(ws_zaiseki, session)
    quality_warnings = export_quality_warnings(session)
    low_confidence_exclusions = _low_confidence_exclusion_rows(session)
    low_confidence_exclusion_count = 0
    if low_confidence_exclusions:
        ws_exclusions = wb.create_sheet(LOW_CONFIDENCE_EXCLUSION_SHEET)
        low_confidence_exclusion_count = _write_low_confidence_exclusions(ws_exclusions, low_confidence_exclusions)
    if any(quality_warnings.values()):
        log.warning("excel_export_quality_warnings", **quality_warnings)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    results = {
        "採録状況": sairoku_count,
        "対象比率": taisho_count,
        "学科別": gakka_count,
        "在籍のみ抜粋": zaiseki_count,
        **({LOW_CONFIDENCE_EXCLUSION_SHEET: low_confidence_exclusion_count} if low_confidence_exclusion_count else {}),
        **{f"quality_{key}": value for key, value in quality_warnings.items()},
    }
    log.info("export_complete", results=results, output=str(output_path))
    return results


def diff_workbooks(exported_path: Path, original_path: Path) -> dict[str, dict[str, int]]:
    """Compare row counts between exported and original Excel files.

    Args:
        exported_path: Path to the exported workbook.
        original_path: Path to the original reference workbook.

    Returns:
        Dict mapping sheet name to {exported, original, diff}.
    """
    wb_exp = openpyxl.load_workbook(exported_path, read_only=True, data_only=True)
    wb_orig = openpyxl.load_workbook(original_path, read_only=True, data_only=True)

    results: dict[str, dict[str, int]] = {}

    all_sheets = set(wb_exp.sheetnames) | set(wb_orig.sheetnames)
    for name in sorted(all_sheets):
        exp_rows = 0
        orig_rows = 0

        if name in wb_exp.sheetnames:
            ws = wb_exp[name]
            exp_rows = ws.max_row or 0

        if name in wb_orig.sheetnames:
            ws = wb_orig[name]
            orig_rows = ws.max_row or 0

        results[name] = {
            "exported": exp_rows,
            "original": orig_rows,
            "diff": exp_rows - orig_rows,
        }

    wb_exp.close()
    wb_orig.close()

    return results


def _is_number(value: object) -> TypeGuard[int | float]:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _cell_values_equal(exported: object, original: object, *, numeric_tolerance: float) -> bool:
    if exported == original:
        return True
    if numeric_tolerance > 0 and _is_number(exported) and _is_number(original):
        return abs(float(exported) - float(original)) <= numeric_tolerance
    return False


def _excel_cell_ref(row: int, column: int) -> str:
    letters = ""
    current = column
    while current:
        current, remainder = divmod(current - 1, 26)
        letters = f"{chr(65 + remainder)}{letters}"
    return f"{letters}{row}"


def _iter_sheet_values(worksheet: Any, *, max_row: int, max_column: int) -> Iterator[tuple[object, ...]]:
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True):
        yield tuple(row)


def diff_workbook_values(
    exported_path: Path,
    original_path: Path,
    *,
    max_diffs: int = 20,
    numeric_tolerance: float = 0.0,
) -> WorkbookValueDiff:
    """Compare cell values between exported and original Excel workbooks."""
    if max_diffs < 0:
        raise ValueError("max_diffs must be non-negative")
    if numeric_tolerance < 0:
        raise ValueError("numeric_tolerance must be non-negative")

    wb_exp = openpyxl.load_workbook(exported_path, read_only=True, data_only=True)
    wb_orig = openpyxl.load_workbook(original_path, read_only=True, data_only=True)

    try:
        exp_sheets = set(wb_exp.sheetnames)
        orig_sheets = set(wb_orig.sheetnames)
        missing_sheets = sorted(orig_sheets - exp_sheets)
        extra_sheets = sorted(exp_sheets - orig_sheets)
        samples: list[WorkbookValueDiffSample] = []
        differing_cells = 0

        for name in sorted(exp_sheets & orig_sheets):
            ws_exp = wb_exp[name]
            ws_orig = wb_orig[name]
            max_row = max(ws_exp.max_row or 0, ws_orig.max_row or 0)
            max_column = max(ws_exp.max_column or 0, ws_orig.max_column or 0)
            exported_rows = _iter_sheet_values(ws_exp, max_row=max_row, max_column=max_column)
            original_rows = _iter_sheet_values(ws_orig, max_row=max_row, max_column=max_column)
            empty_row: tuple[object, ...] = ()
            for row_index, (exported_row, original_row) in enumerate(
                zip_longest(exported_rows, original_rows, fillvalue=empty_row),
                start=1,
            ):
                if exported_row == original_row:
                    continue
                for column_index in range(1, max_column + 1):
                    exported = exported_row[column_index - 1] if column_index <= len(exported_row) else None
                    original = original_row[column_index - 1] if column_index <= len(original_row) else None
                    if _cell_values_equal(exported, original, numeric_tolerance=numeric_tolerance):
                        continue
                    differing_cells += 1
                    if len(samples) < max_diffs:
                        samples.append(
                            {
                                "sheet": name,
                                "cell": _excel_cell_ref(row_index, column_index),
                                "exported": exported,
                                "original": original,
                            }
                        )

        return {
            "ok": not missing_sheets and not extra_sheets and differing_cells == 0,
            "missing_sheets": missing_sheets,
            "extra_sheets": extra_sheets,
            "differing_cells": differing_cells,
            "samples": samples,
        }
    finally:
        wb_exp.close()
        wb_orig.close()


_BUSINESS_DIFF_SHEETS = ("対象比率", "在籍のみ抜粋")


def _stringify_key(key: BusinessKey) -> str:
    return " | ".join("" if part is None else str(part) for part in key)


def _business_soft_key_part(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"[\s\u3000]+", "", text)
    for marker in ("（専）", "専門学校", "専修学校"):
        text = text.replace(marker, "")
    return text


def _business_soft_key(sheet: str, key: BusinessKey) -> tuple[str, ...]:
    if sheet == "対象比率":
        year, prefecture, _corporation, school = key
        return tuple(_business_soft_key_part(part) for part in (year, prefecture, school))
    if sheet == "在籍のみ抜粋":
        prefecture, _corporation, school, course, department, daynight, years = key
        parts = (prefecture, school, course, department, daynight, years)
        return tuple(_business_soft_key_part(part) for part in parts)
    return tuple(_business_soft_key_part(part) for part in key)


def _normalize_business_value(value: object) -> object:
    return "" if value is None else value


def _business_diff_category(exported: object, original: object) -> str:
    non_numeric_legacy_markers = {"#DIV/0!", "#VALUE!", "#N/A", "不明"}
    if exported == "" and original in non_numeric_legacy_markers:
        return "export_blank_vs_original_error_or_unknown"
    if original == "" and exported in non_numeric_legacy_markers:
        return "export_error_or_unknown_vs_original_blank"
    if exported == "" and original != "":
        return "export_blank_vs_original_value"
    if original == "" and exported != "":
        return "export_value_vs_original_blank"
    if _is_number(exported) and _is_number(original):
        return "numeric_mismatch"
    if isinstance(exported, str) and isinstance(original, str):
        return "text_mismatch"
    return f"type_mismatch:{type(exported).__name__}->{type(original).__name__}"


def _load_taisho_business_table(worksheet: Any) -> tuple[BusinessTable, list[str]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}, []
    metrics = [
        "前年在籍",
        "前半期",
        "前半期_第Ⅰ区分",
        "前半期_第Ⅱ区分",
        "前半期_第Ⅲ区分",
        "前半期_第Ⅳ区分",
        "後半期",
        "後半期_第Ⅰ区分",
        "後半期_第Ⅱ区分",
        "後半期_第Ⅲ区分",
        "後半期_第Ⅳ区分",
        "年間",
        "家計急変多子世帯",
        "総計",
        "備考",
        "受給比率",
    ]
    metric_columns = dict(zip(metrics, range(6, 22), strict=False))
    table: BusinessTable = {}
    for raw in rows[1:]:
        row = tuple(raw)
        if not any(value is not None for value in row):
            continue
        key = tuple(row[index] if index < len(row) else None for index in (1, 3, 4, 5))
        table[key] = {
            metric: _normalize_business_value(row[index] if index < len(row) else None)
            for metric, index in metric_columns.items()
        }
    return table, list(metric_columns)


def _forward_fill(values: tuple[object, ...]) -> list[object]:
    filled: list[object] = []
    current: object = None
    for value in values:
        if value is not None:
            current = value
        filled.append(current)
    return filled


def _load_zaiseki_business_table(worksheet: Any) -> tuple[BusinessTable, list[str]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}, []
    group_row = _forward_fill(tuple(rows[0]))
    year_row = tuple(rows[1])
    field_columns: list[tuple[str, int]] = []
    for index in range(7, max(len(group_row), len(year_row))):
        group = group_row[index] if index < len(group_row) else None
        year = year_row[index] if index < len(year_row) else None
        if group in {"在籍者数", "留学生数"} and year is not None:
            field_columns.append((f"{group}:{year}", index))

    table: BusinessTable = {}
    for raw in rows[2:]:
        row = tuple(raw)
        if not any(value is not None for value in row):
            continue
        key = tuple(row[index] if index < len(row) else None for index in range(7))
        table[key] = {
            field: _normalize_business_value(row[index] if index < len(row) else None)
            for field, index in field_columns
        }
    return table, [field for field, _index in field_columns]


def _load_business_table(worksheet: Any, sheet: str) -> tuple[BusinessTable, list[str]]:
    if sheet == "対象比率":
        return _load_taisho_business_table(worksheet)
    if sheet == "在籍のみ抜粋":
        return _load_zaiseki_business_table(worksheet)
    raise ValueError(f"unsupported business diff sheet: {sheet}")


def diff_workbook_business_values(
    exported_path: Path,
    original_path: Path,
    *,
    sheets: list[str] | None = None,
    max_diffs: int = 20,
    numeric_tolerance: float = 0.0,
) -> WorkbookBusinessDiff:
    """Compare Excel values after aligning rows by business keys."""
    if max_diffs < 0:
        raise ValueError("max_diffs must be non-negative")
    if numeric_tolerance < 0:
        raise ValueError("numeric_tolerance must be non-negative")

    target_sheets = list(_BUSINESS_DIFF_SHEETS if sheets is None else sheets)
    unsupported = sorted(set(target_sheets) - set(_BUSINESS_DIFF_SHEETS))
    if unsupported:
        raise ValueError(f"unsupported business diff sheets: {', '.join(unsupported)}")

    wb_exp = openpyxl.load_workbook(exported_path, read_only=True, data_only=True)
    wb_orig = openpyxl.load_workbook(original_path, read_only=True, data_only=True)
    try:
        exp_sheets = set(wb_exp.sheetnames)
        orig_sheets = set(wb_orig.sheetnames)
        missing_sheets = [sheet for sheet in target_sheets if sheet in orig_sheets and sheet not in exp_sheets]
        extra_sheets = [sheet for sheet in target_sheets if sheet in exp_sheets and sheet not in orig_sheets]
        missing_fields: dict[str, list[str]] = {}
        extra_fields: dict[str, list[str]] = {}
        missing_rows = 0
        extra_rows = 0
        differing_fields = 0
        sheet_summaries: dict[str, WorkbookBusinessDiffSheetSummary] = {}
        samples: list[WorkbookBusinessDiffSample] = []

        for sheet in target_sheets:
            if sheet not in exp_sheets or sheet not in orig_sheets:
                continue
            exp_table, exp_fields = _load_business_table(wb_exp[sheet], sheet)
            orig_table, orig_fields = _load_business_table(wb_orig[sheet], sheet)
            exp_field_set = set(exp_fields)
            orig_field_set = set(orig_fields)
            missing_fields[sheet] = sorted(orig_field_set - exp_field_set)
            extra_fields[sheet] = sorted(exp_field_set - orig_field_set)
            common_fields = [field for field in exp_fields if field in orig_field_set]

            exp_keys = set(exp_table)
            orig_keys = set(orig_table)
            missing_row_keys = sorted(orig_keys - exp_keys, key=_stringify_key)
            extra_row_keys = sorted(exp_keys - orig_keys, key=_stringify_key)
            missing_rows += len(missing_row_keys)
            extra_rows += len(extra_row_keys)
            exp_soft_keys = {_business_soft_key(sheet, key) for key in exp_keys}
            orig_soft_keys = {_business_soft_key(sheet, key) for key in orig_keys}
            missing_rows_soft_matched = sum(
                1 for key in missing_row_keys if _business_soft_key(sheet, key) in exp_soft_keys
            )
            extra_rows_soft_matched = sum(
                1 for key in extra_row_keys if _business_soft_key(sheet, key) in orig_soft_keys
            )
            missing_row_samples = sorted(
                missing_row_keys,
                key=lambda key: (_business_soft_key(sheet, key) not in exp_soft_keys, _stringify_key(key)),
            )
            extra_row_samples = sorted(
                extra_row_keys,
                key=lambda key: (_business_soft_key(sheet, key) not in orig_soft_keys, _stringify_key(key)),
            )
            category_counts: Counter[str] = Counter()
            field_counts: Counter[str] = Counter()

            for key in sorted(exp_keys & orig_keys, key=_stringify_key):
                exp_row = exp_table[key]
                orig_row = orig_table[key]
                for field in common_fields:
                    exported = exp_row.get(field, "")
                    original = orig_row.get(field, "")
                    if _cell_values_equal(exported, original, numeric_tolerance=numeric_tolerance):
                        continue
                    differing_fields += 1
                    category_counts[_business_diff_category(exported, original)] += 1
                    field_counts[field] += 1
                    if len(samples) < max_diffs:
                        samples.append(
                            {
                                "sheet": sheet,
                                "key": _stringify_key(key),
                                "field": field,
                                "exported": exported,
                                "original": original,
                            }
                        )
            sheet_summaries[sheet] = {
                "missing_rows": len(missing_row_keys),
                "extra_rows": len(extra_row_keys),
                "missing_rows_soft_matched": missing_rows_soft_matched,
                "extra_rows_soft_matched": extra_rows_soft_matched,
                "category_counts": dict(category_counts.most_common()),
                "field_counts": dict(field_counts.most_common()),
                "missing_row_samples": [_stringify_key(key) for key in missing_row_samples[:max_diffs]],
                "extra_row_samples": [_stringify_key(key) for key in extra_row_samples[:max_diffs]],
            }

        return {
            "ok": (
                not missing_sheets
                and not extra_sheets
                and missing_rows == 0
                and extra_rows == 0
                and differing_fields == 0
                and all(not fields for fields in missing_fields.values())
                and all(not fields for fields in extra_fields.values())
            ),
            "missing_sheets": missing_sheets,
            "extra_sheets": extra_sheets,
            "missing_rows": missing_rows,
            "extra_rows": extra_rows,
            "differing_fields": differing_fields,
            "missing_fields": missing_fields,
            "extra_fields": extra_fields,
            "sheet_summaries": sheet_summaries,
            "samples": samples,
        }
    finally:
        wb_exp.close()
        wb_orig.close()
