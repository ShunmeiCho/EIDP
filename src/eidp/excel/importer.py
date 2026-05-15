"""Excel importer — reads 4 sheets from master Excel into DB.

Sheets: 採録状況, 対象比率, 学科別, 在籍のみ抜粋 (snapshot, import skipped — re-derivable)
"""

import unicodedata
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
import structlog
from sqlalchemy.orm import Session

from eidp.db.models import Department, DepartmentYearly, School, SchoolAlias, SchoolYearStatus, SupportRecipient
from eidp.department_normalization import normalize_course_name
from eidp.fiscal_year import current_fiscal_year, fiscal_year_from_japanese_era_text

log = structlog.get_logger()

ImportStats = dict[str, int | str]


def _norm(s: str) -> str:
    """NFKC normalize and strip ALL whitespace for consistent matching.

    Aligned with school_matcher._normalize() to prevent whitespace-based
    mismatches from creating phantom School rows.
    """
    if not s:
        return ""
    import re
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", "", s)
    return s


class SchoolResolver:
    """Multi-level school lookup with auto-create for cross-sheet matching.

    Lookup cascade:
    1. Exact (prefecture, corporation_name, school_name) match
    2. NFKC-normalized exact match
    3. (prefecture, school_name) match — handles corporation name variations
    4. Auto-create new School record from sheet data
    """

    def __init__(self, session: Session) -> None:
        self._session = session
        self._exact: dict[tuple[str, str, str], int] = {}
        self._norm: dict[tuple[str, str, str], int] = {}
        self._pref_name: dict[tuple[str, str], int] = {}
        self._name_only: dict[str, list[int]] = {}  # name -> [school_ids]
        self._auto_created = 0

    def build(self) -> None:
        """Build all lookup indices from current School table."""
        schools = self._session.query(School).all()
        self._exact.clear()
        self._norm.clear()
        self._pref_name.clear()
        self._name_only.clear()
        for s in schools:
            key = (s.prefecture, s.corporation_name, s.school_name)
            self._exact[key] = s.id

            norm_key = (_norm(s.prefecture), _norm(s.corporation_name), _norm(s.school_name))
            self._norm[norm_key] = s.id

            pn_key = (_norm(s.prefecture), _norm(s.school_name))
            if pn_key not in self._pref_name:
                self._pref_name[pn_key] = s.id

            norm_name = _norm(s.school_name)
            self._name_only.setdefault(norm_name, []).append(s.id)

        log.info("school_resolver_built", exact=len(self._exact),
                 norm=len(self._norm), pref_name=len(self._pref_name))

    def resolve(
        self,
        prefecture: str,
        corporation_name: str,
        school_name: str,
        *,
        reconcile_prefecture: bool = False,
    ) -> int | None:
        """Resolve a school to its DB id using cascading lookup.

        Cascade: exact -> NFKC normalized -> (pref+name) -> name-only (unique) -> auto-create.
        Levels 3-4 record a SchoolAlias so the mapping is visible and auditable.
        Level 5 (auto-create) only fires when the school name is completely new.
        """
        if not school_name or len(school_name.strip()) < 2:
            return None

        # Level 1: exact match
        key = (prefecture, corporation_name, school_name)
        sid = self._exact.get(key)
        if sid is not None:
            return sid

        # Level 2: NFKC + whitespace-normalized match
        norm_key = (_norm(prefecture), _norm(corporation_name), _norm(school_name))
        sid = self._norm.get(norm_key)
        if sid is not None:
            self._exact[key] = sid
            return sid

        # Level 3: (prefecture, school_name) match — corp name differs between sheets
        pn_key = (_norm(prefecture), _norm(school_name))
        sid = self._pref_name.get(pn_key)
        if sid is not None:
            self._exact[key] = sid
            self._record_alias(sid, school_name, "pref_name_match")
            return sid

        # Level 4: name-only match — only when unique (1 school with this name)
        norm_name = _norm(school_name)
        candidates = self._name_only.get(norm_name, [])
        if len(candidates) == 1:
            sid = candidates[0]
            if reconcile_prefecture:
                self._reconcile_prefecture_for_unique_name_match(sid, prefecture, school_name)
            self._exact[key] = sid
            self._norm[norm_key] = sid
            self._pref_name[pn_key] = sid
            self._record_alias(sid, school_name, "name_only_match")
            return sid

        # Level 5: auto-create — school name is genuinely new
        school = School(
            prefecture=prefecture,
            corporation_name=corporation_name,
            school_name=school_name,
            school_type="専門学校",
        )
        self._session.add(school)
        self._session.flush()
        self._auto_created += 1

        # Update all indices
        self._exact[key] = school.id
        self._norm[norm_key] = school.id
        if pn_key not in self._pref_name:
            self._pref_name[pn_key] = school.id
        self._name_only.setdefault(norm_name, []).append(school.id)

        log.debug("school_auto_created", prefecture=prefecture,
                  corporation_name=corporation_name, school_name=school_name,
                  school_id=school.id)
        return school.id

    def _record_alias(self, school_id: int, alias_name: str, source: str) -> None:
        """Record a SchoolAlias for audit trail when fuzzy match is used."""
        existing = (
            self._session.query(SchoolAlias)
            .filter(SchoolAlias.school_id == school_id, SchoolAlias.alias_name == alias_name)
            .first()
        )
        if not existing:
            alias = SchoolAlias(
                school_id=school_id,
                alias_name=alias_name,
                alias_type="cross_sheet",
                source=source,
            )
            self._session.add(alias)

    def _reconcile_prefecture_for_unique_name_match(
        self,
        school_id: int,
        prefecture: str,
        school_name: str,
    ) -> None:
        """Repair master.xlsx cross-sheet prefecture drift for a unique school name."""
        if not prefecture:
            return

        school = self._session.get(School, school_id)
        if school is None or school.prefecture == prefecture:
            return

        old_prefecture = school.prefecture
        school.prefecture = prefecture
        log.info(
            "school_prefecture_reconciled",
            school_id=school_id,
            school_name=school_name,
            old_prefecture=old_prefecture,
            new_prefecture=prefecture,
            source="name_only_match",
        )

    @property
    def auto_created_count(self) -> int:
        return self._auto_created

# Year columns in 採録状況 — computed dynamically to match exporter
def _compute_sairoku_years() -> list[int]:
    from datetime import datetime
    now = datetime.now()
    current_fy = now.year if now.month >= 4 else now.year - 1
    return list(range(2019, current_fy + 1))

SAIROKU_YEARS = _compute_sairoku_years()

# Status mapping: Excel free-text -> DB status enum
LEGACY_TO_STATUS = {
    "〇": "collected",
    "〇（一部欠損）": "collected",
    "〇（一部昨年？）": "collected",
    "△": "collected",
    "△（不足）": "collected",
    "△（前年データ）": "stale",
    "△（前年）": "stale",
    "△（同一データ）": "stale",
    "対象外": "excluded",
    "学校なし": "excluded",
    "統合": "excluded",
    "統廃合": "excluded",
    "閉校": "excluded",
    "募集停止": "excluded",
    "リンクミス": "error",
    "職実": "collected",
    "職実代用": "collected",
    "一部職実": "collected",
    "一部職実代用": "collected",
    "一部学科職実": "collected",
    "不足": "collected",
    "欠損データ": "error",
    "データなし": "error",
    "前年データ": "stale",
    "日付は変更されるが内容同じ": "stale",
    "情報公開": "collected",
    "事業報告": "collected",
    "新規申請": "pending",
    "不明": "pending",
    "他法人": "excluded",
}

EXCLUDED_REASONS = {"対象外", "学校なし", "統合", "統廃合", "閉校", "募集停止", "他法人"}

# 学科別 year block layout per field-spec
# 2019: 10 cols (no 備考), 2020-2025: 11 cols (with 備考)
YEAR_BLOCK_FIELDS_NO_BIKO = [
    "capacity", "enrollment", "intl_students", "graduates",
    "advanced", "employed", "other", "prev_enrollment",
    "dropouts", "dropout_rate",
]
YEAR_BLOCK_FIELDS_WITH_BIKO = YEAR_BLOCK_FIELDS_NO_BIKO + ["notes"]

GAKKA_KEY_COLS = 7  # 都道府県, 法人名, 学校名, 課程名, 学科名, 昼夜, 年限
DEPARTMENT_YEARLY_IMPORT_FIELDS = tuple(YEAR_BLOCK_FIELDS_WITH_BIKO)


def _safe_int(val: object) -> int | None:
    if val is None or val == "" or val == "-":
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _safe_float(val: object) -> float | None:
    if val is None or val == "" or val == "-":
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def _safe_str(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _same_department_yearly_import_values(
    row: DepartmentYearly,
    block_data: dict[str, int | float | str | None],
) -> bool:
    for field_name in DEPARTMENT_YEARLY_IMPORT_FIELDS:
        current = getattr(row, field_name)
        incoming = block_data.get(field_name)
        if current is None and incoming is None:
            continue
        if isinstance(current, float) or isinstance(incoming, float):
            if current is None or incoming is None or float(current) != float(incoming):
                return False
            continue
        if current != incoming:
            return False
    return True


def _department_yearly_from_block(
    *,
    department_id: int,
    fiscal_year: int,
    revision: int,
    block_data: dict[str, int | float | str | None],
) -> DepartmentYearly:
    return DepartmentYearly(
        department_id=department_id,
        fiscal_year=fiscal_year,
        revision=revision,
        is_current=True,
        capacity=block_data.get("capacity"),
        enrollment=block_data.get("enrollment"),
        intl_students=block_data.get("intl_students"),
        graduates=block_data.get("graduates"),
        advanced=block_data.get("advanced"),
        employed=block_data.get("employed"),
        other=block_data.get("other"),
        prev_enrollment=block_data.get("prev_enrollment"),
        dropouts=block_data.get("dropouts"),
        dropout_rate=block_data.get("dropout_rate"),
        notes=block_data.get("notes"),
        extraction_method="excel_import",
    )


def import_sairoku(ws: openpyxl.worksheet.worksheet.Worksheet, session: Session) -> dict[str, int]:
    """Import 採録状況 sheet -> school + school_year_status tables."""
    school_cache: dict[tuple[str, str, str], int] = {}
    stats = {"schools": 0, "statuses": 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        prefecture = _safe_str(row[0])
        corp_name = _safe_str(row[1])
        school_name = _safe_str(row[2])

        if not school_name:
            continue

        cache_key = (prefecture, corp_name, school_name)
        if cache_key in school_cache:
            school_id = school_cache[cache_key]
        else:
            # Upsert: find existing or create
            existing = (
                session.query(School)
                .filter(
                    School.prefecture == prefecture,
                    School.corporation_name == corp_name,
                    School.school_name == school_name,
                )
                .first()
            )
            if existing:
                school_id = existing.id
            else:
                school = School(
                    prefecture=prefecture,
                    corporation_name=corp_name,
                    school_name=school_name,
                    school_type="専門学校",
                )
                session.add(school)
                session.flush()
                school_id = school.id
                stats["schools"] += 1
            school_cache[cache_key] = school_id

        # Year status columns: cols 3-9 (0-indexed) for 2019-2025
        for i, year in enumerate(SAIROKU_YEARS):
            col_idx = 3 + i
            raw_val = _safe_str(row[col_idx]) if col_idx < len(row) else ""
            if not raw_val:
                status = "pending"
                legacy = None
            else:
                legacy = raw_val
                status = LEGACY_TO_STATUS.get(raw_val, "pending")

            excluded_reason = raw_val if raw_val in EXCLUDED_REASONS else None

            # Append-only upsert (Sprint 8.2.1). Earlier this branch did
            # in-place .first()-and-update which silently mutated old
            # revisions — incompatible with the new
            # UNIQUE(school_id, fiscal_year, revision) + partial unique on
            # is_current=true contract. Now: demote prior current row,
            # insert revision = max + 1 with is_current=True.
            existing_rows = (
                session.query(SchoolYearStatus)
                .filter(
                    SchoolYearStatus.school_id == school_id,
                    SchoolYearStatus.fiscal_year == year,
                )
                .all()
            )
            current_row = next((r for r in existing_rows if r.is_current), None)
            max_rev = max((r.revision for r in existing_rows), default=0)

            if current_row is not None:
                # Skip the write entirely if the current revision is already
                # equal to the values we'd insert — keeps idempotent re-runs
                # of the master Excel import from churning revisions.
                if (current_row.status == status
                        and current_row.legacy_status == legacy
                        and current_row.excluded_reason == excluded_reason):
                    stats["statuses"] += 1
                    continue
                session.query(SchoolYearStatus).filter(
                    SchoolYearStatus.school_id == school_id,
                    SchoolYearStatus.fiscal_year == year,
                    SchoolYearStatus.is_current == True,  # noqa: E712
                ).update({"is_current": False}, synchronize_session="fetch")

            session.add(
                SchoolYearStatus(
                    school_id=school_id,
                    fiscal_year=year,
                    status=status,
                    legacy_status=legacy,
                    excluded_reason=excluded_reason,
                    revision=max_rev + 1,
                    is_current=True,
                )
            )
            stats["statuses"] += 1

    session.flush()
    log.info("sairoku_imported", **stats)
    return stats


def import_gakka(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    session: Session,
    resolver: SchoolResolver,
) -> dict[str, int]:
    """Import 学科別 sheet -> department + department_yearly tables.

    Multi-row header: row 1 = year groups, row 2 = field names. Data starts row 3.
    """
    stats = {
        "departments": 0,
        "yearly_rows": 0,
        "school_misses": 0,
        "yearly_dupes": 0,
        "yearly_skipped_non_excel_current": 0,
        "auto_created": 0,
    }
    dept_cache: dict[tuple[int, str, str, str, float | None], int] = {}
    yearly_seen: set[tuple[int, int]] = set()  # (department_id, fiscal_year)

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        prefecture = _safe_str(row[0])
        corp_name = _safe_str(row[1])
        school_name = _safe_str(row[2])
        course_name = normalize_course_name(_safe_str(row[3])) or ""  # 課程名
        dept_name = _safe_str(row[4])    # 学科名
        day_night = _safe_str(row[5])    # 昼夜
        duration_raw = _safe_float(row[6])   # 年限 (supports 1.5, 2.4 etc.)
        # Round to 1 decimal to match DB Numeric(3,1) precision
        duration = round(duration_raw, 1) if duration_raw is not None else None

        if not dept_name:
            continue

        school_id = resolver.resolve(prefecture, corp_name, school_name, reconcile_prefecture=True)
        if school_id is None:
            stats["school_misses"] += 1
            continue

        # Upsert department
        dept_key = (school_id, dept_name, day_night, course_name, duration)
        if dept_key in dept_cache:
            dept_id = dept_cache[dept_key]
        else:
            existing_dept = (
                session.query(Department)
                .filter(
                    Department.school_id == school_id,
                    Department.canonical_name == dept_name,
                    Department.course_type == (day_night if day_night else None),
                    Department.course_name == (course_name if course_name else None),
                    Department.duration_years == duration,
                )
                .first()
            )
            if existing_dept:
                dept_id = existing_dept.id
            else:
                dept = Department(
                    school_id=school_id,
                    course_name=course_name if course_name else None,
                    canonical_name=dept_name,
                    course_type=day_night if day_night else None,
                    duration_years=duration,
                )
                session.add(dept)
                session.flush()
                dept_id = dept.id
                stats["departments"] += 1
            dept_cache[dept_key] = dept_id

        # Parse year blocks
        col_offset = GAKKA_KEY_COLS  # start after key columns

        for year in SAIROKU_YEARS:
            if year == 2019:
                fields = YEAR_BLOCK_FIELDS_NO_BIKO
                block_size = 10
            else:
                fields = YEAR_BLOCK_FIELDS_WITH_BIKO
                block_size = 11

            block_data: dict[str, int | float | str | None] = {}
            for fi, field_name in enumerate(fields):
                ci = col_offset + fi
                raw = row[ci] if ci < len(row) else None
                if field_name == "dropout_rate":
                    block_data[field_name] = _safe_float(raw)
                elif field_name == "notes":
                    block_data[field_name] = _safe_str(raw) if raw else None
                else:
                    block_data[field_name] = _safe_int(raw)

            col_offset += block_size

            # Only insert if there's any data
            has_data = any(
                v is not None and v != ""
                for k, v in block_data.items()
                if k != "notes"
            )
            if not has_data:
                continue

            yearly_key = (dept_id, year)
            if yearly_key in yearly_seen:
                stats["yearly_dupes"] += 1
                continue
            yearly_seen.add(yearly_key)

            existing_rows = (
                session.query(DepartmentYearly)
                .filter(
                    DepartmentYearly.department_id == dept_id,
                    DepartmentYearly.fiscal_year == year,
                )
                .all()
            )
            existing_dy = next((r for r in existing_rows if r.is_current), None)
            max_rev = max((r.revision for r in existing_rows), default=0)
            if existing_dy:
                if _same_department_yearly_import_values(existing_dy, block_data):
                    stats["yearly_rows"] += 1
                    continue
                if existing_dy.extraction_method not in (None, "excel_import"):
                    stats["yearly_skipped_non_excel_current"] += 1
                    continue
                session.query(DepartmentYearly).filter(
                    DepartmentYearly.department_id == dept_id,
                    DepartmentYearly.fiscal_year == year,
                    DepartmentYearly.is_current == True,  # noqa: E712
                ).update({"is_current": False}, synchronize_session="fetch")

            session.add(
                _department_yearly_from_block(
                    department_id=dept_id,
                    fiscal_year=year,
                    revision=max_rev + 1,
                    block_data=block_data,
                )
            )
            stats["yearly_rows"] += 1

    session.flush()
    log.info("gakka_imported", **stats)
    return stats


def import_taisho_hiritu(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    session: Session,
    resolver: SchoolResolver,
) -> dict[str, int]:
    """Import 対象比率 sheet -> support_recipient table.

    Each Excel row = one DB row (period='full').
    Columns: 番号, 年度, 学校番号, 都道府県, 法人名, 学校名,
    前年在籍, 前半期, 第Ⅰ区分x4, 後半期, 第Ⅰ区分x4,
    年間, 家計急変多子世帯, 総計, 備考, 受給比率
    """
    stats = {"rows": 0, "school_misses": 0, "duplicates": 0, "auto_created": 0, "invalid_year": 0}
    seen: set[tuple[int, int]] = set()  # (school_id, fiscal_year)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        year_str = _safe_str(row[1])
        school_number = _safe_str(row[2])
        prefecture = _safe_str(row[3])
        corp_name = _safe_str(row[4])
        school_name = _safe_str(row[5])

        if not school_name or not year_str:
            continue

        fiscal_year = _parse_fiscal_year(year_str)
        if fiscal_year is None:
            stats["invalid_year"] += 1
            continue

        school_id = resolver.resolve(prefecture, corp_name, school_name)
        if school_id is None:
            stats["school_misses"] += 1
            continue

        dedup_key = (school_id, fiscal_year)
        if dedup_key in seen:
            stats["duplicates"] += 1
            continue
        seen.add(dedup_key)

        # Append-only upsert (Sprint 8.2.1). Same rationale as
        # SchoolYearStatus above — must not silently overwrite an old
        # revision now that revision/is_current are part of the schema.
        existing_rows = (
            session.query(SupportRecipient)
            .filter(SupportRecipient.school_id == school_id, SupportRecipient.fiscal_year == fiscal_year)
            .all()
        )
        current_row = next((r for r in existing_rows if r.is_current), None)
        max_rev = max((r.revision for r in existing_rows), default=0)

        # Build the new revision's field set up front so we can short-circuit
        # the equality check (Sprint 8.2.2). Mirrors import_sairoku's no-op
        # path: a re-import of identical 対象比率 must NOT churn revisions.
        new_fields = {
            "school_number": school_number if school_number else None,
            "prev_enrollment": _safe_int(row[6]),
            "first_half_total": _safe_int(row[7]),
            "first_half_cat1": _safe_int(row[8]),
            "first_half_cat2": _safe_int(row[9]),
            "first_half_cat3": _safe_int(row[10]),
            "first_half_cat4": _safe_int(row[11]),
            "second_half_total": _safe_int(row[12]),
            "second_half_cat1": _safe_int(row[13]) if len(row) > 13 else None,
            "second_half_cat2": _safe_int(row[14]) if len(row) > 14 else None,
            "second_half_cat3": _safe_int(row[15]) if len(row) > 15 else None,
            "second_half_cat4": _safe_int(row[16]) if len(row) > 16 else None,
            "annual_total": _safe_int(row[17]) if len(row) > 17 else None,
            "household_change": _safe_int(row[18]) if len(row) > 18 else None,
            "grand_total": _safe_int(row[19]) if len(row) > 19 else None,
            "recipient_rate": _safe_float(row[21]) if len(row) > 21 else None,
            "notes": _safe_str(row[20]) if len(row) > 20 and row[20] else None,
        }

        if current_row is not None:
            # Equality short-circuit: identical content → no new revision.
            if all(
                getattr(current_row, field) == value
                for field, value in new_fields.items()
            ):
                stats["rows"] += 1
                continue
            session.query(SupportRecipient).filter(
                SupportRecipient.school_id == school_id,
                SupportRecipient.fiscal_year == fiscal_year,
                SupportRecipient.is_current == True,  # noqa: E712
            ).update({"is_current": False}, synchronize_session="fetch")

        sr = SupportRecipient(
            school_id=school_id,
            fiscal_year=fiscal_year,
            revision=max_rev + 1,
            is_current=True,
            **new_fields,
        )
        session.add(sr)
        stats["rows"] += 1

    session.flush()
    log.info("taisho_hiritu_imported", **stats)
    return stats


def _import_fiscal_year_upper_bound(max_fiscal_year: int | None) -> int:
    return max_fiscal_year if max_fiscal_year is not None else current_fiscal_year() + 1


def _bounded_import_fiscal_year(year: int | None, *, max_fiscal_year: int | None = None) -> int | None:
    if year is None:
        return None
    if year > _import_fiscal_year_upper_bound(max_fiscal_year):
        return None
    return year


def _parse_fiscal_year(val: str, *, max_fiscal_year: int | None = None) -> int | None:
    """Parse fiscal year from various formats."""
    import re

    val = val.strip()

    # "2024年度" or just "2024"
    m = re.search(r"(20\d{2})", val)
    if m:
        return _bounded_import_fiscal_year(int(m.group(1)), max_fiscal_year=max_fiscal_year)

    fiscal_year = fiscal_year_from_japanese_era_text(val)
    if fiscal_year is not None:
        return _bounded_import_fiscal_year(fiscal_year, max_fiscal_year=max_fiscal_year)

    return None


def import_all(excel_path: Path, session: Session) -> dict[str, ImportStats]:
    """Import all 4 sheets from master Excel. Returns stats per sheet."""
    log.info("import_start", path=str(excel_path))

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    try:
        results: dict[str, ImportStats] = {}

        # Sheet 1: 採録状況 -> school + school_year_status
        ws_sairoku = wb["採録状況"]
        results["採録状況"] = dict(import_sairoku(ws_sairoku, session))

        # Build multi-level school resolver for cross-sheet matching
        resolver = SchoolResolver(session)
        resolver.build()

        # Sheet 2: 対象比率 -> support_recipient
        ws_taisho = wb["対象比率"]
        taisho_stats: ImportStats = dict(import_taisho_hiritu(ws_taisho, session, resolver))
        taisho_stats["auto_created"] = resolver.auto_created_count
        results["対象比率"] = taisho_stats

        # Rebuild resolver indices after sheet 2 auto-creates
        if resolver.auto_created_count > 0:
            resolver.build()

        pre_gakka_auto = resolver.auto_created_count

        # Sheet 3: 学科別 -> department + department_yearly
        ws_gakka = wb["学科別"]
        gakka_stats: ImportStats = dict(import_gakka(ws_gakka, session, resolver))
        gakka_stats["auto_created"] = resolver.auto_created_count - pre_gakka_auto
        results["学科別"] = gakka_stats

        # Sheet 4: 在籍のみ抜粋 — snapshot, skip import (re-derivable from department_yearly)
        results["在籍のみ抜粋"] = {"skipped": 1, "reason": "re-derivable from department_yearly"}

        log.info("import_complete", results=results,
                 total_auto_created=resolver.auto_created_count)
        return results
    finally:
        wb.close()
