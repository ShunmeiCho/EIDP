"""Streamlit page: Excel プレビュー (Sprint 8.4.c.3).

Operator-facing dry-run before the master workbook is downloaded /
distributed. Generates the same 4-sheet workbook ``export_master_workbook``
produces, but in memory (BytesIO) so nothing hits disk until the operator
clicks the download button.

Architecture
------------
Same shape as 8.4.c.1 / 8.4.c.2. Pure helpers under unit test:

  * ``build_preview_workbook(session)`` — in-memory openpyxl workbook
    builder that mirrors ``export_master_workbook`` but returns
    BytesIO + a counts dict. No filesystem writes.
  * ``format_sheet_preview(workbook, sheet_name, max_rows)`` — slices
    the first N rows of a sheet into a list-of-rows for the UI table.
  * ``count_unmatched_and_gap(session)`` — surfaces 採録状況 unmatched
    and 学科別 gap counts so the operator sees what's missing before
    downloading.

Lock contract: this page is read-only — Excel is generated from
already-committed data. No lock needed. The page surfaces a probe
banner for visibility only.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
from sqlalchemy.orm import Session

from eidp.config import settings
from eidp.db.locking import probe_lock
from eidp.excel.exporter import (
    _write_gakka,
    _write_sairoku,
    _write_taisho_hiritu,
    _write_zaiseki,
    excel_auto_flag_extraction_confidence,
    excel_min_extraction_confidence,
    export_quality_warnings,
)
from eidp.fiscal_year import format_fiscal_year_label
from eidp.reports.coverage import ExportGapReport, gap_report_for_export
from eidp.review.school_scope import OPERATOR_SCHOOL_SCOPE_LABEL, OPERATOR_SCHOOL_TYPE_SCOPE
from eidp.review.target_year_status import target_year_overview

# Sheets the master workbook carries, in display order.
SHEET_ORDER: tuple[str, ...] = ("採録状況", "対象比率", "学科別", "在籍のみ抜粋")


@dataclass
class PreviewWorkbook:
    """In-memory workbook + per-sheet row counts."""

    workbook: openpyxl.Workbook
    counts: dict[str, int] = field(default_factory=dict)
    quality_warnings: dict[str, int] = field(default_factory=dict)

    def to_bytes(self) -> bytes:
        """Serialize to bytes for st.download_button. Calling this does
        NOT close the workbook — caller controls lifecycle."""
        buf = io.BytesIO()
        self.workbook.save(buf)
        return buf.getvalue()

    def close(self) -> None:
        """Release any workbook resources held by openpyxl."""
        close = getattr(self.workbook, "close", None)
        if callable(close):
            close()


# ---------------------------------------------------------------------------
# In-memory workbook builder
# ---------------------------------------------------------------------------


def build_preview_workbook(session: Session) -> PreviewWorkbook:
    """Build the 4-sheet master workbook in memory.

    Mirrors ``eidp.excel.exporter.export_master_workbook`` but returns a
    PreviewWorkbook with counts attached. No filesystem writes — caller
    decides whether to materialize via ``to_bytes()``.
    """
    wb = openpyxl.Workbook()

    ws_sairoku = wb.active
    ws_sairoku.title = "採録状況"
    sairoku_count = _write_sairoku(ws_sairoku, session)

    ws_taisho = wb.create_sheet("対象比率")
    taisho_count = _write_taisho_hiritu(ws_taisho, session)

    ws_gakka = wb.create_sheet("学科別")
    gakka_count = _write_gakka(ws_gakka, session)

    ws_zaiseki = wb.create_sheet("在籍のみ抜粋")
    zaiseki_count = _write_zaiseki(ws_zaiseki, session)

    return PreviewWorkbook(
        workbook=wb,
        counts={
            "採録状況": sairoku_count,
            "対象比率": taisho_count,
            "学科別": gakka_count,
            "在籍のみ抜粋": zaiseki_count,
        },
        quality_warnings=export_quality_warnings(session),
    )


# ---------------------------------------------------------------------------
# Sheet preview slicing
# ---------------------------------------------------------------------------


def format_sheet_preview(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    *,
    max_rows: int = 30,
) -> list[list[Any]]:
    """Return the first N rows of ``sheet_name`` as a list-of-rows."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet {sheet_name!r} not in workbook ({workbook.sheetnames})")
    ws = workbook[sheet_name]
    out: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        out.append(list(row))
    return out


def format_sheet_preview_from_bytes(
    workbook_bytes: bytes,
    sheet_name: str,
    *,
    max_rows: int = 30,
) -> list[list[Any]]:
    """Load a preview workbook for one table render, then close it.

    Streamlit reruns keep ``st.session_state`` alive, so the UI stores only
    serialized XLSX bytes and reconstructs a read-only workbook for the small
    visible table slice.
    """
    workbook = openpyxl.load_workbook(
        io.BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        return format_sheet_preview(workbook, sheet_name, max_rows=max_rows)
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


# ---------------------------------------------------------------------------
# Coverage / gap counts
# ---------------------------------------------------------------------------


def count_unmatched_and_gap(
    session: Session,
    *,
    fiscal_year: int | None = None,
    school_type: str | None = "専門学校",
) -> ExportGapReport:
    """Return target-FY export readiness counters.

    Historical data can remain in the workbook, but readiness must be based on
    the configured target fiscal year. Otherwise a school with only old-year
    rows would make the preview look safe when the current-year task is still
    incomplete.
    """
    target_fiscal_year = fiscal_year if fiscal_year is not None else settings.target_fiscal_year
    return gap_report_for_export(
        session,
        fiscal_year=target_fiscal_year,
        school_type=school_type,
    )


# ---------------------------------------------------------------------------
# Streamlit render
# ---------------------------------------------------------------------------


def _low_confidence_message(row_count: int) -> str:
    return (
        f"confidence {excel_min_extraction_confidence():.2f} 未満の current 行が {row_count} 件あります。"
        "これらはExcel出力から除外されます。PDF確認または手入力で修正してください。"
    )


def _auto_flag_confidence_message(row_count: int) -> str:
    return (
        "confidence "
        f"{excel_min_extraction_confidence():.2f}以上"
        f"{excel_auto_flag_extraction_confidence():.2f}未満の要確認行が {row_count} 件あります。"
        "Excelには含まれますが、配布前にPDF確認画面で内容を確認してください。"
    )


def render(session: Session, *, lock_path: Path) -> None:  # pragma: no cover - thin streamlit shell
    """Top-level Streamlit render for the Excel プレビュー page."""
    import streamlit as st

    st.subheader("Excel プレビュー")
    status = probe_lock(lock_path)
    if status.held:
        st.info(
            f"週次処理中 (owner={status.owner})。"
            " このページは読み取り専用です。"
        )

    target_label = format_fiscal_year_label(settings.target_fiscal_year)
    target = target_year_overview(
        session,
        target_fiscal_year=settings.target_fiscal_year,
        school_type=OPERATOR_SCHOOL_TYPE_SCOPE,
    )
    st.caption(f"対象年度: {target_label} / 対象範囲: {OPERATOR_SCHOOL_SCOPE_LABEL}")
    target_cols = st.columns(5)
    target_cols[0].metric("対象年度PDFあり", target.current_target_schools)
    target_cols[1].metric("旧年度fallback", target.stale_target_documents)
    target_cols[2].metric("未採録校", target.missing_current_target_schools)
    target_cols[3].metric("来年度以降PDF", target.future_target_documents)
    target_cols[4].metric("要確認キュー", target.review_queue_documents)
    if target.current_target_documents == 0 and target.stale_target_documents > 0:
        st.warning(
            f"{target_label} のPDFが未採録です。旧年度fallbackはExcel成果として扱わず、"
            "先に週次再取得またはURL追加を行ってください。"
        )

    export_gap = count_unmatched_and_gap(
        session,
        fiscal_year=settings.target_fiscal_year,
        school_type=OPERATOR_SCHOOL_TYPE_SCOPE,
    )
    cols = st.columns(4)
    cols[0].metric("現在年度PDF採録率", f"{export_gap.target_pdf_rate:.0%}")
    cols[1].metric("抽出済み学校", export_gap.extracted_schools)
    cols[2].metric("Excel出力可", export_gap.excel_ready_schools)
    cols[3].metric("Excel対象行", export_gap.target_yearly_rows)
    gap_cols = st.columns(4)
    gap_cols[0].metric("URLなし", export_gap.no_url_schools)
    gap_cols[1].metric("旧年度fallback校", export_gap.stale_fallback_schools)
    gap_cols[2].metric("未採録校", export_gap.missing_target_pdf_schools)
    gap_cols[3].metric("対象校", export_gap.total_schools)

    if not export_gap.has_target_year_data:
        st.error(
            f"{target_label} の在籍者数など転記対象データが 0 件です。"
            "旧年度データを成果としてダウンロードしないでください。"
            "先に学校別タスクでURL取得、PDF確認、年度修正を進めてください。"
        )
    elif export_gap.excel_ready_schools < export_gap.total_schools:
        st.warning(
            f"{target_label} は未完了です。"
            f"Excel出力可 {export_gap.excel_ready_schools}/{export_gap.total_schools} 校の状態で出力します。"
        )
    quality_warnings = export_quality_warnings(session)
    low_confidence_rows = (
        quality_warnings["department_yearly_low_confidence_current"]
        + quality_warnings["support_recipient_low_confidence_current"]
    )
    auto_flag_rows = (
        quality_warnings["department_yearly_auto_flag_current"]
        + quality_warnings["support_recipient_auto_flag_current"]
    )
    if low_confidence_rows:
        st.error(_low_confidence_message(low_confidence_rows))
    if auto_flag_rows:
        st.warning(_auto_flag_confidence_message(auto_flag_rows))

    can_generate = export_gap.has_target_year_data
    if st.button("プレビュー workbook を生成", type="primary", disabled=status.held or not can_generate):
        with st.spinner("生成中..."):
            preview = build_preview_workbook(session)
            try:
                st.session_state["excel_preview_bytes"] = preview.to_bytes()
                st.session_state["excel_preview_counts"] = preview.counts
                st.session_state["excel_preview_gap"] = export_gap
                st.session_state["excel_preview_quality_warnings"] = preview.quality_warnings
                st.session_state.pop("excel_preview_workbook", None)
            finally:
                preview.close()
        st.rerun()

    if "excel_preview_bytes" in st.session_state:
        workbook_bytes = st.session_state["excel_preview_bytes"]
        counts_map = st.session_state["excel_preview_counts"]
        st.caption(
            "シート行数: " + " / ".join(
                f"{name}={counts_map.get(name, 0)}" for name in SHEET_ORDER
            )
        )
        sheet_name = st.selectbox("シート選択", options=list(SHEET_ORDER))
        rows = format_sheet_preview_from_bytes(workbook_bytes, sheet_name, max_rows=30)
        if rows:
            st.table(rows)
        else:
            st.caption("(空)")
        st.download_button(
            label="Excel ダウンロード",
            data=st.session_state["excel_preview_bytes"],
            file_name="eidp_master.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
