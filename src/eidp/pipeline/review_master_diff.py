"""Diff reviewed extraction rows against a read-only master expected subset.

This is a Linux/Web review hardening layer. It produces mismatch reports for
reviewed rows and never writes final Excel output.
"""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path

from eidp.pdf.master_ground_truth import fy_metric_columns, normalize_text
from eidp.pipeline.department_join import (
    COURSE_GRANULARITY_COLLISION_REASON,
    is_course_granularity_collision,
    join_key_label,
    make_join_key,
    values_equal,
)
from eidp.pipeline.extraction_review import ReviewStatus
from eidp.pipeline.review_report import ReviewedExtractionRow

__all__ = [
    "DIFF_REPORT_COLUMNS",
    "DiffResultRow",
    "MatchStatus",
    "MasterExpectedRow",
    "diff_report_csv",
    "diff_reviewed_against_master",
    "diff_summary",
    "load_master_expected_subset",
]


class MatchStatus(StrEnum):
    MATCH = "match"
    MISSING_IN_EXTRACTION = "missing_in_extraction"
    MISSING_IN_MASTER = "missing_in_master"
    VALUE_MISMATCH = "value_mismatch"
    NEEDS_REVIEW_NOT_COMPARABLE = "needs_review_not_comparable"
    EXCLUDED_NOT_COMPARABLE = "excluded_not_comparable"
    AMBIGUOUS_KEY = "ambiguous_key"


DIFF_REPORT_COLUMNS: tuple[str, ...] = (
    "key",
    "school_name",
    "school_id",
    "fiscal_year",
    "department_name",
    "field_category",
    "course_name",
    "metric",
    "extracted_value",
    "expected_value",
    "match_status",
    "mismatch_reason",
    "review_status",
    "original_value",
    "corrected_value",
    "confidence",
    "source_pdf",
    "page_no",
    "table_index",
    "row_index",
    "col_index",
    "raw_label",
    "raw_value",
    "canonical_metric",
    "review_note",
    "reviewed_by",
    "reviewed_at",
    "master_row_id",
    "operator_mapping_id",
)

_DiffKey = tuple[str, str, str, int, str]


@dataclass(frozen=True)
class MasterExpectedRow:
    school_name: str
    school_id: str | None
    fiscal_year: int
    department_name: str
    metric: str
    expected_value: int | float | str | None
    field_category: str | None = None
    course_name: str | None = None
    day_or_evening: str | None = None
    duration_years: str | None = None
    master_row_id: str | None = None
    operator_mapping_id: str | None = None
    source_sheet: str | None = None
    source_cell: str | None = None


@dataclass(frozen=True)
class DiffResultRow:
    key: str
    school_name: str
    school_id: str | None
    fiscal_year: int
    department_name: str
    field_category: str | None
    course_name: str | None
    metric: str
    extracted_value: int | float | str | None
    expected_value: int | float | str | None
    match_status: MatchStatus
    mismatch_reason: str
    review_status: ReviewStatus | None
    original_value: int | None
    corrected_value: int | None
    confidence: float | None
    source_pdf: str | None
    page_no: int | None
    table_index: int | None
    row_index: int | None
    col_index: int | None
    raw_label: str | None
    raw_value: str | None
    canonical_metric: str | None
    review_note: str | None
    reviewed_by: str | None
    reviewed_at: str | None
    master_row_id: str | None
    operator_mapping_id: str | None


def load_master_expected_subset(
    master_path: Path | str,
    *,
    corporation_name: str,
    school_name: str,
    fiscal_year: int,
    school_id: str | None = None,
    prefecture: str | None = None,
) -> list[MasterExpectedRow]:
    """Load one school/FY expected subset from data/master.xlsx read-only."""
    import openpyxl  # type: ignore[import-untyped]  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]  # noqa: PLC0415

    cap_col, enr_col, intl_col = fy_metric_columns(fiscal_year)
    target_corp = normalize_text(corporation_name)
    target_school = normalize_text(school_name)
    target_pref = normalize_text(prefecture) if prefecture else None
    metric_columns = (("capacity", cap_col), ("enrollment", enr_col), ("intl_students", intl_col))

    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    expected: list[MasterExpectedRow] = []
    try:
        worksheet = wb["学科別"]
        for row_number, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            if row is None or len(row) <= intl_col:
                continue
            if normalize_text(str(row[1] or "")) != target_corp:
                continue
            if normalize_text(str(row[2] or "")) != target_school:
                continue
            if target_pref and normalize_text(str(row[0] or "")) != target_pref:
                continue
            field_category = str(row[3] or "")
            department_name = str(row[4] or "")
            for metric, col in metric_columns:
                expected.append(
                    MasterExpectedRow(
                        school_name=school_name,
                        school_id=school_id,
                        fiscal_year=fiscal_year,
                        department_name=department_name,
                        field_category=field_category,
                        course_name=None,
                        day_or_evening=_optional_text(row[5] if len(row) > 5 else None),
                        duration_years=_optional_text(row[6] if len(row) > 6 else None),
                        metric=metric,
                        expected_value=_safe_int(row[col]),
                        master_row_id=f"学科別!{row_number}",
                        operator_mapping_id=None,
                        source_sheet="学科別",
                        source_cell=f"{get_column_letter(col + 1)}{row_number}",
                    )
                )
        return expected
    finally:
        wb.close()


def diff_reviewed_against_master(
    reviewed_rows: list[ReviewedExtractionRow],
    expected_rows: list[MasterExpectedRow],
) -> list[DiffResultRow]:
    reviewed_by_key: dict[_DiffKey, list[ReviewedExtractionRow]] = defaultdict(list)
    expected_by_key: dict[_DiffKey, list[MasterExpectedRow]] = defaultdict(list)
    for candidate_reviewed_row in reviewed_rows:
        if candidate_reviewed_row.metric and candidate_reviewed_row.department_name:
            reviewed_by_key[_reviewed_key(candidate_reviewed_row)].append(candidate_reviewed_row)
    for expected_row in expected_rows:
        expected_by_key[_expected_key(expected_row)].append(expected_row)

    results: list[DiffResultRow] = []

    for key in sorted(set(reviewed_by_key) | set(expected_by_key), key=repr):
        reviewed_group = reviewed_by_key.get(key, [])
        expected_group = expected_by_key.get(key, [])
        if len(reviewed_group) > 1 or len(expected_group) > 1:
            results.extend(_ambiguous_rows(key, reviewed_group, expected_group))
            continue
        reviewed_row = reviewed_group[0] if reviewed_group else None
        expected = expected_group[0] if expected_group else None
        if reviewed_row is None and expected is not None:
            results.append(_missing_extraction_row(expected))
            continue
        if reviewed_row is None:
            continue
        if reviewed_row.review_status == ReviewStatus.EXCLUDED:
            results.append(
                _diff_row(reviewed_row, expected, MatchStatus.EXCLUDED_NOT_COMPARABLE, "review_status=excluded")
            )
            continue
        if reviewed_row.final_review_value is None:
            results.append(
                _diff_row(
                    reviewed_row,
                    expected,
                    MatchStatus.NEEDS_REVIEW_NOT_COMPARABLE,
                    "no final reviewed value",
                )
            )
            continue
        if expected is None:
            results.append(
                _diff_row(reviewed_row, None, MatchStatus.MISSING_IN_MASTER, "stable key absent from master subset")
            )
            continue
        if is_course_granularity_collision(reviewed_row.department_name, expected.department_name):
            # The loose department_key joined these two rows, but the strict key (which never
            # collapses a bare trailing コース) says they are DISTINCT granularities. Certifying a
            # match here would false-merge a course track with its parent 科, so refuse it and
            # surface the pair for human disambiguation instead.
            results.append(
                _diff_row(
                    reviewed_row,
                    expected,
                    MatchStatus.AMBIGUOUS_KEY,
                    COURSE_GRANULARITY_COLLISION_REASON,
                )
            )
            continue
        if values_equal(reviewed_row.final_review_value, expected.expected_value):
            results.append(_diff_row(reviewed_row, expected, MatchStatus.MATCH, "reviewed value matches master"))
            continue
        results.append(
            _diff_row(reviewed_row, expected, MatchStatus.VALUE_MISMATCH, "reviewed value differs from master")
        )

    return results


def diff_summary(rows: list[DiffResultRow]) -> dict[str, int]:
    counts = {status.value: 0 for status in MatchStatus}
    for row in rows:
        counts[row.match_status.value] += 1
    return counts


def diff_report_csv(rows: list[DiffResultRow]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=DIFF_REPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(_diff_report_row(row) for row in rows)
    return output.getvalue()


def _reviewed_key(row: ReviewedExtractionRow) -> _DiffKey:
    return make_join_key(
        row.school_name,
        row.field_category,
        row.department_name or "",
        row.fiscal_year,
        row.metric or "",
    )


def _expected_key(row: MasterExpectedRow) -> _DiffKey:
    return make_join_key(row.school_name, row.field_category, row.department_name, row.fiscal_year, row.metric)


def _diff_row(
    row: ReviewedExtractionRow,
    expected: MasterExpectedRow | None,
    status: MatchStatus,
    reason: str,
) -> DiffResultRow:
    key = _reviewed_key(row)
    return DiffResultRow(
        key=join_key_label(key),
        school_name=row.school_name,
        school_id=row.school_id,
        fiscal_year=row.fiscal_year,
        department_name=row.department_name or "",
        field_category=row.field_category,
        course_name=row.course_name,
        metric=row.metric or "",
        extracted_value=row.final_review_value,
        expected_value=expected.expected_value if expected is not None else None,
        match_status=status,
        mismatch_reason=reason,
        review_status=row.review_status,
        original_value=row.original_value,
        corrected_value=row.corrected_value,
        confidence=row.confidence,
        source_pdf=row.source_pdf,
        page_no=row.page_no,
        table_index=row.table_index,
        row_index=row.row_index,
        col_index=row.col_index,
        raw_label=row.raw_label,
        raw_value=row.raw_value,
        canonical_metric=row.canonical_metric,
        review_note=row.review_note,
        reviewed_by=row.reviewed_by,
        reviewed_at=row.reviewed_at,
        master_row_id=expected.master_row_id if expected is not None else None,
        operator_mapping_id=expected.operator_mapping_id if expected is not None else None,
    )


def _missing_extraction_row(
    expected: MasterExpectedRow,
    *,
    status: MatchStatus = MatchStatus.MISSING_IN_EXTRACTION,
    reason: str = "stable key absent from reviewed extraction rows",
) -> DiffResultRow:
    key = _expected_key(expected)
    return DiffResultRow(
        key=join_key_label(key),
        school_name=expected.school_name,
        school_id=expected.school_id,
        fiscal_year=expected.fiscal_year,
        department_name=expected.department_name,
        field_category=expected.field_category,
        course_name=expected.course_name,
        metric=expected.metric,
        extracted_value=None,
        expected_value=expected.expected_value,
        match_status=status,
        mismatch_reason=reason,
        review_status=None,
        original_value=None,
        corrected_value=None,
        confidence=None,
        source_pdf=None,
        page_no=None,
        table_index=None,
        row_index=None,
        col_index=None,
        raw_label=None,
        raw_value=None,
        canonical_metric=None,
        review_note=None,
        reviewed_by=None,
        reviewed_at=None,
        master_row_id=expected.master_row_id,
        operator_mapping_id=expected.operator_mapping_id,
    )


def _diff_report_row(row: DiffResultRow) -> dict[str, object]:
    return {
        "key": row.key,
        "school_name": row.school_name,
        "school_id": row.school_id or "",
        "fiscal_year": row.fiscal_year,
        "department_name": row.department_name,
        "field_category": row.field_category or "",
        "course_name": row.course_name or "",
        "metric": row.metric,
        "extracted_value": row.extracted_value if row.extracted_value is not None else "",
        "expected_value": row.expected_value if row.expected_value is not None else "",
        "match_status": row.match_status.value,
        "mismatch_reason": row.mismatch_reason,
        "review_status": row.review_status.value if row.review_status is not None else "",
        "original_value": row.original_value if row.original_value is not None else "",
        "corrected_value": row.corrected_value if row.corrected_value is not None else "",
        "confidence": row.confidence if row.confidence is not None else "",
        "source_pdf": row.source_pdf or "",
        "page_no": row.page_no if row.page_no is not None else "",
        "table_index": row.table_index if row.table_index is not None else "",
        "row_index": row.row_index if row.row_index is not None else "",
        "col_index": row.col_index if row.col_index is not None else "",
        "raw_label": row.raw_label or "",
        "raw_value": row.raw_value or "",
        "canonical_metric": row.canonical_metric or "",
        "review_note": row.review_note or "",
        "reviewed_by": row.reviewed_by or "",
        "reviewed_at": row.reviewed_at or "",
        "master_row_id": row.master_row_id or "",
        "operator_mapping_id": row.operator_mapping_id or "",
    }


def _ambiguous_rows(
    key: _DiffKey,
    reviewed_rows: list[ReviewedExtractionRow],
    expected_rows: list[MasterExpectedRow],
) -> list[DiffResultRow]:
    reason_parts: list[str] = []
    if len(reviewed_rows) > 1:
        reason_parts.append(f"duplicate reviewed rows={len(reviewed_rows)}")
    if len(expected_rows) > 1:
        reason_parts.append(f"duplicate master rows={len(expected_rows)}")
    reason = "; ".join(reason_parts)
    if reviewed_rows:
        expected = expected_rows[0] if len(expected_rows) == 1 else None
        return [_diff_row(row, expected, MatchStatus.AMBIGUOUS_KEY, reason) for row in reviewed_rows]
    return [
        _missing_extraction_row(expected, status=MatchStatus.AMBIGUOUS_KEY, reason=reason)
        for expected in expected_rows
    ]


def _safe_int(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "‐", "―"}:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
