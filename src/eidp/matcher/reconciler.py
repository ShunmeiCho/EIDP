"""School identity reconciliation — Step 4.

Cross-references DB schools against MEXT target institution list,
resolves unmatched/conflicting identities, produces verification report.
"""

import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import structlog
from sqlalchemy import func
from sqlalchemy.orm import Session

from eidp.db.models import School, SchoolAlias

log = structlog.get_logger()


@dataclass(frozen=True)
class TargetInstitution:
    school_code: str
    category: str  # 国立/公立/私立
    school_type: str  # 大学/専門学校/etc.
    name: str
    prefecture: str
    setter_name: str  # 設置者名称


@dataclass
class ReconcileCandidate:
    school_id: int
    school_name: str
    prefecture: str
    corporation_name: str
    candidate_code: str | None = None
    candidate_name: str | None = None
    match_method: str | None = None
    confidence: float = 0.0
    resolution: str | None = None  # assigned, excluded, duplicate, missing


@dataclass
class ReconcileReport:
    already_resolved: int = 0
    auto_assigned: list[ReconcileCandidate] = field(default_factory=list)
    needs_manual: list[ReconcileCandidate] = field(default_factory=list)
    excluded: list[ReconcileCandidate] = field(default_factory=list)
    missing_from_db: list[TargetInstitution] = field(default_factory=list)


def _norm(name: str) -> str:
    if not name:
        return ""
    name = unicodedata.normalize("NFKC", name)
    return name.replace(" ", "").replace("\u3000", "")


def load_target_institutions(path: Path) -> list[TargetInstitution]:
    """Load 専門学校 entries from MEXT target institution list."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    results: list[TargetInstitution] = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        code = str(row[0]).strip() if row[0] else ""
        school_type = str(row[2]).strip() if row[2] else ""

        if school_type != "専門学校":
            continue

        pref_raw = str(row[6]).strip() if row[6] else ""
        setter_name = str(row[8]).strip() if row[8] else ""

        results.append(
            TargetInstitution(
                school_code=code,
                category=str(row[1]).strip() if row[1] else "",
                school_type=school_type,
                name=str(row[3]).strip() if row[3] else "",
                prefecture=pref_raw,
                setter_name=setter_name,
            )
        )

    wb.close()
    log.info("target_institutions_loaded", count=len(results))
    return results


def reconcile(session: Session, data_dir: Path) -> ReconcileReport:
    """Reconcile unmatched schools against target institution list."""
    targets = load_target_institutions(data_dir / "target_institutions.xlsx")
    report = ReconcileReport()

    # Build target lookup by code and by normalized name+prefecture
    target_by_code: dict[str, TargetInstitution] = {t.school_code: t for t in targets}
    target_by_norm_pref: dict[tuple[str, str], list[TargetInstitution]] = defaultdict(list)
    for t in targets:
        target_by_norm_pref[(_norm(t.name), t.prefecture)].append(t)

    # Get all schools without school_code
    unresolved = session.query(School).filter(School.school_code.is_(None)).all()
    report.already_resolved = session.query(func.count(School.id)).filter(School.school_code.isnot(None)).scalar() or 0

    log.info("reconcile_start", unresolved=len(unresolved), targets=len(targets))

    # Also check aliases for existing matches
    existing_aliases = session.query(SchoolAlias).all()
    alias_by_school: dict[int, list[str]] = defaultdict(list)
    for a in existing_aliases:
        alias_by_school[a.school_id].append(a.alias_name)

    # Already-used codes
    used_codes: set[str] = set()
    for s in session.query(School).filter(School.school_code.isnot(None)).all():
        if s.school_code:
            used_codes.add(s.school_code)

    for school in unresolved:
        candidate = ReconcileCandidate(
            school_id=school.id,
            school_name=school.school_name,
            prefecture=school.prefecture,
            corporation_name=school.corporation_name,
        )

        # Strategy 1: Match by normalized name + prefecture against target list
        norm_name = _norm(school.school_name)
        matches = target_by_norm_pref.get((norm_name, school.prefecture), [])
        if matches:
            t = matches[0]
            if t.school_code not in used_codes:
                candidate.candidate_code = t.school_code
                candidate.candidate_name = t.name
                candidate.match_method = "target_exact"
                candidate.confidence = 1.0
                candidate.resolution = "assigned"
                report.auto_assigned.append(candidate)
                used_codes.add(t.school_code)
                continue

        # Strategy 2: Fuzzy matches go to needs_manual (not auto-applied per design)
        corp_norm = _norm(school.corporation_name)
        pref_targets = [t for t in targets if t.prefecture == school.prefecture and t.school_code not in used_codes]
        best_candidate_t: TargetInstitution | None = None
        best_score = 0.0
        best_method = ""

        for t in pref_targets:
            setter_norm = _norm(t.setter_name)
            t_name_norm = _norm(t.name)

            if corp_norm and setter_norm and (corp_norm in setter_norm or setter_norm in corp_norm):
                if norm_name and t_name_norm and (norm_name in t_name_norm or t_name_norm in norm_name):
                    score = min(len(norm_name), len(t_name_norm)) / max(len(norm_name), len(t_name_norm))
                    if score > best_score:
                        best_score = score
                        best_candidate_t = t
                        best_method = "setter_containment"

            if norm_name and t_name_norm:
                if norm_name in t_name_norm or t_name_norm in norm_name:
                    score = min(len(norm_name), len(t_name_norm)) / max(len(norm_name), len(t_name_norm))
                    if score > best_score:
                        best_score = score
                        best_candidate_t = t
                        best_method = "name_containment"

        if best_candidate_t is not None and best_score >= 0.6:
            candidate.candidate_code = best_candidate_t.school_code
            candidate.candidate_name = best_candidate_t.name
            candidate.match_method = best_method
            candidate.confidence = round(best_score, 3)
            # Fuzzy matches always go to manual review, never auto-applied
            report.needs_manual.append(candidate)
        else:
            # Check if school is excluded (閉校/統合/etc.)
            from eidp.db.models import SchoolYearStatus
            latest_status = (
                session.query(SchoolYearStatus)
                .filter(SchoolYearStatus.school_id == school.id)
                .order_by(SchoolYearStatus.fiscal_year.desc())
                .first()
            )
            if latest_status and latest_status.excluded_reason:
                candidate.resolution = "excluded"
                candidate.match_method = f"excluded:{latest_status.excluded_reason}"
                report.excluded.append(candidate)
            else:
                report.needs_manual.append(candidate)

    # Check for target-list schools missing from DB
    db_codes = used_codes | {c.candidate_code for c in report.auto_assigned if c.candidate_code}
    for t in targets:
        if t.school_code not in db_codes:
            report.missing_from_db.append(t)

    log.info(
        "reconcile_complete",
        already_resolved=report.already_resolved,
        auto_assigned=len(report.auto_assigned),
        needs_manual=len(report.needs_manual),
        excluded=len(report.excluded),
        missing_from_db=len(report.missing_from_db),
    )
    return report


def apply_reconciliation(session: Session, report: ReconcileReport) -> dict[str, int]:
    """Apply auto-assigned reconciliation results to DB."""
    stats = {"codes_assigned": 0, "aliases_created": 0}

    for c in report.auto_assigned:
        if not c.candidate_code:
            continue

        school = session.get(School, c.school_id)
        if school is None:
            continue

        # Check for code conflict
        existing = session.query(School).filter(School.school_code == c.candidate_code).first()
        if existing and existing.id != school.id:
            log.warning("reconcile_conflict", code=c.candidate_code, school=school.school_name)
            continue

        school.school_code = c.candidate_code
        stats["codes_assigned"] += 1

        if c.candidate_name and c.candidate_name != school.school_name:
            # Check alias doesn't already exist
            existing_alias = (
                session.query(SchoolAlias)
                .filter(SchoolAlias.school_id == school.id, SchoolAlias.alias_name == c.candidate_name)
                .first()
            )
            if not existing_alias:
                alias = SchoolAlias(
                    school_id=school.id,
                    alias_name=c.candidate_name,
                    alias_type="formal",
                    source="target_list",
                )
                session.add(alias)
                stats["aliases_created"] += 1

    session.flush()
    log.info("reconciliation_applied", **stats)
    return stats


def verify_identity(session: Session, data_dir: Path) -> dict[str, object]:
    """Verification gate: check identity completeness including target list coverage."""
    from sqlalchemy import text

    from eidp.db.models import SchoolYearStatus

    total = session.query(func.count(School.id)).scalar() or 0
    with_code = session.query(func.count(School.id)).filter(School.school_code.isnot(None)).scalar() or 0
    without_code = total - with_code

    # Check for duplicate codes
    dupes = session.execute(
        text("SELECT school_code, count(*) FROM school WHERE school_code IS NOT NULL GROUP BY school_code HAVING count(*) > 1")
    ).fetchall()

    # Count excluded schools (no code needed)
    excluded_ids = set()
    for row in session.query(SchoolYearStatus.school_id).filter(SchoolYearStatus.excluded_reason.isnot(None)).distinct():
        excluded_ids.add(row[0])

    # Schools without code that are NOT excluded = truly unresolved
    no_code_schools = session.query(School).filter(School.school_code.is_(None)).all()
    truly_unresolved = [s for s in no_code_schools if s.id not in excluded_ids]

    # Target list gap: MEXT target 専門学校 codes not in our DB
    target_list_path = data_dir / "target_institutions.xlsx"
    target_gap = 0
    if target_list_path.exists():
        targets = load_target_institutions(target_list_path)
        db_codes = set()
        for row in session.query(School.school_code).filter(School.school_code.isnot(None)):
            if row[0]:
                db_codes.add(row[0])
        target_codes = {t.school_code for t in targets}
        target_gap = len(target_codes - db_codes)

    result = {
        "total_schools": total,
        "with_code": with_code,
        "without_code": without_code,
        "excluded_no_code_needed": len([s for s in no_code_schools if s.id in excluded_ids]),
        "truly_unresolved": len(truly_unresolved),
        "duplicate_codes": len(dupes),
        "target_list_gap": target_gap,
        "pass": len(truly_unresolved) == 0 and len(dupes) == 0 and target_gap == 0,
    }
    return result
